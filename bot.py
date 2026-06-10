import os
import atexit
import json
import signal
import threading
import uuid
import hashlib
import re
import requests
from typing import Optional
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import io
import logging
import asyncio
import time
import functools
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import subprocess
    subprocess.run(["pip", "install", "openpyxl", "-q"], check=True)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, ReplyKeyboardRemove, KeyboardButton, CopyTextButton
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes
from datetime import datetime

# ── Performance: thread-safe DB helper with WAL mode ──────────────────────────
async def _run_db(func, *args):
    """Run a blocking DB function in a thread so the event loop stays free."""
    return await asyncio.to_thread(func, *args)

# ── Performance: membership-check TTL cache ────────────────────────────────────
_MEMBER_CACHE: dict = {}          # user_id → (unjoined_list, expire_at)
_MEMBER_CACHE_TTL = 300           # 5 minutes

def _cache_get(user_id: int):
    entry = _MEMBER_CACHE.get(user_id)
    if entry and time.monotonic() < entry[1]:
        return entry[0]
    return None

def _cache_set(user_id: int, unjoined: list):
    _MEMBER_CACHE[user_id] = (unjoined, time.monotonic() + _MEMBER_CACHE_TTL)

def _cache_invalidate(user_id: int = None):
    """Invalidate one user or the whole cache (call when channels change)."""
    if user_id is None:
        _MEMBER_CACHE.clear()
    else:
        _MEMBER_CACHE.pop(user_id, None)

# ── is_admin() cache (60 s TTL) ───────────────────────────────────────────────
_ADMIN_CACHE: dict = {}          # user_id → (bool, expire_at)
_ADMIN_CACHE_TTL = 60

def _admin_cache_get(user_id: int):
    entry = _ADMIN_CACHE.get(user_id)
    if entry and time.monotonic() < entry[1]:
        return entry[0]
    return None

def _admin_cache_set(user_id: int, result: bool):
    _ADMIN_CACHE[user_id] = (result, time.monotonic() + _ADMIN_CACHE_TTL)

def _admin_cache_clear(user_id: int = None):
    if user_id is None:
        _ADMIN_CACHE.clear()
    else:
        _ADMIN_CACHE.pop(user_id, None)

# ── get_referral_settings() cache (60 s TTL) ──────────────────────────────────
_RSETTINGS_CACHE: list = []      # [result, expire_at]

def _rsettings_cache_get():
    if _RSETTINGS_CACHE and time.monotonic() < _RSETTINGS_CACHE[1]:
        return _RSETTINGS_CACHE[0]
    return None

def _rsettings_cache_set(result):
    if _RSETTINGS_CACHE:
        _RSETTINGS_CACHE[0] = result
        _RSETTINGS_CACHE[1] = time.monotonic() + 60
    else:
        _RSETTINGS_CACHE.extend([result, time.monotonic() + 60])

def _rsettings_cache_clear():
    _RSETTINGS_CACHE.clear()

# ── get_user_balance_data() cache (30 s TTL) ──────────────────────────────────
_BAL_CACHE: dict = {}            # user_id → (result, expire_at)

def _bal_cache_get(user_id: int):
    entry = _BAL_CACHE.get(user_id)
    if entry and time.monotonic() < entry[1]:
        return entry[0]
    return None

def _bal_cache_set(user_id: int, result):
    _BAL_CACHE[user_id] = (result, time.monotonic() + 30)

def _bal_cache_clear(user_id: int = None):
    if user_id is None:
        _BAL_CACHE.clear()
    else:
        _BAL_CACHE.pop(user_id, None)

# ── get_countries() cache (120 s TTL) ─────────────────────────────────────────
_COUNTRIES_CACHE: list = []

def _countries_cache_get():
    if _COUNTRIES_CACHE and time.monotonic() < _COUNTRIES_CACHE[1]:
        return _COUNTRIES_CACHE[0]
    return None

def _countries_cache_set(result):
    if _COUNTRIES_CACHE:
        _COUNTRIES_CACHE[0] = result; _COUNTRIES_CACHE[1] = time.monotonic() + 120
    else:
        _COUNTRIES_CACHE.extend([result, time.monotonic() + 120])

def _countries_cache_clear():
    _COUNTRIES_CACHE.clear()

# ── get_numbers_count_by_country() cache (45 s TTL) ───────────────────────────
_NUMCOUNT_CACHE: dict = {}          # country_id → (total, available, expire_at)

def _numcount_cache_get(country_id):
    e = _NUMCOUNT_CACHE.get(country_id)
    if e and time.monotonic() < e[2]:
        return e[0], e[1]
    return None

def _numcount_cache_set(country_id, total, available):
    _NUMCOUNT_CACHE[country_id] = (total, available, time.monotonic() + 45)

def _numcount_cache_clear(country_id=None):
    if country_id is None:
        _NUMCOUNT_CACHE.clear()
    else:
        _NUMCOUNT_CACHE.pop(country_id, None)

# ── get_numbers_per_request() cache (300 s TTL) ───────────────────────────────
_NPR_CACHE: list = []

def _npr_cache_get():
    if _NPR_CACHE and time.monotonic() < _NPR_CACHE[1]:
        return _NPR_CACHE[0]
    return None

def _npr_cache_set(val):
    if _NPR_CACHE:
        _NPR_CACHE[0] = val; _NPR_CACHE[1] = time.monotonic() + 300
    else:
        _NPR_CACHE.extend([val, time.monotonic() + 300])

def _npr_cache_clear():
    _NPR_CACHE.clear()

# ── get_otp_link() cache (300 s TTL) ──────────────────────────────────────────
_OTP_CACHE: list = []

def _otp_cache_get():
    if _OTP_CACHE and time.monotonic() < _OTP_CACHE[1]:
        return _OTP_CACHE[0]
    return None

def _otp_cache_set(val):
    if _OTP_CACHE:
        _OTP_CACHE[0] = val; _OTP_CACHE[1] = time.monotonic() + 300
    else:
        _OTP_CACHE.extend([val, time.monotonic() + 300])

def _otp_cache_clear():
    _OTP_CACHE.clear()

# ── get_services() cache (120 s TTL) ──────────────────────────────────────────
_SERVICES_CACHE: list = []

def _services_cache_get():
    if _SERVICES_CACHE and time.monotonic() < _SERVICES_CACHE[1]:
        return _SERVICES_CACHE[0]
    return None

def _services_cache_set(val):
    if _SERVICES_CACHE:
        _SERVICES_CACHE[0] = val; _SERVICES_CACHE[1] = time.monotonic() + 120
    else:
        _SERVICES_CACHE.extend([val, time.monotonic() + 120])

def _services_cache_clear():
    _SERVICES_CACHE.clear()

# ── get_join_channels() cache (60 s TTL) ──────────────────────────────────────
_JOIN_CHANNELS_CACHE: list = []

def _join_channels_cache_get():
    if _JOIN_CHANNELS_CACHE and time.monotonic() < _JOIN_CHANNELS_CACHE[1]:
        return _JOIN_CHANNELS_CACHE[0]
    return None

def _join_channels_cache_set(val):
    if _JOIN_CHANNELS_CACHE:
        _JOIN_CHANNELS_CACHE[0] = val; _JOIN_CHANNELS_CACHE[1] = time.monotonic() + 60
    else:
        _JOIN_CHANNELS_CACHE.extend([val, time.monotonic() + 60])

def _join_channels_cache_clear():
    _JOIN_CHANNELS_CACHE.clear()

# ── service_countries cache (60 s TTL) ────────────────────────────────────────
_SVC_COUNTRIES_CACHE: dict = {}  # service_id → (list, expire_at)

def _svc_countries_cache_get(service_id):
    e = _SVC_COUNTRIES_CACHE.get(service_id)
    if e and time.monotonic() < e[1]:
        return e[0]
    return None

def _svc_countries_cache_set(service_id, val):
    _SVC_COUNTRIES_CACHE[service_id] = (val, time.monotonic() + 60)

def _svc_countries_cache_clear(service_id=None):
    if service_id is None:
        _SVC_COUNTRIES_CACHE.clear()
    else:
        _SVC_COUNTRIES_CACHE.pop(service_id, None)

# ── get_check_interval() cache (60 s TTL) ─────────────────────────────────────
_CI_CACHE: list = []
def _ci_cache_get():
    if _CI_CACHE and time.monotonic() < _CI_CACHE[1]: return _CI_CACHE[0]
    return None
def _ci_cache_set(v):
    if _CI_CACHE: _CI_CACHE[0]=v; _CI_CACHE[1]=time.monotonic()+60
    else: _CI_CACHE.extend([v, time.monotonic()+60])
def _ci_cache_clear(): _CI_CACHE.clear()

# ── get_custom_message() cache (60 s TTL) ─────────────────────────────────────
_CUSTMSG_CACHE: list = []
def _custmsg_cache_get():
    if _CUSTMSG_CACHE and time.monotonic() < _CUSTMSG_CACHE[1]: return (_CUSTMSG_CACHE[0],)
    return None
def _custmsg_cache_set(v):
    if _CUSTMSG_CACHE: _CUSTMSG_CACHE[0]=v; _CUSTMSG_CACHE[1]=time.monotonic()+60
    else: _CUSTMSG_CACHE.extend([v, time.monotonic()+60])
def _custmsg_cache_clear(): _CUSTMSG_CACHE.clear()

# ── get_user_count() cache (60 s TTL) ─────────────────────────────────────────
_UCOUNT_CACHE: list = []
def _ucount_cache_get():
    if _UCOUNT_CACHE and time.monotonic() < _UCOUNT_CACHE[1]: return _UCOUNT_CACHE[0]
    return None
def _ucount_cache_set(v):
    if _UCOUNT_CACHE: _UCOUNT_CACHE[0]=v; _UCOUNT_CACHE[1]=time.monotonic()+60
    else: _UCOUNT_CACHE.extend([v, time.monotonic()+60])
def _ucount_cache_clear(): _UCOUNT_CACHE.clear()

# ── get_total_referral_stats() cache (30 s TTL) ───────────────────────────────
_REF_STATS_CACHE: list = []
def _ref_stats_cache_get():
    if _REF_STATS_CACHE and time.monotonic() < _REF_STATS_CACHE[1]: return _REF_STATS_CACHE[0]
    return None
def _ref_stats_cache_set(v):
    if _REF_STATS_CACHE: _REF_STATS_CACHE[0]=v; _REF_STATS_CACHE[1]=time.monotonic()+30
    else: _REF_STATS_CACHE.extend([v, time.monotonic()+30])
def _ref_stats_cache_clear(): _REF_STATS_CACHE.clear()

# ── get_withdraw_config() cache (60 s TTL) ────────────────────────────────────
_WCONFIG_CACHE: list = []
def _wconfig_cache_get():
    if _WCONFIG_CACHE and time.monotonic() < _WCONFIG_CACHE[1]: return _WCONFIG_CACHE[0]
    return None
def _wconfig_cache_set(v):
    if _WCONFIG_CACHE: _WCONFIG_CACHE[0]=v; _WCONFIG_CACHE[1]=time.monotonic()+60
    else: _WCONFIG_CACHE.extend([v, time.monotonic()+60])
def _wconfig_cache_clear(): _WCONFIG_CACHE.clear()

# ── get_withdraw_stats() cache (30 s TTL) ─────────────────────────────────────
_WSTATS_CACHE: list = []
def _wstats_cache_get():
    if _WSTATS_CACHE and time.monotonic() < _WSTATS_CACHE[1]: return _WSTATS_CACHE[0]
    return None
def _wstats_cache_set(v):
    if _WSTATS_CACHE: _WSTATS_CACHE[0]=v; _WSTATS_CACHE[1]=time.monotonic()+30
    else: _WSTATS_CACHE.extend([v, time.monotonic()+30])
def _wstats_cache_clear(): _WSTATS_CACHE.clear()

# ── Session-level verified-user guard ──────────────────────────────────────────
_ALREADY_VERIFIED: set = set()

# Bot configuration
BOT_TOKEN = '7965821240:AAGCoLHvEwKzEVaf7IvQZ2UR-TdxNaSWSRc'

PROTECTED_ADMIN_IDS = [6013092363, 6991526772]  # These admin user IDs cannot be removed
CHANNEL_USERNAME = "https://t.me/sxchannel143"  # Your channel username
OTP_GROUP_LINK = "https://t.me/sxotp99"  # Updated OTP group link

# New channel links
CHANNEL_ONE_LINK = "https://t.me/+fU203rxxA9xkNjk1"
CHANNEL_TWO_LINK = "https://t.me/+fU203rxxA9xkNjk1"

import re as _re

# Country name → Unicode flag (used in inline buttons where HTML can't render)
COUNTRY_UNICODE_FLAGS = {
    'afghanistan':'🇦🇫','albania':'🇦🇱','algeria':'🇩🇿','andorra':'🇦🇩',
    'angola':'🇦🇴','argentina':'🇦🇷','armenia':'🇦🇲','australia':'🇦🇺',
    'austria':'🇦🇹','azerbaijan':'🇦🇿','bahamas':'🇧🇸','bahrain':'🇧🇭',
    'bangladesh':'🇧🇩','barbados':'🇧🇧','belarus':'🇧🇾','belgium':'🇧🇪',
    'belize':'🇧🇿','benin':'🇧🇯','bhutan':'🇧🇹','bolivia':'🇧🇴',
    'bosnia':'🇧🇦','bosnia and herzegovina':'🇧🇦','botswana':'🇧🇼',
    'brazil':'🇧🇷','brunei':'🇧🇳','bulgaria':'🇧🇬','burkina faso':'🇧🇫',
    'burundi':'🇧🇮','cambodia':'🇰🇭','cameroon':'🇨🇲','canada':'🇨🇦',
    'cape verde':'🇨🇻','cabo verde':'🇨🇻','central african republic':'🇨🇫',
    'chad':'🇹🇩','chile':'🇨🇱','china':'🇨🇳','colombia':'🇨🇴',
    'comoros':'🇰🇲','congo':'🇨🇬','croatia':'🇭🇷','cuba':'🇨🇺',
    'cyprus':'🇨🇾','czech republic':'🇨🇿','czechia':'🇨🇿','denmark':'🇩🇰',
    'djibouti':'🇩🇯','dominican republic':'🇩🇴','drc':'🇨🇩',
    'democratic republic of congo':'🇨🇩','ecuador':'🇪🇨','egypt':'🇪🇬',
    'el salvador':'🇸🇻','england':'🏴󠁧󠁢󠁥󠁮󠁧󠁿','eritrea':'🇪🇷','estonia':'🇪🇪',
    'ethiopia':'🇪🇹','fiji':'🇫🇯','finland':'🇫🇮','france':'🇫🇷',
    'gabon':'🇬🇦','gambia':'🇬🇲','georgia':'🇬🇪','germany':'🇩🇪',
    'ghana':'🇬🇭','greece':'🇬🇷','guatemala':'🇬🇹','guinea':'🇬🇳',
    'guinea bissau':'🇬🇼','guyana':'🇬🇾','haiti':'🇭🇹','honduras':'🇭🇳',
    'hong kong':'🇭🇰','hungary':'🇭🇺','iceland':'🇮🇸','india':'🇮🇳',
    'indonesia':'🇮🇩','iran':'🇮🇷','iraq':'🇮🇶','ireland':'🇮🇪',
    'israel':'🇮🇱','italy':'🇮🇹','ivory coast':'🇨🇮','jamaica':'🇯🇲',
    'japan':'🇯🇵','jordan':'🇯🇴','kazakhstan':'🇰🇿','kenya':'🇰🇪',
    'kosovo':'🇽🇰','kuwait':'🇰🇼','kyrgyzstan':'🇰🇬','laos':'🇱🇦',
    'latvia':'🇱🇻','lebanon':'🇱🇧','lesotho':'🇱🇸','liberia':'🇱🇷',
    'libya':'🇱🇾','liechtenstein':'🇱🇮','lithuania':'🇱🇹','luxembourg':'🇱🇺',
    'macau':'🇲🇴','madagascar':'🇲🇬','malawi':'🇲🇼','malaysia':'🇲🇾',
    'maldives':'🇲🇻','mali':'🇲🇱','malta':'🇲🇹','mauritania':'🇲🇷',
    'mauritius':'🇲🇺','mexico':'🇲🇽','moldova':'🇲🇩','monaco':'🇲🇨',
    'mongolia':'🇲🇳','montenegro':'🇲🇪','morocco':'🇲🇦','mozambique':'🇲🇿',
    'myanmar':'🇲🇲','namibia':'🇳🇦','nepal':'🇳🇵','netherlands':'🇳🇱',
    'new zealand':'🇳🇿','nicaragua':'🇳🇮','niger':'🇳🇪','nigeria':'🇳🇬',
    'north korea':'🇰🇵','north macedonia':'🇲🇰','norway':'🇳🇴','oman':'🇴🇲',
    'pakistan':'🇵🇰','palestine':'🇵🇸','panama':'🇵🇦','papua new guinea':'🇵🇬',
    'paraguay':'🇵🇾','peru':'🇵🇪','philippines':'🇵🇭','poland':'🇵🇱',
    'portugal':'🇵🇹','qatar':'🇶🇦','romania':'🇷🇴','russia':'🇷🇺',
    'rwanda':'🇷🇼','saudi arabia':'🇸🇦','senegal':'🇸🇳','serbia':'🇷🇸',
    'seychelles':'🇸🇨','sierra leone':'🇸🇱','singapore':'🇸🇬','slovakia':'🇸🇰',
    'somalia':'🇸🇴','south africa':'🇿🇦','south korea':'🇰🇷','south sudan':'🇸🇸',
    'spain':'🇪🇸','sri lanka':'🇱🇰','sudan':'🇸🇩','suriname':'🇸🇷',
    'sweden':'🇸🇪','switzerland':'🇨🇭','syria':'🇸🇾','taiwan':'🇹🇼',
    'tajikistan':'🇹🇯','tanzania':'🇹🇿','thailand':'🇹🇭','timor leste':'🇹🇱',
    'east timor':'🇹🇱','togo':'🇹🇬','trinidad and tobago':'🇹🇹',
    'tunisia':'🇹🇳','turkey':'🇹🇷','turkiye':'🇹🇷','turkmenistan':'🇹🇲',
    'uganda':'🇺🇬','ukraine':'🇺🇦','united arab emirates':'🇦🇪','uae':'🇦🇪',
    'united kingdom':'🇬🇧','uk':'🇬🇧','united states':'🇺🇸',
    'usa':'🇺🇸','america':'🇺🇸','uruguay':'🇺🇾','uzbekistan':'🇺🇿',
    'venezuela':'🇻🇪','vietnam':'🇻🇳','yemen':'🇾🇪','zambia':'🇿🇲',
    'zimbabwe':'🇿🇼','san marino':'🇸🇲','vatican':'🇻🇦','costa rica':'🇨🇷',
}

# Unicode flag → FlagSXSponsor animated custom emoji ID
_FLAG_ANIMATED_IDS = {
    '🇦🇫':'6068927950483366509','🇦🇱':'6068850147150799507','🇩🇿':'6068720013936697475',
    '🇦🇩':'6068613361308802907','🇦🇴':'6068732757104666619','🇦🇬':'6068835673111011150',
    '🇦🇷':'6068907480669233831','🇦🇲':'6068639337271008823','🇦🇺':'6068989626213738674',
    '🇦🇹':'6068984738540953165','🇧🇸':'6068939456700751075','🇧🇭':'6071034288344670682',
    '🇧🇩':'6071169073008351765','🇧🇧':'6068610165853135804','🇧🇾':'6069026790565748639',
    '🇧🇪':'6069107235303203770','🇧🇿':'6068806149505819224','🇧🇯':'6068924673423318975',
    '🇧🇴':'6068668040537449187','🇧🇦':'6068755833963945453','🇧🇼':'6068645315865485515',
    '🇧🇷':'6068810435883181359','🇧🇳':'6068791108530348141','🇧🇬':'6068782574430331375',
    '🇧🇫':'6069016778996981094','🇧🇮':'6068790356911070862','🇰🇭':'6068712076837138187',
    '🇨🇲':'6069048351301574234','🇨🇦':'6068905861466562243','🇨🇻':'6068627912658001434',
    '🇨🇫':'6068740144448413981','🇹🇩':'6068761357291889363','🇨🇱':'6068831073201037205',
    '🇨🇳':'6068915941754806919','🇨🇴':'6068606459296359433','🇰🇲':'6071338857360533636',
    '🇨🇬':'6069017358817566023','🇨🇩':'6068683712873111327','🇭🇷':'6068775075417432634',
    '🇨🇺':'6068879752360369822','🇨🇾':'6068826103923877221','🇨🇿':'6068712549283536903',
    '🇩🇰':'6068794578863923495','🇩🇯':'6068900260829207669','🇩🇴':'6068847986782249403',
    '🇩🇲':'6069120133089992671','🇪🇨':'6071399382039666958','🇪🇬':'6068671489396187942',
    '🇸🇻':'6069151069739425958','🏴󠁧󠁢󠁥󠁮󠁧󠁿':'6069059629885693764','🇪🇷':'6068850705496547535',
    '🇪🇪':'6068739796556064126','🇪🇹':'6069060050792488610','🇪🇺':'6068829904969933582',
    '🇫🇯':'6068861073547599331','🇫🇮':'6068854880204758456','🇫🇷':'6068852440663337195',
    '🇬🇦':'6068612205962600590','🇬🇲':'6068670965410176631','🇬🇪':'6069006874802397901',
    '🇩🇪':'6069114515272767964','🇬🇭':'6068713262248110916','🇬🇷':'6068612064228681415',
    '🇬🇹':'6068690855403725013','🇬🇼':'6068741488773179908','🇬🇾':'6068943498264977216',
    '🇬🇳':'6069094994646409126','🇭🇳':'6068650895028004419','🇭🇺':'6068987302636430069',
    '🇮🇸':'6071358566965452179','🇮🇩':'6068817020068044593','🇮🇷':'6070979016410537269',
    '🇮🇶':'6068769479075047597','🇮🇪':'6068826662269624401','🇮🇱':'6069123083732525257',
    '🇮🇹':'6068923015565943055','🇯🇲':'6068662388360486601','🇯🇵':'6068626512498663220',
    '🇯🇴':'6068651659532182346','🇰🇿':'6068607004757204932','🇰🇪':'6068699745986028043',
    '🇽🇰':'6068829342329216753','🇰🇼':'6068951791846825925','🇰🇬':'6068969495702019112',
    '🇱🇦':'6069114317704274654','🇱🇻':'6068872996376813421','🇱🇧':'6068677012724130653',
    '🇱🇸':'6068958221412868398','🇱🇷':'6068645015217773293','🇱🇾':'6070868000095870508',
    '🇱🇮':'6069160930984338118','🇱🇹':'6068955498403604216','🇱🇺':'6068924690603188765',
    '🇲🇰':'6068917466468198032','🇲🇬':'6069085897905675283','🇲🇼':'6068712240045890811',
    '🇲🇾':'6068811595524350491','🇲🇻':'6068914670444487419','🇲🇱':'6069064341464817393',
    '🇲🇹':'6068625486001480089','🇲🇷':'6069092331766686459','🇲🇺':'6068888170496269303',
    '🇲🇽':'6068859505884538177','🇲🇩':'6071339647634512652','🇲🇨':'6071181867715927382',
    '🇲🇳':'6068912029039599214','🇲🇦':'6068827401003999277','🇲🇲':'6068752853256643546',
    '🇳🇦':'6069061360757514214','🇳🇵':'6068668276760648922','🇳🇱':'6069162593136682062',
    '🇳🇿':'6068612227437437727','🇳🇪':'6068898491302683941','🇳🇬':'6068902352478280873',
    '🇰🇵':'6069030737640694177','🇳🇴':'6069000642804850808','🇴🇲':'6070952323188793206',
    '🇵🇰':'6068673946117481090','🇵🇸':'6069126158929111291','🇵🇦':'6069105508726349994',
    '🇵🇬':'6068864333427777276','🇵🇾':'6068710968735571795','🇵🇪':'6068663191519373386',
    '🇵🇭':'6068738503770906845','🇵🇱':'6068710247181064806','🇵🇹':'6068758136066415801',
    '🇶🇦':'6071180373067308851','🇷🇴':'6068749417282805793','🇷🇺':'6068656491370389370',
    '🇷🇼':'6068851903792423534','🇸🇦':'6069080507721719535','🇸🇳':'6068901772657695888',
    '🇷🇸':'6068635901297174101','🇸🇨':'6069112165925657872','🇸🇱':'6069155695419203519',
    '🇸🇬':'6068982887410048870','🇸🇰':'6068926528849191367','🇸🇧':'6068732980442963481',
    '🇸🇴':'6071048341477662902','🇿🇦':'6069064246975536894','🇰🇷':'6068937631339651142',
    '🇸🇸':'6070931857669627058','🇪🇸':'6068658166407635084','🇱🇰':'6068998379357086095',
    '🇸🇷':'6070865642158824426','🇸🇩':'6071210051291323763','🇸🇪':'6068856507997363828',
    '🇨🇭':'6068705797594946946','🇸🇾':'6068732434982116796','🇹🇼':'6068967154944843903',
    '🇹🇯':'6069117680663666200','🇹🇿':'6068897232877264377','🇹🇭':'6069161334711263904',
    '🇹🇱':'6069017393177305558','🇹🇬':'6068823157576310784','🇹🇹':'6068876260551959598',
    '🇹🇳':'6071171396585660619','🇹🇷':'6069086065409400728','🇹🇲':'6069152036107066544',
    '🇺🇬':'6068958315902149071','🇺🇦':'6068774242193778244','🇦🇪':'6069102162946829028',
    '🇺🇸':'6068663066965319449','🇺🇾':'6068742175967945374','🇺🇿':'6068989278321384647',
    '🇻🇦':'6068687080127471441','🇻🇪':'6068905234401337504','🇾🇪':'6068934100876537105',
    '🇿🇼':'6070975928329050777','🇬🇧':'6068980155810848896','🇨🇮':'6068671824403636419',
    '🇸🇲':'6068691761641823924','🇸🇮':'6068865651982736345','🌐':'6068742614054607952',
}

# SXEmojisSXSponsor animated custom emoji IDs
_SX = {
    '🤖':'6069064298515144683','🚩':'6068725142127648372','📣':'6068984648346641323',
    '👑':'6068633264187253513','🔑':'6068876810307772648','💰':'6068607339764653118',
    '✅':'6069044743529046223','🌟':'6068709087539896145','📱':'6069136299346895222',
    '📊':'6070915764427168620','🔔':'6068616324836238233','🎁':'6069098151447371295',
    '💸':'6069031789907681337','🔗':'6068709083244927986','❌':'6068625520361217686',
    '✔️':'6069081100427206871','➕':'6068772120479932812','➖':'6068768181994921479',
    '🟢':'6068931596910599428','🔴':'6069038726279864720','👨‍💻':'6069003460303395931',
    '💎':'6068735273955498815','⭐':'6069081697427659834','🔍':'6068705166234755288',
    '📌':'6068651693891918945','⚠️':'6068960751148604995','🏦':'6068894247874993666',
    '💵':'6068951821911596988','🌐':'6068742614054607952','🔄':'6068743984149175274',
    '⌛':'6068649679552257644','👍':'6068653716821516079','🎉':'6068944949963923893',
    '🗑':'6068801498056244942','📥':'6071131436209937959','👌':'6068923758595284398',
    '💲':'6068678932574510264','👥':'6069003460303395931','🗺':'6068725142127648372',
    '💬':'6068727156467309816','📞':'6068652724684070452','🔴_2':'6069042858038403362',
    '✏️':'6069153612360066093','📢':'6068984648346641323',
}

def _sx(emoji: str) -> str:
    """Return an animated <tg-emoji> HTML tag from SXEmojisSXSponsor pack."""
    eid = _SX.get(emoji)
    if not eid:
        return emoji
    return f'<tg-emoji emoji-id="{eid}">{emoji}</tg-emoji>'

def _clean_name(name: str) -> str:
    """Strip any leading emoji/whitespace to get the plain country name."""
    return _re.sub(r'^[\U00010000-\U0010ffff\U00002600-\U000027ff\s]+', '', name).strip()

def get_unicode_flag(country_name: str) -> str:
    """Return Unicode flag emoji for a country name (safe for buttons)."""
    key = _clean_name(country_name).lower()
    if key in COUNTRY_UNICODE_FLAGS:
        return COUNTRY_UNICODE_FLAGS[key]
    for k, v in COUNTRY_UNICODE_FLAGS.items():
        if k in key or key in k:
            return v
    return '🌍'

def get_animated_flag_html(country_name: str) -> str:
    """Return animated tg-emoji HTML tag from FlagSXSponsor pack."""
    unicode_flag = get_unicode_flag(country_name)
    emoji_id = _FLAG_ANIMATED_IDS.get(unicode_flag)
    if emoji_id:
        return f'<tg-emoji emoji-id="{emoji_id}">{unicode_flag}</tg-emoji>'
    return unicode_flag  # fallback to Unicode if not in pack

# ──────────────────────────────────────────────
#  SERVICE ANIMATED EMOJI  (APPEmojiSXSponsor)
# ──────────────────────────────────────────────

# Service name keyword → Unicode emoji (fallback display)
SERVICE_UNICODE_EMOJIS = {
    'whatsapp':    '💬',
    'telegram':    '✈️',
    'instagram':   '📸',
    'facebook':    '📘',
    'messenger':   '💬',
    'gmail':       '📧',
    'google':      '🔍',
    'youtube':     '▶️',
    'tiktok':      '🎵',
    'discord':     '🎮',
    'twitter':     '🐦',
    'x':           '🐦',
    'snapchat':    '👻',
    'netflix':     '🎬',
    'spotify':     '🎧',
    'amazon':      '📦',
    'microsoft':   '🪟',
    'apple':       '🍎',
    'icloud':      '☁️',
    'paypal':      '💳',
    'linkedin':    '💼',
    'reddit':      '🤖',
    'viber':       '📞',
    'wechat':      '💚',
    'signal':      '🔒',
    'line':        '💚',
    'skype':       '💙',
    'zoom':        '📹',
    'pinterest':   '📌',
    'tumblr':      '📝',
    'clubhouse':   '🎙️',
    'binance':     '💰',
    'coinbase':    '💰',
    'okx':         '📊',
    'bybit':       '📊',
    'kraken':      '🦑',
    'kucoin':      '💰',
    'huobi':       '💰',
    'gate':        '🚪',
    'mexc':        '📊',
    'bitget':      '📊',
    'uber':        '🚗',
    'lyft':        '🚗',
    'airbnb':      '🏠',
    'booking':     '🏨',
    'tinder':      '❤️',
    'bumble':      '🐝',
    'badoo':       '💛',
    'hinge':       '🔗',
    'yahoo':       '📬',
    'outlook':     '📨',
    'proton':      '🛡️',
    'bluesky':     '🦋',
    'twitter':     '🐦',
    'threads':     '🧵',
    'imo':         '📱',
    'bigo':        '🎥',
    'likee':       '🎬',
    'sharechat':   '💬',
    'truecaller':  '📞',
    'textplus':    '💬',
    'textfree':    '📱',
    'twilio':      '📲',
    '2degrees':    '📡',
    'numbrix':     '🔢',
    'hushed':      '🤫',
    'burner':      '🔥',
    'googlevoice': '📞',
    'vonage':      '📲',
    'dingtalk':    '🔔',
    'kakao':       '🟡',
    'zalo':        '🔵',
    'naver':       '🟢',
    'steam':       '🎮',
    'epic':        '🎮',
    'riot':        '🎮',
    'blizzard':    '🎮',
    'microsoft':   '🪟',
    'office':      '📄',
    'dropbox':     '📦',
    'box':         '📦',
    'onedrive':    '☁️',
    'gdrive':      '📁',
    'notion':      '📓',
    'slack':       '💬',
    'trello':      '📋',
    'asana':       '📌',
}

# App-specific custom animated sticker IDs from APPEmojiSXSponsor pack
# Mapped by dominant color analysis of 158 sticker thumbnails
_APP_EMOJI_CUSTOM_IDS: dict = {
    # ── Messaging ──────────────────────────────────────────────────────────
    'whatsapp':          '6300761828330840482',  # bright green  #109
    'whatsapp business': '6298480844214379008',  # green         #106
    'telegram':          '6068867859595927448',  # sky blue      #65
    'signal':            '6071133841391623449',  # green         #6
    'viber':             '6068811466675331885',  # purple-blue   #60
    'line':              '6068991180991896864',  # light green   #22
    'wechat':            '6068999968494985887',  # teal-green    #64
    'imo':               '6071133841391623449',  # green         #6
    'skype':             '6298289026679971933',  # light blue    #112
    'discord':           '6069134280712267236',  # blurple       #79
    'slack':             '6068807674219207819',  # dark          #8
    'teams':             '6069122684300565166',  # blue          #17
    'messenger':         '6068903074032786803',  # blue          #24
    'zalo':              '6068813158892445814',  # blue          #57
    'kakao':             '6069107140813921396',  # yellow        #12
    'kakaotalk':         '6069107140813921396',  # yellow        #12
    'snapchat':          '6069134126093443500',  # bright yellow #117
    # ── Social Media ──────────────────────────────────────────────────────
    'instagram':         '6068863345585299844',  # pink          #108
    'facebook':          '6068813158892445814',  # blue          #57
    'twitter':           '6298786920763761414',  # light blue    #85
    'x':                 '6298786920763761414',  # light blue    #85
    'tiktok':            '6068819665767898174',  # dark          #42
    'pinterest':         '6068629317112307916',  # dark red-pink #48
    'reddit':            '6068691972095221372',  # orange        #86
    'linkedin':          '6069122684300565166',  # blue          #17
    'tumblr':            '6068724914494381835',  # dark navy     #50
    'youtube':           '6068663500757016577',  # red           #121
    'twitch':            '6298717548451996369',  # purple        #114
    # ── Streaming & Entertainment ──────────────────────────────────────────
    'netflix':           '6068612699883839769',  # dark          #41
    'spotify':           '6298480844214379008',  # green         #106
    'amazon prime':      '6068904302393433242',  # orange        #128
    'prime video':       '6068904302393433242',  # orange        #128
    'apple tv':          '6300762360906784639',  # gray          #148
    'apple music':       '6069023191383155305',  # red-pink      #18
    'deezer':            '6068762130386000620',  # cyan-green    #70
    'soundcloud':        '6068691972095221372',  # orange        #86
    # ── Google ────────────────────────────────────────────────────────────
    'google':            '6069097687590903402',  # blue          #89
    'gmail':             '6068824403116826236',  # red           #103
    'google drive':      '6069136295051928666',  # green         #63
    'gdrive':            '6069136295051928666',  # green         #63
    'google meet':       '6071287910458466057',  # light blue    #20
    'google pay':        '6068798431449587943',  # blue          #71
    'gpay':              '6068798431449587943',  # blue          #71
    'youtube music':     '6068663500757016577',  # red           #121
    # ── Shopping ──────────────────────────────────────────────────────────
    'amazon':            '6068904302393433242',  # orange        #128
    'shopee':            '6068922362730913720',  # orange        #80
    'aliexpress':        '6068691972095221372',  # orange-red    #86
    'flipkart':          '6068997979925126231',  # blue          #4
    'ebay':              '6068824403116826236',  # red           #103
    'lazada':            '6069023191383155305',  # pink-red      #18
    'temu':              '6068824403116826236',  # red           #103
    # ── Finance & Crypto ──────────────────────────────────────────────────
    'paypal':            '6068997979925126231',  # blue          #4
    'binance':           '6071090251768537478',  # golden yellow #67
    'coinbase':          '6068903074032786803',  # blue          #24
    'okx':               '6068965101950477168',  # dark          #45
    'bybit':             '6068807674219207819',  # dark          #8
    'kucoin':            '6298480844214379008',  # green         #106
    'kraken':            '6068866562515804338',  # dark blue     #37
    'bkash':             '6068810882559778610',  # pink-red      #62
    'nagad':             '6068824403116826236',  # red           #103
    'rocket':            '6068922362730913720',  # orange        #80
    'crypto':            '6068811466675331885',  # blue          #60
    # ── Dating ────────────────────────────────────────────────────────────
    'tinder':            '6068810882559778610',  # red-pink      #62
    'bumble':            '6069107140813921396',  # yellow        #12
    'badoo':             '6069023191383155305',  # red           #18
    'hinge':             '6068810882559778610',  # pink          #62
    # ── Ride & Delivery ───────────────────────────────────────────────────
    'uber':              '6068814550461849153',  # dark          #13
    'lyft':              '6068810882559778610',  # pink          #62
    'pathao':            '6068824403116826236',  # red           #103
    # ── Email ─────────────────────────────────────────────────────────────
    'outlook':           '6068997979925126231',  # blue          #4
    'yahoo mail':        '6068717883632917743',  # purple        #102
    'yahoo':             '6068717883632917743',  # purple        #102
    'protonmail':        '6068810882559778610',  # red           #62
    # ── Cloud & Productivity ──────────────────────────────────────────────
    'dropbox':           '6068903074032786803',  # blue          #24
    'onedrive':          '6068997979925126231',  # blue          #4
    'notion':            '6068814550461849153',  # dark          #13
    'zoom':              '6069027920142147440',  # dark blue     #34
    'microsoft':         '6068982165855543196',  # orange-red    #1
    'office':            '6068982165855543196',  # orange-red    #1
    'trello':            '6068997979925126231',  # blue          #4
    # ── Tech Giants ───────────────────────────────────────────────────────
    'apple':             '6300762360906784639',  # gray          #148
    'icloud':            '6298721637260863189',  # light blue    #7
    'truecaller':        '6298289026679971933',  # teal          #112
    'shazam':            '6068997979925126231',  # blue          #4
    'airbnb':            '6068810882559778610',  # pink-red      #62
    'booking':           '6068997979925126231',  # blue          #4
    'steam':             '6069027920142147440',  # dark          #34
}

def get_service_emoji(service_name: str) -> str:
    """Return Unicode emoji for a service name."""
    key = _re.sub(r'[^\w\s]', '', service_name).lower().strip()
    if key in SERVICE_UNICODE_EMOJIS:
        return SERVICE_UNICODE_EMOJIS[key]
    for k, v in SERVICE_UNICODE_EMOJIS.items():
        if k in key or key in k:
            return v
    return '📱'

def get_service_emoji_html(service_name: str) -> str:
    """Return animated tg-emoji HTML — hardcoded pack first, then unicode."""
    key = _re.sub(r'[^\w\s]', '', service_name).lower().strip()
    # 1) Try hardcoded animated ID pack
    custom_id = None
    if key in _APP_EMOJI_CUSTOM_IDS:
        custom_id = _APP_EMOJI_CUSTOM_IDS[key]
    else:
        for k, v in _APP_EMOJI_CUSTOM_IDS.items():
            if k in key or key in k:
                custom_id = v
                break
    if custom_id:
        fallback = get_service_emoji(service_name)
        return f'<tg-emoji emoji-id="{custom_id}">{fallback}</tg-emoji>'
    # 3) Fallback: unicode emoji
    return get_service_emoji(service_name)


# ════════════════════════════════════════════════════════════════════════
# (Old Firestore layer removed — local JSON DB used instead)
# ════════════════════════════════════════════════════════════════════════

# DATABASE LAYER  (local JSON — formerly bot_db.py — inlined for single-file deploy)
# ════════════════════════════════════════════════════════════════════════
_DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "users.json")
_lock = threading.RLock()
_dirty = threading.Event()          # set whenever _DB is modified
_shutdown = threading.Event()       # signals the flush thread to stop

_PROTECTED_IDS: list = []

_DB: dict = {}

_DEFAULT_DB = {
    "settings": {
        "numbers_per_request": "3",
        "otp_link": "",
        "check_interval_minutes": "60",
        "referral_enabled": "1",
        "referral_reward": "1",
        "referral_label": "BDT",
        "withdraw_enabled": "1",
        "withdraw_min_amount": "100",
        "withdraw_group_chat_id": "",
        "custom_messages": {},
    },
    "join_channels": [],
    "countries": [],
    "numbers": {},
    "services": [],
    "service_countries": [],
    "admins": {},
    "users": {},
    "referrals": {},
    "withdraw_requests": {},
    "custom_emoji": {},
    "banned_users": {},
    "sms_panels": {},
    "number_assignments": {},
    "seen_sms_messages": {},
    "last_panel_messages": {},
    "global_forward_chat_id": "",
    "panel_send_counts": {},
    "country_otp_rewards": {},
    "group_number_btn_link": "https://t.me/UnofficialNumberBOT",
    "group_channel_btn_link": "https://t.me/sxchannel143",
}

# ─── Persistence ───────────────────────────────────────────────────────────────

def _load():
    global _DB
    if os.path.exists(_DB_PATH):
        try:
            with open(_DB_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            for k, v in _DEFAULT_DB.items():
                if k not in data:
                    data[k] = v
            data["settings"].setdefault("custom_messages", {})
            _DB = data
            return
        except Exception as e:
            print(f"⚠️ users.json load error: {e} — starting fresh")
    _DB = json.loads(json.dumps(_DEFAULT_DB))


def _flush():
    """Write _DB to disk atomically (tmp file → rename)."""
    tmp = _DB_PATH + ".tmp"
    try:
        with _lock:
            snapshot = json.dumps(_DB, ensure_ascii=False, indent=2)
        with open(tmp, "w", encoding="utf-8") as f:
            f.write(snapshot)
        os.replace(tmp, _DB_PATH)
        _dirty.clear()
    except Exception as e:
        print(f"⚠️ users.json flush error: {e}")


def _mark_dirty():
    """Mark that _DB has unsaved changes."""
    _dirty.set()


def _background_flush_loop(interval: int = 30):
    """Flush to disk every `interval` seconds when dirty."""
    while not _shutdown.wait(timeout=interval):
        if _dirty.is_set():
            _flush()
    if _dirty.is_set():       # final flush on shutdown
        _flush()


def _on_shutdown():
    _shutdown.set()


atexit.register(_on_shutdown)
try:
    _orig_sigterm = signal.getsignal(signal.SIGTERM)
    def _sigterm_handler(sig, frame):
        _on_shutdown()
        if callable(_orig_sigterm):
            _orig_sigterm(sig, frame)
    signal.signal(signal.SIGTERM, _sigterm_handler)
except Exception:
    pass


def db_init(otp_link="", channel_one_link="", channel_two_link="", protected_admin_ids=None):
    global _PROTECTED_IDS
    if protected_admin_ids:
        _PROTECTED_IDS = list(protected_admin_ids)

    _load()

    with _lock:
        s = _DB["settings"]
        if not s.get("otp_link"):
            s["otp_link"] = otp_link
        links = {ch["link"] for ch in _DB["join_channels"]}
        next_id = max((ch["id"] for ch in _DB["join_channels"]), default=0) + 1
        if channel_one_link and channel_one_link not in links:
            _DB["join_channels"].append({"id": next_id, "title": "Channel One", "link": channel_one_link})
            next_id += 1
        if channel_two_link and channel_two_link != channel_one_link and channel_two_link not in links:
            _DB["join_channels"].append({"id": next_id, "title": "Channel Two", "link": channel_two_link})
        _mark_dirty()

    _flush()   # immediate first flush

    t = threading.Thread(target=_background_flush_loop, args=(30,), daemon=True, name="json-flush")
    t.start()

    print("✅ users.json database initialized! (background flush every 30s)")


# ─── Settings helpers ──────────────────────────────────────────────────────────

def _get_setting(key, default=""):
    return _DB["settings"].get(key, default)


def _set_setting(key, value):
    with _lock:
        _DB["settings"][key] = str(value)
        _mark_dirty()


# ─── Bot Settings ──────────────────────────────────────────────────────────────

def get_numbers_per_request():
    try:
        return int(_get_setting("numbers_per_request", "1"))
    except Exception:
        return 1


def set_numbers_per_request(n):
    _set_setting("numbers_per_request", str(n))
    _npr_cache_clear()


def get_otp_link():
    val = _get_setting("otp_link", "")
    _otp_cache_set(val)
    return val


def set_otp_link(link):
    _set_setting("otp_link", link)
    _otp_cache_clear()


def get_group_number_btn_link() -> str:
    with _lock:
        return _DB.get("group_number_btn_link", "https://t.me/UnofficialNumberBOT")

def set_group_number_btn_link(link: str):
    with _lock:
        _DB["group_number_btn_link"] = link
        _mark_dirty()

def get_group_channel_btn_link() -> str:
    with _lock:
        return _DB.get("group_channel_btn_link", "https://t.me/sxchannel143")

def set_group_channel_btn_link(link: str):
    with _lock:
        _DB["group_channel_btn_link"] = link
        _mark_dirty()


def get_check_interval():
    try:
        return int(_get_setting("check_interval_minutes", "60"))
    except Exception:
        return 60


def set_check_interval(minutes: int):
    _set_setting("check_interval_minutes", str(minutes))
    _ci_cache_clear()


# ─── Custom Messages ───────────────────────────────────────────────────────────

def get_custom_messages() -> list:
    raw = _DB["settings"].get("custom_messages", {}) or {}
    msgs = sorted(raw.items(), key=lambda x: x[0])
    _custmsg_cache_set(msgs)
    return msgs


def get_custom_message() -> str | None:
    msgs = get_custom_messages()
    return "\n\n".join(t for _, t in msgs) if msgs else None


def add_custom_message(msg_text: str):
    msg_id = str(int(time.time() * 1000))
    with _lock:
        _DB["settings"].setdefault("custom_messages", {})[msg_id] = msg_text
        _mark_dirty()
    _custmsg_cache_clear()


def set_custom_message(msg_text: str):
    add_custom_message(msg_text)


def remove_custom_message_by_id(msg_id: str):
    with _lock:
        _DB["settings"].get("custom_messages", {}).pop(msg_id, None)
        _mark_dirty()
    _custmsg_cache_clear()


def remove_custom_message():
    with _lock:
        _DB["settings"]["custom_messages"] = {}
        _mark_dirty()
    _custmsg_cache_clear()


# ─── Join Channels ─────────────────────────────────────────────────────────────

def get_join_channels():
    result = [(ch["id"], ch["title"], ch["link"]) for ch in _DB["join_channels"]]
    _join_channels_cache_set(result)
    return result


def add_join_channel(title, link):
    with _lock:
        if any(ch["link"] == link for ch in _DB["join_channels"]):
            return False
        next_id = max((ch["id"] for ch in _DB["join_channels"]), default=0) + 1
        _DB["join_channels"].append({"id": next_id, "title": title, "link": link})
        for u in _DB["users"].values():
            u["is_verified"] = False
        _mark_dirty()
    _cache_invalidate()
    _join_channels_cache_clear()
    return True


def remove_join_channel(channel_id):
    with _lock:
        _DB["join_channels"] = [ch for ch in _DB["join_channels"] if ch["id"] != channel_id]
        _mark_dirty()
    _cache_invalidate()
    _join_channels_cache_clear()


# ─── Numbers & Countries ───────────────────────────────────────────────────────

def get_available_numbers_by_country(country_id, count=1, user_id=None):
    with _lock:
        available = [
            num for num, info in _DB["numbers"].items()
            if info["country_id"] == country_id and not info.get("used", False)
        ]
        picked = available[:count]
        for num in picked:
            _DB["numbers"][num]["used"] = True
            if user_id:
                _DB["number_assignments"][_normalize_number(num)] = user_id
        if picked:
            _mark_dirty()
    return picked


def _normalize_number(number: str) -> str:
    """Strip leading + and whitespace for consistent lookup."""
    return str(number).strip().lstrip('+')


def assign_number_to_user(number: str, user_id: int):
    """Explicitly record that a phone number was given to a Telegram user."""
    key = _normalize_number(number)
    with _lock:
        _DB["number_assignments"][key] = user_id
        _mark_dirty()


def get_user_id_by_number(number: str):
    """Return the Telegram user_id who received this phone number, or None."""
    key = _normalize_number(number)
    return _DB["number_assignments"].get(key)


def get_country_name_by_number(number: str) -> str:
    """Return country name for a phone number by looking it up in DB numbers."""
    key = _normalize_number(number)
    info = _DB["numbers"].get(key)
    if not info:
        return "Unknown"
    country_id = info.get("country_id")
    for c in _DB["countries"]:
        if c["id"] == country_id:
            return c["name"]
    return "Unknown"


def extract_otp_from_message(message: str) -> str:
    """Extract OTP (4-8 digit code, with or without dashes) from an SMS message."""
    import re
    patterns = [
        r'\b(\d{3}-\d{3})\b',   # 3-3 dash e.g. 914-094
        r'\b(\d{4}-\d{4})\b',   # 4-4 dash e.g. 1234-5678
        r'\b(\d{3}-\d{4})\b',   # 3-4 dash e.g. 123-4567
        r'\b(\d{4}-\d{3})\b',   # 4-3 dash e.g. 1234-567
        r'\b(\d{2}-\d{2}-\d{2})\b',  # 2-2-2 dash e.g. 12-34-56
        r'\b(\d{6})\b',          # 6-digit
        r'\b(\d{4})\b',          # 4-digit
        r'\b(\d{8})\b',          # 8-digit
        r'\b(\d{5})\b',          # 5-digit
        r'\b(\d{7})\b',          # 7-digit
    ]
    for pat in patterns:
        m = re.search(pat, message)
        if m:
            return m.group(1)
    return ""


def get_available_number_by_country(country_id):
    nums = get_available_numbers_by_country(country_id, count=1)
    return nums[0] if nums else None


def get_countries():
    result = [(c["id"], c["name"]) for c in sorted(_DB["countries"], key=lambda x: x["name"])]
    _countries_cache_set(result)
    return result


def get_numbers_count_by_country(country_id):
    nums = [v for v in _DB["numbers"].values() if v["country_id"] == country_id]
    total = len(nums)
    available = sum(1 for v in nums if not v.get("used", False))
    _numcount_cache_set(country_id, total, available)
    return total, available


def add_country(country_name):
    with _lock:
        if any(c["name"] == country_name for c in _DB["countries"]):
            return False
        next_id = max((c["id"] for c in _DB["countries"]), default=0) + 1
        _DB["countries"].append({"id": next_id, "name": country_name})
        _mark_dirty()
    _countries_cache_clear()
    return True


def add_numbers_to_country(country_id, numbers_list):
    added = 0
    with _lock:
        for raw in numbers_list:
            num = raw.strip()
            if not num or num in _DB["numbers"]:
                continue
            _DB["numbers"][num] = {
                "country_id": country_id, "used": False,
                "created_at": datetime.utcnow().isoformat()
            }
            added += 1
        if added:
            _mark_dirty()
    _numcount_cache_clear(country_id)
    return added


def delete_number(number):
    num = number.strip()
    with _lock:
        if num not in _DB["numbers"]:
            return False
        del _DB["numbers"][num]
        _mark_dirty()
    return True


def delete_all_numbers_from_country(country_id):
    with _lock:
        to_del = [k for k, v in _DB["numbers"].items() if v["country_id"] == country_id]
        for k in to_del:
            del _DB["numbers"][k]
        if to_del:
            _mark_dirty()
    return len(to_del)


def delete_country(country_id):
    nums_deleted = delete_all_numbers_from_country(country_id)
    with _lock:
        _DB["service_countries"] = [sc for sc in _DB["service_countries"] if sc["country_id"] != country_id]
        _DB["countries"] = [c for c in _DB["countries"] if c["id"] != country_id]
        _mark_dirty()
    _countries_cache_clear()
    _numcount_cache_clear()
    return nums_deleted, True


def get_country_stats():
    return [
        (c["name"], *get_numbers_count_by_country(c["id"]))
        for c in sorted(_DB["countries"], key=lambda x: x["name"])
    ]


# ─── Services ──────────────────────────────────────────────────────────────────

def add_service(name, custom_emoji_id=""):
    with _lock:
        if any(s["name"] == name for s in _DB["services"]):
            return False
        next_id = max((s["id"] for s in _DB["services"]), default=0) + 1
        _DB["services"].append({"id": next_id, "name": name, "custom_emoji_id": custom_emoji_id})
        _mark_dirty()
    _services_cache_clear()
    return True


def get_services():
    result = [(s["id"], s["name"]) for s in sorted(_DB["services"], key=lambda x: x["name"])]
    _services_cache_set(result)
    return result


def get_services_with_emoji():
    """Return (id, name, custom_emoji_id) — used for animated display."""
    return [
        (s["id"], s["name"], s.get("custom_emoji_id", ""))
        for s in sorted(_DB["services"], key=lambda x: x["name"])
    ]


def _svc_animated_tag(name: str, custom_emoji_id: str = "") -> str:
    """Return <tg-emoji> HTML for a service.
    Only uses manually stored custom_emoji_id — no auto-lookup by name."""
    if custom_emoji_id:
        fallback = get_service_emoji(name)
        return f'<tg-emoji emoji-id="{custom_emoji_id}">{fallback}</tg-emoji>'
    return ""


def add_country_to_service(service_id, country_id):
    with _lock:
        if any(sc["service_id"] == service_id and sc["country_id"] == country_id for sc in _DB["service_countries"]):
            return False
        _DB["service_countries"].append({"service_id": service_id, "country_id": country_id})
        _mark_dirty()
    _svc_countries_cache_clear(service_id)
    return True


def get_countries_by_service(service_id):
    ids = {sc["country_id"] for sc in _DB["service_countries"] if sc["service_id"] == service_id}
    all_c = {c["id"]: c["name"] for c in _DB["countries"]}
    result = sorted([(cid, all_c[cid]) for cid in ids if cid in all_c], key=lambda x: x[1])
    _svc_countries_cache_set(service_id, result)
    return result


def get_service_map():
    all_c = {c["id"]: c["name"] for c in _DB["countries"]}
    svc_map: dict = {}
    for sc in _DB["service_countries"]:
        svc_map.setdefault(sc["service_id"], []).append(all_c.get(sc["country_id"], ""))
    return [
        (s["id"], s["name"], sorted(svc_map.get(s["id"], [])))
        for s in sorted(_DB["services"], key=lambda x: x["name"])
    ]


def unlink_country_from_service(service_id, country_id):
    with _lock:
        before = len(_DB["service_countries"])
        _DB["service_countries"] = [
            sc for sc in _DB["service_countries"]
            if not (sc["service_id"] == service_id and sc["country_id"] == country_id)
        ]
        changed = len(_DB["service_countries"]) < before
        if changed:
            _mark_dirty()
    _svc_countries_cache_clear(service_id)
    return changed


def delete_service(name):
    with _lock:
        svc = next((s for s in _DB["services"] if s["name"].lower() == name.lower()), None)
        if not svc:
            return False
        sid = svc["id"]
        _DB["service_countries"] = [sc for sc in _DB["service_countries"] if sc["service_id"] != sid]
        _DB["services"] = [s for s in _DB["services"] if s["id"] != sid]
        _mark_dirty()
    _services_cache_clear()
    return True


# ─── Admins ────────────────────────────────────────────────────────────────────

def add_admin(user_id_int):
    key = str(user_id_int)
    with _lock:
        if key in _DB["admins"]:
            return False
        _DB["admins"][key] = {"user_id": user_id_int, "added_at": datetime.utcnow().isoformat()}
        _mark_dirty()
    _admin_cache_clear(user_id_int)
    return True


def remove_admin(user_id_int):
    if user_id_int in _PROTECTED_IDS:
        return False, "Cannot remove protected admin"
    key = str(user_id_int)
    with _lock:
        if key not in _DB["admins"]:
            return False, "Admin not found"
        del _DB["admins"][key]
        _mark_dirty()
    _admin_cache_clear(user_id_int)
    return True, "Admin removed successfully"


def get_all_admins():
    db_admins = [v["user_id"] for v in _DB["admins"].values()]
    return list(set(db_admins + _PROTECTED_IDS))


def get_all_admin_ids():
    return get_all_admins()


def is_admin(username, user_id=None):
    if user_id and user_id in _PROTECTED_IDS:
        return True
    if not user_id:
        return False
    cached = _admin_cache_get(user_id)
    if cached is not None:
        return cached
    result = str(user_id) in _DB["admins"]
    _admin_cache_set(user_id, result)
    return result


# ─── User Management ───────────────────────────────────────────────────────────

def add_user(user_id, username, first_name, last_name):
    key = str(user_id)
    with _lock:
        existing = _DB["users"].get(key, {})
        _DB["users"][key] = {
            "user_id": user_id,
            "username": username or "",
            "first_name": first_name or "",
            "last_name": last_name or "",
            "joined_at": existing.get("joined_at", datetime.utcnow().isoformat()),
            "is_verified": existing.get("is_verified", False),
            "verified_at": existing.get("verified_at"),
            "balance": existing.get("balance", 0),
            "total_earned": existing.get("total_earned", 0),
            "last_active": existing.get("last_active"),
            "usage_count": existing.get("usage_count", 0),
        }
        _mark_dirty()
    _ucount_cache_clear()


def update_user_activity(user_id):
    key = str(user_id)
    with _lock:
        u = _DB["users"].get(key)
        if u is None:
            return
        u["last_active"] = datetime.utcnow().isoformat()
        u["usage_count"] = u.get("usage_count", 0) + 1
        _mark_dirty()


def get_activity_stats():
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_ago  = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_ago  = datetime(now.year, now.month, now.day) - __import__('datetime').timedelta(days=7)
    month_ago = datetime(now.year, now.month, now.day) - __import__('datetime').timedelta(days=30)

    today_count = week_count = inactive_count = never_count = 0
    for u in _DB["users"].values():
        la = u.get("last_active")
        if not la:
            never_count += 1
            inactive_count += 1
            continue
        try:
            la_dt = datetime.fromisoformat(la)
        except Exception:
            never_count += 1
            inactive_count += 1
            continue
        if la_dt >= today_start:
            today_count += 1
        if la_dt >= week_ago:
            week_count += 1
        if la_dt < month_ago:
            inactive_count += 1

    return {
        "total": len(_DB["users"]),
        "today": today_count,
        "week": week_count,
        "inactive_30d": inactive_count,
        "never": never_count,
    }


def get_inactive_users(days=30):
    from datetime import timedelta
    cutoff = datetime.utcnow() - timedelta(days=days)
    result = []
    for u in _DB["users"].values():
        la = u.get("last_active")
        if not la:
            result.append(u)
            continue
        try:
            if datetime.fromisoformat(la) < cutoff:
                result.append(u)
        except Exception:
            result.append(u)
    return result


def get_all_users():
    return list(_DB["users"].values())


# ─── Banned users ───────────────────────────────────────────────────────────────

def ban_user(user_id: int, reason: str = "", banned_by: str = ""):
    with _lock:
        _DB.setdefault("banned_users", {})[str(user_id)] = {
            "user_id": user_id,
            "reason": reason,
            "banned_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
            "banned_by": banned_by,
        }
        _mark_dirty()

def unban_user(user_id: int) -> bool:
    with _lock:
        if str(user_id) in _DB.get("banned_users", {}):
            del _DB["banned_users"][str(user_id)]
            _mark_dirty()
            return True
        return False

def is_user_banned(user_id: int) -> bool:
    return str(user_id) in _DB.get("banned_users", {})

def get_banned_users() -> list:
    return list(_DB.get("banned_users", {}).values())


def get_user_count():
    val = len(_DB["users"])
    _ucount_cache_set(val)
    return val


def is_user_verified(user_id):
    return bool(_DB["users"].get(str(user_id), {}).get("is_verified", False))


def verify_user(user_id):
    key = str(user_id)
    with _lock:
        if key in _DB["users"]:
            _DB["users"][key]["is_verified"] = True
            _DB["users"][key]["verified_at"] = datetime.utcnow().isoformat()
            _mark_dirty()
    _cache_invalidate(user_id)
    return True


# ─── Referral ──────────────────────────────────────────────────────────────────

def get_referral_settings():
    s = _DB["settings"]
    result = {
        "enabled": s.get("referral_enabled", "1") == "1",
        "reward": int(s.get("referral_reward", "1")),
        "label": s.get("referral_label", "BDT"),
    }
    _rsettings_cache_set(result)
    return result


def set_referral_setting(field, value):
    key_map = {
        "enabled": "referral_enabled",
        "reward_per_referral": "referral_reward",
        "reward_label": "referral_label",
    }
    db_key = key_map.get(field, field)
    if field == "enabled":
        value = "1" if value else "0"
    _set_setting(db_key, value)
    _rsettings_cache_clear()


def get_user_balance_data(user_id):
    cached = _bal_cache_get(user_id)
    if cached is not None:
        return cached
    u = _DB["users"].get(str(user_id), {})
    result = {"balance": u.get("balance", 0), "total_earned": u.get("total_earned", 0)}
    _bal_cache_set(user_id, result)
    return result


def update_user_balance(user_id, amount):
    key = str(user_id)
    with _lock:
        if key in _DB["users"]:
            _DB["users"][key]["balance"] = _DB["users"][key].get("balance", 0) + amount
            _mark_dirty()
    _bal_cache_clear(user_id)


def _add_balance_db(user_id, amount):
    key = str(user_id)
    with _lock:
        if key in _DB["users"]:
            u = _DB["users"][key]
            u["balance"] = u.get("balance", 0) + amount
            u["total_earned"] = u.get("total_earned", 0) + amount
            _mark_dirty()
    _bal_cache_clear(user_id)


def get_country_otp_reward(country_name: str) -> float:
    """Return the OTP reward amount for a country (0 if not set)."""
    key = (country_name or "").strip().lower()
    rewards = _DB.get("country_otp_rewards", {})
    return float(rewards.get(key, 0))


def set_country_otp_reward(country_name: str, amount: float):
    """Set the OTP reward amount for a country."""
    key = (country_name or "").strip().lower()
    with _lock:
        _DB.setdefault("country_otp_rewards", {})[key] = amount
        _mark_dirty()


def get_all_country_otp_rewards() -> dict:
    """Return all country OTP rewards as {country_name_lower: amount}."""
    return dict(_DB.get("country_otp_rewards", {}))


def record_referral(referrer_id, referred_id):
    if referrer_id == referred_id:
        return False
    key = str(referred_id)
    with _lock:
        if key in _DB["referrals"]:
            return False
        _DB["referrals"][key] = {
            "referrer_id": referrer_id,
            "referred_id": referred_id,
            "created_at": datetime.utcnow().isoformat(),
        }
        _mark_dirty()
    settings = get_referral_settings()
    if settings["enabled"] and settings["reward"] > 0:
        _add_balance_db(referrer_id, settings["reward"])
    _ref_stats_cache_clear()
    return True


def get_user_referral_count(user_id):
    return sum(1 for v in _DB["referrals"].values() if v["referrer_id"] == user_id)


def get_referral_leaderboard(limit=10):
    counts: dict = {}
    for v in _DB["referrals"].values():
        rid = v["referrer_id"]
        counts[rid] = counts.get(rid, 0) + 1
    top = sorted(counts.items(), key=lambda x: x[1], reverse=True)[:limit]
    result = []
    for uid, cnt in top:
        u = _DB["users"].get(str(uid), {})
        result.append((uid, u.get("username", ""), u.get("first_name", "User"), cnt, u.get("total_earned", 0)))
    return result


def get_total_referral_stats():
    cached = _ref_stats_cache_get()
    if cached is not None:
        return cached
    total = len(_DB["referrals"])
    unique_ref = len({v["referrer_id"] for v in _DB["referrals"].values()})
    total_rewards = sum(u.get("total_earned", 0) for u in _DB["users"].values())
    result = (total, unique_ref, total_rewards)
    _ref_stats_cache_set(result)
    return result


# ─── Withdraw Config ───────────────────────────────────────────────────────────

def get_withdraw_config():
    cached = _wconfig_cache_get()
    if cached is not None:
        return cached
    s = _DB["settings"]
    result = {
        "enabled": s.get("withdraw_enabled", "1") == "1",
        "min_amount": int(s.get("withdraw_min_amount", "0")),
        "group_chat_id": s.get("withdraw_group_chat_id", ""),
    }
    _wconfig_cache_set(result)
    return result


def set_withdraw_enabled(enabled: bool):
    _set_setting("withdraw_enabled", "1" if enabled else "0")
    _wconfig_cache_clear()


def set_withdraw_min_amount(amount: int):
    _set_setting("withdraw_min_amount", str(amount))
    _wconfig_cache_clear()


def set_withdraw_group_chat_id(chat_id: str):
    _set_setting("withdraw_group_chat_id", chat_id)
    _wconfig_cache_clear()


def get_withdraw_stats():
    cached = _wstats_cache_get()
    if cached is not None:
        return cached
    today = datetime.utcnow().strftime("%Y-%m-%d")
    all_wr = list(_DB["withdraw_requests"].values())
    today_data = [r for r in all_wr if str(r.get("created_at", "")).startswith(today)]
    result = {
        "today_count": len(today_data),
        "today_pending": sum(1 for r in today_data if r.get("status") == "pending"),
        "total_count": len(all_wr),
        "total_pending": sum(1 for r in all_wr if r.get("status") == "pending"),
        "total_approved": sum(1 for r in all_wr if r.get("status") == "approved"),
        "total_rejected": sum(1 for r in all_wr if r.get("status") == "rejected"),
    }
    _wstats_cache_set(result)
    return result


# ─── Custom Emoji ───────────────────────────────────────────────────────────────

def get_all_custom_emojis():
    return [
        (k, v["emoji_id"], v.get("fallback_unicode", "📱"))
        for k, v in sorted(_DB["custom_emoji"].items())
    ]


def save_custom_emoji(service_name: str, emoji_id: str, fallback: str = "📱"):
    key = service_name.lower().strip()
    with _lock:
        _DB["custom_emoji"][key] = {"emoji_id": emoji_id.strip(), "fallback_unicode": fallback}
        _mark_dirty()


def delete_custom_emoji(service_name: str):
    key = service_name.lower().strip()
    with _lock:
        if key not in _DB["custom_emoji"]:
            return False
        del _DB["custom_emoji"][key]
        _mark_dirty()
    return True


# ─── Withdraw Requests ─────────────────────────────────────────────────────────

def create_withdraw_request(user_id, username, wallet_type, wallet_address, amount, label):
    key_u = str(user_id)
    with _lock:
        current = _DB["users"].get(key_u, {}).get("balance", 0)
        if current < amount:
            return None, "insufficient"
        _DB["users"][key_u]["balance"] = current - amount
        req_id = str(uuid.uuid4())[:12]
        _DB["withdraw_requests"][req_id] = {
            "id": req_id, "user_id": user_id, "username": username or "",
            "wallet_type": wallet_type, "wallet_address": wallet_address,
            "amount": amount, "label": label, "status": "pending",
            "created_at": datetime.utcnow().isoformat(),
        }
        _mark_dirty()
    _bal_cache_clear(user_id)
    return req_id, "ok"


def get_pending_withdrawals():
    rows = [
        (r["id"], r["user_id"], r.get("username", ""), r["wallet_type"],
         r["wallet_address"], r["amount"], r.get("label", "BDT"), r.get("created_at", ""))
        for r in _DB["withdraw_requests"].values()
        if r.get("status") == "pending"
    ]
    return sorted(rows, key=lambda x: x[7], reverse=True)


def get_withdraw_by_id(req_id):
    r = _DB["withdraw_requests"].get(str(req_id))
    if not r:
        return None
    return (r["id"], r["user_id"], r.get("username", ""), r["wallet_type"],
            r["wallet_address"], r["amount"], r.get("label", "BDT"), r.get("status", "pending"))


def update_withdraw_status(req_id, status, refund_user_id=None, refund_amount=0):
    with _lock:
        r = _DB["withdraw_requests"].get(str(req_id))
        if r:
            r["status"] = status
        if status == "rejected" and refund_user_id and refund_amount > 0:
            key_u = str(refund_user_id)
            if key_u in _DB["users"]:
                _DB["users"][key_u]["balance"] = _DB["users"][key_u].get("balance", 0) + refund_amount
        _mark_dirty()
    _wstats_cache_clear()
    if refund_user_id:
        _bal_cache_clear(refund_user_id)


def get_all_withdrawals(limit=30):
    rows = [
        (r["id"], r["user_id"], r.get("username", ""), r["wallet_type"],
         r["wallet_address"], r["amount"], r.get("label", "BDT"),
         r.get("status", "pending"), r.get("created_at", ""))
        for r in _DB["withdraw_requests"].values()
    ]
    return sorted(rows, key=lambda x: x[8], reverse=True)[:limit]


def get_user_withdraw_history(user_id, limit=5):
    rows = [
        (r["id"], r["wallet_type"], r["wallet_address"], r["amount"],
         r.get("label", "BDT"), r.get("status", "pending"), r.get("created_at", ""))
        for r in _DB["withdraw_requests"].values()
        if r["user_id"] == user_id
    ]
    return sorted(rows, key=lambda x: x[6], reverse=True)[:limit]


# ─── Admin Balance ─────────────────────────────────────────────────────────────

def admin_add_balance(user_id, amount):
    key = str(user_id)
    with _lock:
        if key in _DB["users"]:
            u = _DB["users"][key]
            u["balance"] = u.get("balance", 0) + amount
            u["total_earned"] = u.get("total_earned", 0) + amount
            _mark_dirty()
    _bal_cache_clear(user_id)
    _ref_stats_cache_clear()
    return _DB["users"].get(key, {}).get("balance", 0)


def admin_remove_balance(user_id, amount):
    key = str(user_id)
    if key not in _DB["users"]:
        return None, "not_found"
    current = _DB["users"][key].get("balance", 0)
    if amount > current:
        return current, "insufficient"
    with _lock:
        _DB["users"][key]["balance"] = current - amount
        _mark_dirty()
    _bal_cache_clear(user_id)
    return current - amount, None


# ─── SMS Panels ────────────────────────────────────────────────────────────────

def add_sms_panel(name: str, login_url: str, message_url: str,
                  username: str, password: str, column_map: dict = None) -> str:
    panel_id = str(uuid.uuid4())[:12]
    with _lock:
        _DB["sms_panels"][panel_id] = {
            "id": panel_id, "name": name,
            "login_url": login_url, "message_url": message_url,
            "username": username, "password": password,
            "enabled": True,
            "created_at": datetime.utcnow().isoformat(),
            "column_map": column_map or {},
        }
        _mark_dirty()
    return panel_id


def get_sms_panels() -> list:
    return list(_DB["sms_panels"].values())


def get_sms_panel(panel_id: str) -> dict | None:
    return _DB["sms_panels"].get(str(panel_id))


def delete_sms_panel(panel_id: str) -> bool:
    with _lock:
        if panel_id not in _DB["sms_panels"]:
            return False
        del _DB["sms_panels"][panel_id]
        _DB["seen_sms_messages"].pop(panel_id, None)
        _mark_dirty()
    return True


def toggle_sms_panel(panel_id: str) -> bool | None:
    with _lock:
        p = _DB["sms_panels"].get(panel_id)
        if not p:
            return None
        p["enabled"] = not p["enabled"]
        _mark_dirty()
    return p["enabled"]


def set_global_forward_chat(chat_id: str):
    """Set the global auto-forward chat ID for ALL panels."""
    with _lock:
        _DB["global_forward_chat_id"] = str(chat_id).strip()
        _mark_dirty()


def get_global_forward_chat() -> str:
    """Return the global auto-forward chat ID, or empty string."""
    return _DB.get("global_forward_chat_id", "")


def remove_global_forward_chat():
    """Remove the global auto-forward chat ID."""
    with _lock:
        _DB["global_forward_chat_id"] = ""
        _mark_dirty()


def set_panel_forward_chat(panel_id: str, chat_id: str) -> bool:
    """Set the auto-forward chat ID for a panel."""
    with _lock:
        p = _DB["sms_panels"].get(panel_id)
        if not p:
            return False
        p["forward_chat_id"] = str(chat_id).strip()
        _mark_dirty()
    return True


def get_panel_forward_chat(panel_id: str) -> str | None:
    """Return the auto-forward chat ID for a panel, or None."""
    p = _DB["sms_panels"].get(panel_id)
    return p.get("forward_chat_id") if p else None


def remove_panel_forward_chat(panel_id: str) -> bool:
    """Remove the auto-forward chat ID from a panel."""
    with _lock:
        p = _DB["sms_panels"].get(panel_id)
        if not p:
            return False
        p.pop("forward_chat_id", None)
        _mark_dirty()
    return True


def update_sms_panel_credentials(panel_id: str, username: str = None, password: str = None) -> bool:
    with _lock:
        p = _DB["sms_panels"].get(panel_id)
        if not p:
            return False
        if username is not None:
            p["username"] = username
        if password is not None:
            p["password"] = password
        _mark_dirty()
    return True


def update_sms_panel_interval(panel_id: str, seconds: int) -> bool:
    with _lock:
        p = _DB["sms_panels"].get(panel_id)
        if not p:
            return False
        p["poll_interval"] = seconds
        _mark_dirty()
    return True


def get_sms_panel_interval(panel_id: str) -> int:
    """Return poll interval in seconds for a panel (default 3s)."""
    p = _DB["sms_panels"].get(panel_id)
    if not p:
        return 3
    return int(p.get("poll_interval", 3))


# ─── Seen SMS tracking ─────────────────────────────────────────────────────────

def is_sms_seen(panel_id: str, sms_id: str) -> bool:
    return sms_id in _DB["seen_sms_messages"].get(panel_id, {})


def mark_sms_seen(panel_id: str, sms_id: str):
    with _lock:
        _DB["seen_sms_messages"].setdefault(panel_id, {})[sms_id] = int(time.time())
        _mark_dirty()


def get_today_panel_message_count(panel_id: str) -> int:
    """Count how many messages have been received in the last 24 hours for a given panel."""
    cutoff_ts = int(time.time()) - 86400  # last 24 hours
    seen = _DB["seen_sms_messages"].get(panel_id, {})
    return sum(1 for ts in seen.values() if ts >= cutoff_ts)


def increment_panel_user_count(panel_id: str):
    """Increment the count of messages sent to users for a panel."""
    with _lock:
        _DB.setdefault("panel_send_counts", {}).setdefault(panel_id, {"user": 0, "group": 0})
        _DB["panel_send_counts"][panel_id]["user"] += 1
        _mark_dirty()


def increment_panel_group_count(panel_id: str):
    """Increment the count of messages forwarded to group for a panel."""
    with _lock:
        _DB.setdefault("panel_send_counts", {}).setdefault(panel_id, {"user": 0, "group": 0})
        _DB["panel_send_counts"][panel_id]["group"] += 1
        _mark_dirty()


def get_panel_send_counts(panel_id: str) -> dict:
    """Return {user: int, group: int} send counts for a panel."""
    return _DB.get("panel_send_counts", {}).get(panel_id, {"user": 0, "group": 0})


def store_last_panel_message(panel_id: str, number: str, message: str, country: str = "", date: str = "", sender: str = ""):
    """Store the most recent SMS received by a panel."""
    with _lock:
        _DB.setdefault("last_panel_messages", {})[panel_id] = {
            "number": number,
            "message": message,
            "country": country,
            "sender": sender,
            "time": date if date else datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
        }
        _mark_dirty()


def get_last_panel_message(panel_id: str) -> dict | None:
    """Return the last stored SMS for a panel, or None."""
    return _DB.get("last_panel_messages", {}).get(panel_id)


def cleanup_seen_sms(max_age_seconds: int = 86400 * 7):
    """Remove seen SMS records older than max_age_seconds (default 7 days)."""
    cutoff = int(time.time()) - max_age_seconds
    with _lock:
        for panel_id in list(_DB["seen_sms_messages"].keys()):
            _DB["seen_sms_messages"][panel_id] = {
                k: v for k, v in _DB["seen_sms_messages"][panel_id].items()
                if v > cutoff
            }
        _mark_dirty()


# ─── Cache stubs ───────────────────────────────────────────────────────────────

def _npr_cache_clear(): pass
def _otp_cache_set(v): pass
def _otp_cache_clear(): pass
def _ci_cache_clear(): pass
def _custmsg_cache_set(v): pass
def _custmsg_cache_clear(): pass
def _join_channels_cache_set(v): pass
def _join_channels_cache_clear(): pass
def _cache_invalidate(uid=None): pass
def _numcount_cache_set(cid, t, a): pass
def _numcount_cache_clear(cid=None): pass
def _countries_cache_set(v): pass
def _countries_cache_clear(): pass
def _services_cache_set(v): pass
def _services_cache_clear(): pass
def _svc_countries_cache_set(sid, v): pass
def _svc_countries_cache_clear(sid=None): pass
def _admin_cache_get(uid): return None
def _admin_cache_set(uid, v): pass
def _admin_cache_clear(uid=None): pass
def _ucount_cache_set(v): pass
def _ucount_cache_clear(): pass
def _rsettings_cache_set(v): pass
def _rsettings_cache_clear(): pass
def _bal_cache_get(uid): return None
def _bal_cache_set(uid, v): pass
def _bal_cache_clear(uid=None): pass
def _ref_stats_cache_get(): return None
def _ref_stats_cache_set(v): pass
def _ref_stats_cache_clear(): pass
def _wconfig_cache_get(): return None
def _wconfig_cache_set(v): pass
def _wconfig_cache_clear(): pass
def _wstats_cache_get(): return None
def _wstats_cache_set(v): pass
def _wstats_cache_clear(): pass

# ════════════════════════════════════════════════════════════════════════
# SMS PANEL CHECKER  (formerly sms_panel_checker.py — inlined)
# ════════════════════════════════════════════════════════════════════════
_sms_panel_logger = logging.getLogger('sms_panel_checker')
_smshadi_last_api_call: dict = {}   # key: login_url → last call timestamp (float)
_SMSHADI_API_MIN_INTERVAL = 16      # smshadi panel requires 15s minimum between CDR API calls

# phone number pattern — match typical international / local numbers
_PHONE_RE = re.compile(r'\+?\d[\d\s\-().]{6,17}\d')

_active_tasks: dict[str, asyncio.Task] = {}   # panel_id → running Task
_panel_sessions: dict = {}                      # panel_id → authenticated requests.Session


def extract_service_from_message(message: str) -> str:
    """Extract service/app name from an SMS message."""
    if not message:
        return ""
    # Pattern: [#][ServiceName] e.g. "[#][TikTok]"
    m = re.search(r'\[#\]\[([^\]]+)\]', message)
    if m:
        return m.group(1).strip()
    # Pattern: [ServiceName] alone
    m = re.search(r'\[([A-Za-z][A-Za-z0-9\s\-\.]{1,30})\]', message)
    if m:
        return m.group(1).strip()
    # Pattern: first capitalized word/phrase before OTP keyword
    m = re.search(r'^([A-Z][A-Za-z0-9\s]{1,25}?)[\s\-:]*(?:code|otp|verification|is your|one-time)', message, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return ""


# ─── Login helper ─────────────────────────────────────────────────────────────

def _solve_math_captcha(html: str) -> Optional[int]:
    """Parse 'What is N OP M = ?' style captcha and return the answer.
    Handles +, -, *, x, × operators.
    """
    from bs4 import BeautifulSoup as _BS
    # Extract plain text so HTML entities / tags don't interfere
    plain = _BS(html, 'html.parser').get_text(' ')
    # Primary: "What is N OP M" pattern
    m = re.search(r'[Ww]hat\s+is\s+(\d+)\s*([\+\-\*x×])\s*(\d+)', plain)
    if not m:
        # Fallback: any "N OP M =" pattern in raw HTML
        m = re.search(r'(\d+)\s*([\+\-\*x×])\s*(\d+)\s*=', html)
    if m:
        a, op, b = int(m.group(1)), m.group(2).strip(), int(m.group(3))
        if op == '+':
            return a + b
        if op == '-':
            return a - b
        if op in ('*', 'x', '×'):
            return a * b
    return None


def _parse_dashboard_stats(html: str) -> dict:
    """
    Parse SMS stats from the panel dashboard page.
    Returns dict with keys: yesterday, this_week, this_month, this_year, today_total
    """
    from bs4 import BeautifulSoup as _BS
    text = _BS(html, 'html.parser').get_text(separator='\n', strip=True)
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    stats = {}
    for i, line in enumerate(lines):
        nxt = lines[i+1] if i+1 < len(lines) else ''
        def _int(s):
            try: return int(s.replace(',','').replace('.',''))
            except: return None
        if 'Yesterday SMS' in line:
            v = _int(nxt)
            if v is not None: stats['yesterday'] = v
        elif 'SMS This Week' in line:
            v = _int(nxt)
            if v is not None: stats['this_week'] = v
        elif 'SMS This Month' in line:
            v = _int(nxt)
            if v is not None: stats['this_month'] = v
        elif 'SMS This Year' in line:
            v = _int(nxt)
            if v is not None: stats['this_year'] = v
        elif 'Today SMS' in line or 'SMS Today' in line:
            v = _int(nxt)
            if v is not None: stats['today'] = v
    return stats


def _fetch_panel_stats(panel: dict) -> dict | None:
    """Login to the panel and return dashboard stats dict, or None on failure."""
    from urllib.parse import urlparse, urljoin
    session = requests.Session()
    session.headers.update({'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'})
    ok = _do_login(session, panel['login_url'], panel['username'], panel['password'])
    if not ok:
        return None
    parsed = urlparse(panel['login_url'])
    base_url = urljoin(f"{parsed.scheme}://{parsed.netloc}", '/agent/')
    try:
        resp = session.get(base_url, timeout=15, allow_redirects=True)
        if 'login' in resp.url.lower():
            return None
        stats = _parse_dashboard_stats(resp.text)
        return stats if stats else None
    except Exception as e:
        _sms_panel_logger.warning(f"_fetch_panel_stats error: {e}")
        return None


def _country_from_range(range_name: str) -> str:
    """Extract country name from SMS range name like 'BANGLADESH 58' → 'Bangladesh'."""
    if not range_name:
        return ""
    first_word = range_name.strip().split()[0]
    return first_word.title()


# ─── Group Forward Message Builder ────────────────────────────────────────────

COUNTRY_SHORT_CODES = {
    'afghanistan':'AF','albania':'AL','algeria':'DZ','andorra':'AD',
    'angola':'AO','argentina':'AR','armenia':'AM','australia':'AU',
    'austria':'AT','azerbaijan':'AZ','bahamas':'BS','bahrain':'BH',
    'bangladesh':'BD','barbados':'BB','belarus':'BY','belgium':'BE',
    'belize':'BZ','benin':'BJ','bhutan':'BT','bolivia':'BO',
    'bosnia':'BA','bosnia and herzegovina':'BA','botswana':'BW',
    'brazil':'BR','brunei':'BN','bulgaria':'BG','burkina faso':'BF',
    'burundi':'BI','cambodia':'KH','cameroon':'CM','canada':'CA',
    'cape verde':'CV','cabo verde':'CV','central african republic':'CF',
    'chad':'TD','chile':'CL','china':'CN','colombia':'CO',
    'comoros':'KM','congo':'CG','croatia':'HR','cuba':'CU',
    'cyprus':'CY','czech republic':'CZ','czechia':'CZ','denmark':'DK',
    'djibouti':'DJ','dominican republic':'DO','drc':'CD',
    'democratic republic of congo':'CD','ecuador':'EC','egypt':'EG',
    'el salvador':'SV','eritrea':'ER','estonia':'EE','ethiopia':'ET',
    'fiji':'FJ','finland':'FI','france':'FR','gabon':'GA',
    'gambia':'GM','georgia':'GE','germany':'DE','ghana':'GH',
    'greece':'GR','guatemala':'GT','guinea':'GN','guinea bissau':'GW',
    'guyana':'GY','haiti':'HT','honduras':'HN','hong kong':'HK',
    'hungary':'HU','iceland':'IS','india':'IN','indonesia':'ID',
    'iran':'IR','iraq':'IQ','ireland':'IE','israel':'IL',
    'italy':'IT','ivory coast':'CI','jamaica':'JM','japan':'JP',
    'jordan':'JO','kazakhstan':'KZ','kenya':'KE','kosovo':'XK',
    'kuwait':'KW','kyrgyzstan':'KG','laos':'LA','latvia':'LV',
    'lebanon':'LB','lesotho':'LS','liberia':'LR','libya':'LY',
    'liechtenstein':'LI','lithuania':'LT','luxembourg':'LU',
    'macau':'MO','madagascar':'MG','malawi':'MW','malaysia':'MY',
    'maldives':'MV','mali':'ML','malta':'MT','mauritania':'MR',
    'mauritius':'MU','mexico':'MX','moldova':'MD','monaco':'MC',
    'mongolia':'MN','montenegro':'ME','morocco':'MA','mozambique':'MZ',
    'myanmar':'MM','namibia':'NA','nepal':'NP','netherlands':'NL',
    'new zealand':'NZ','nicaragua':'NI','niger':'NE','nigeria':'NG',
    'north korea':'KP','north macedonia':'MK','norway':'NO','oman':'OM',
    'pakistan':'PK','palestine':'PS','panama':'PA','papua new guinea':'PG',
    'paraguay':'PY','peru':'PE','philippines':'PH','poland':'PL',
    'portugal':'PT','qatar':'QA','romania':'RO','russia':'RU',
    'rwanda':'RW','saudi arabia':'SA','senegal':'SN','serbia':'RS',
    'seychelles':'SC','sierra leone':'SL','singapore':'SG','slovakia':'SK',
    'somalia':'SO','south africa':'ZA','south korea':'KR','south sudan':'SS',
    'spain':'ES','sri lanka':'LK','sudan':'SD','suriname':'SR',
    'sweden':'SE','switzerland':'CH','syria':'SY','taiwan':'TW',
    'tajikistan':'TJ','tanzania':'TZ','thailand':'TH',
    'timor leste':'TL','east timor':'TL','togo':'TG',
    'trinidad and tobago':'TT','tunisia':'TN','turkey':'TR',
    'turkiye':'TR','turkmenistan':'TM','uganda':'UG','ukraine':'UA',
    'united arab emirates':'AE','uae':'AE','united kingdom':'GB',
    'uk':'GB','united states':'US','usa':'US','america':'US',
    'uruguay':'UY','uzbekistan':'UZ','venezuela':'VE','vietnam':'VN',
    'yemen':'YE','zambia':'ZM','zimbabwe':'ZW','san marino':'SM',
    'vatican':'VA','costa rica':'CR',
}

SERVICE_SHORT_CODES = {
    'whatsapp':'WS','telegram':'TG','facebook':'FB','instagram':'IG',
    'gmail':'GM','google':'GL','tiktok':'TK','twitter':'TX','x':'TX',
    'snapchat':'SC','viber':'VB','wechat':'WC','line':'LN',
    'discord':'DC','linkedin':'LI','uber':'UB','lyft':'LF',
    'amazon':'AM','apple':'AP','microsoft':'MS','netflix':'NF',
    'spotify':'SP','paypal':'PP','binance':'BN','coinbase':'CB',
    'trust wallet':'TW','metamask':'MM','bybit':'BB','okx':'OX',
    'yahoo':'YH','reddit':'RD','pinterest':'PT','twitch':'TC',
    'signal':'SI','imo':'IM','skype':'SK','zoom':'ZM',
    'shopee':'SH','lazada':'LZ','airbnb':'AB','booking':'BK',
    'grab':'GB','grabfood':'GF','foodpanda':'FP','doordash':'DD',
    'door dash':'DD','tinder':'TD','bumble':'BM','hinge':'HG',
    'twitter x':'TX',
}

def _get_country_flag(country: str) -> str:
    key = (country or '').strip().lower()
    return COUNTRY_UNICODE_FLAGS.get(key, '🌐')

def _get_country_code(country: str) -> str:
    key = (country or '').strip().lower()
    code = COUNTRY_SHORT_CODES.get(key)
    if code:
        return code
    # Try first word (e.g. "Iraq 964" → "Iraq")
    first = key.split()[0] if key else ''
    return COUNTRY_SHORT_CODES.get(first, country[:2].upper() if country else 'XX')

def _get_service_code(service: str) -> str:
    key = (service or '').strip().lower()
    return SERVICE_SHORT_CODES.get(key, service[:2].upper() if service else 'XX')

def _mask_number(number: str) -> str:
    """First 3 digits + SX-TEAM + last 3 digits. e.g. '9647828520254' → '964SX-TEAM254'"""
    n = (number or '').lstrip('+').strip()
    if len(n) >= 6:
        return f"{n[:3]}SX-TEAM{n[-3:]}"
    return n

def _build_group_forward(number: str, message: str, country: str,
                          otp: str, service: str, sender: str):
    """
    Build new-style group forward message + inline keyboard.
    Returns (text: str, reply_markup: InlineKeyboardMarkup).
    """
    flag      = _get_country_flag(country)
    ctry_code = _get_country_code(country)
    svc_raw   = service or sender or ''
    svc_code  = _get_service_code(svc_raw)
    masked    = _mask_number(number)

    text = f"{flag} {ctry_code} | {svc_code} <code>{masked}</code>"

    keyboard = []
    if otp:
        keyboard.append([
            InlineKeyboardButton(
                text=f"🔑 {otp}",
                copy_text=CopyTextButton(text=otp),
                style="danger"
            )
        ])
    number_link  = get_group_number_btn_link()
    channel_link = get_group_channel_btn_link()
    keyboard.append([
        InlineKeyboardButton("📱 Number", url=number_link, style="success"),
        InlineKeyboardButton("📢 Channel", url=channel_link, style="primary"),
    ])
    return text, InlineKeyboardMarkup(keyboard)


def _fetch_smshadi_sms(session: requests.Session, login_url: str) -> list[dict]:
    """
    Fetch received SMS from smshadi-type panels via the SMSCDRStats AJAX API
    (agent/res/data_smscdr.php, fg=0 — individual CDR records).

    Column layout returned by the API:
      [0] Date   [1] Range   [2] Number(destination)   [3] CLI(sender)
      [4] Client [5] SMS content   [6] Currency   [7] My Payout   [8] Client Payout

    Returns list of {id, number, message, country, date} dicts, newest first.
    """
    import time
    from urllib.parse import urlparse
    try:
        parsed = urlparse(login_url)
        base = f"{parsed.scheme}://{parsed.netloc}/agent/"

        # ── Rate-limit enforcement (panel requires ≥15 s between CDR API calls) ──
        now = time.time()
        last_call = _smshadi_last_api_call.get(login_url, 0)
        elapsed = now - last_call
        if elapsed < _SMSHADI_API_MIN_INTERVAL:
            wait_s = _SMSHADI_API_MIN_INTERVAL - elapsed
            _sms_panel_logger.debug(
                f"[Smshadi] Rate-limiting: waiting {wait_s:.1f}s before CDR API call"
            )
            time.sleep(wait_s)

        # ── Fetch SMSCDRStats page to extract the sesskey ──
        stats_page = session.get(base + 'SMSCDRStats', timeout=15,
                                  headers={'Referer': base})
        if 'login' in (getattr(stats_page, 'url', '') or '').lower():
            _sms_panel_logger.info("[Smshadi] Session expired on SMSCDRStats fetch")
            return []

        sesskey = ''
        m = re.search(r'sesskey=([A-Za-z0-9+/=]+)', stats_page.text)
        if m:
            sesskey = m.group(1)
        if not sesskey:
            _sms_panel_logger.warning("[Smshadi] Could not extract sesskey from SMSCDRStats page")
            return []

        # ── Build API URL — today's date range, individual records (fg=0) ──
        today = datetime.utcnow().strftime('%Y-%m-%d')
        api_url = (
            f"{base}res/data_smscdr.php"
            f"?fdate1={today}%2000:00:00&fdate2={today}%2023:59:59"
            f"&frange=&fclient=&fnum=&fcli=&fgdate=&fgmonth="
            f"&fgrange=&fgclient=&fgnumber=&fgcli="
            f"&fg=0&sesskey={sesskey}"
            f"&iDisplayStart=0&iDisplayLength=50&sEcho=1&iSortCol_0=0&sSortDir_0=desc"
        )

        _smshadi_last_api_call[login_url] = time.time()
        r = session.get(api_url, timeout=20,
                        headers={'X-Requested-With': 'XMLHttpRequest',
                                 'Referer': base + 'SMSCDRStats'})
        if r.status_code != 200:
            _sms_panel_logger.warning(f"[Smshadi] data_smscdr.php returned HTTP {r.status_code}")
            return []

        # Panel returns a plain-text rate-limit warning instead of JSON when too fast
        if 'atleast 15 second' in r.text or 'Refresh must be done' in r.text:
            _sms_panel_logger.warning("[Smshadi] Panel rate-limit message received; skipping")
            return []

        try:
            data = r.json()
        except Exception:
            _sms_panel_logger.warning(f"[Smshadi] JSON parse error — raw: {r.text[:300]}")
            return []

        results = []
        for row in data.get('aaData', []):
            if not isinstance(row, list) or len(row) < 6:
                continue
            date_str = str(row[0]) if row[0] else ''
            if not re.match(r'\d{4}-\d{2}-\d{2}', date_str):
                continue  # skip header/footer/summary rows
            range_name = str(row[1]) if row[1] else ''
            number     = re.sub(r'[\s\-()+.]', '', str(row[2])) if row[2] else ''
            cli        = str(row[3]) if row[3] else ''
            sms_msg    = str(row[5]) if row[5] else ''   # ← col 5 = actual SMS text

            country = _country_from_range(range_name)
            # Unique ID based on date + number + first 30 chars of message
            uid = hashlib.md5(
                f"{date_str}::{number}::{sms_msg[:30]}".encode()
            ).hexdigest()[:16]

            results.append({
                'id':         uid,
                'number':     number.lstrip('+'),
                'message':    sms_msg,
                'country':    country,
                'range_name': range_name,
                'cli':        cli,
                'date':       date_str,
            })
        _sms_panel_logger.info(
            f"[Smshadi] data_smscdr.php returned {len(results)} record(s) for {today}"
        )
        return results
    except Exception as e:
        _sms_panel_logger.warning(f"_fetch_smshadi_sms error: {e}")
        return []


def _is_smshadi_panel(panel: dict) -> bool:
    """Detect if a panel is smshadi-type (uses data_smscdr.php CDR API)."""
    login_url = panel.get('login_url', '').lower()
    return 'smshadi' in login_url or 'smshadi.net' in login_url


def _is_cdr_panel(panel: dict) -> bool:
    """Detect ANY panel that uses the data_smscdr.php CDR JSON API
    (smshadi, Inteliotech/Seven1Tel, and any other compatible panel software).
    Matches: SMSCDRStats, SMSDashboard, or any URL with /agent/ (Inteliotech pattern)."""
    msg_url = panel.get('message_url', '').lower()
    return ('smscdrstat' in msg_url or 'smscdr' in msg_url
            or '/agent/' in msg_url or 'smsdashboard' in msg_url)


def _get_agent_base(panel: dict) -> str:
    """Derive the agent base URL from the panel's message_url.
    e.g. http://94.23.120.156/ints/agent/SMSCDRStats → http://94.23.120.156/ints/agent/
    e.g. http://smshadi.net/agent/SMSCDRStats        → http://smshadi.net/agent/
    """
    from urllib.parse import urlparse, urlunparse
    msg_url = panel.get('message_url', '')
    parsed = urlparse(msg_url)
    path_parts = parsed.path.rstrip('/').split('/')
    base_path = '/'.join(path_parts[:-1]) + '/'
    return urlunparse(parsed._replace(path=base_path, query='', fragment=''))


def _fetch_cdr_sms(session: requests.Session, panel: dict,
                   skip_rate_limit: bool = False) -> list[dict]:
    """
    Generic CDR API fetch for ALL panels using data_smscdr.php
    (smshadi, Inteliotech/Seven1Tel, and compatible panel software).
    Derives the correct agent base path from panel['message_url'].

    Column layout returned by the API:
      [0] Date   [1] Range   [2] Number(destination)   [3] CLI(sender)
      [4] Client [5] SMS content   [6] Currency   [7] My Payout   [8] Client Payout

    Returns list of {id, number, message, country, date, cli} dicts, newest first.
    skip_rate_limit=True bypasses the 16s minimum interval (use for manual button presses).
    """
    import time
    login_url = panel.get('login_url', '')
    try:
        agent_base = _get_agent_base(panel)

        # Rate-limit guard — skip for manual/on-demand fetches
        if not skip_rate_limit:
            now = time.time()
            last_call = _smshadi_last_api_call.get(login_url, 0)
            elapsed = now - last_call
            if elapsed < _SMSHADI_API_MIN_INTERVAL:
                wait_s = _SMSHADI_API_MIN_INTERVAL - elapsed
                _sms_panel_logger.debug(
                    f"[CDR:{login_url}] Rate-limiting: waiting {wait_s:.1f}s"
                )
                time.sleep(wait_s)

        # Fetch SMSCDRStats page to extract the sesskey
        stats_page = session.get(agent_base + 'SMSCDRStats', timeout=15,
                                  headers={'Referer': agent_base})
        if 'login' in (getattr(stats_page, 'url', '') or '').lower():
            _sms_panel_logger.info(f"[CDR:{login_url}] Session expired on SMSCDRStats fetch")
            return None  # None = session expired (distinct from [] = no results)

        sesskey = ''
        m = re.search(r'sesskey=([A-Za-z0-9+/=]+)', stats_page.text)
        if m:
            sesskey = m.group(1)
        if not sesskey:
            _sms_panel_logger.warning(f"[CDR:{login_url}] Could not extract sesskey")
            return []

        today = datetime.utcnow().strftime('%Y-%m-%d')
        api_url = (
            f"{agent_base}res/data_smscdr.php"
            f"?fdate1={today}%2000:00:00&fdate2={today}%2023:59:59"
            f"&frange=&fclient=&fnum=&fcli=&fgdate=&fgmonth="
            f"&fgrange=&fgclient=&fgnumber=&fgcli="
            f"&fg=0&sesskey={sesskey}"
            f"&iDisplayStart=0&iDisplayLength=50&sEcho=1&iSortCol_0=0&sSortDir_0=desc"
        )

        _smshadi_last_api_call[login_url] = time.time()
        r = session.get(api_url, timeout=20,
                        headers={'X-Requested-With': 'XMLHttpRequest',
                                 'Referer': agent_base + 'SMSCDRStats'})
        if r.status_code != 200:
            _sms_panel_logger.warning(f"[CDR:{login_url}] data_smscdr.php returned HTTP {r.status_code}")
            return []

        if 'atleast 15 second' in r.text or 'Refresh must be done' in r.text:
            _sms_panel_logger.warning(f"[CDR:{login_url}] Panel rate-limit message; skipping")
            return []

        try:
            data = r.json()
        except Exception:
            _sms_panel_logger.warning(f"[CDR:{login_url}] JSON parse error — raw: {r.text[:300]}")
            return []

        results = []
        for row in data.get('aaData', []):
            if not isinstance(row, list) or len(row) < 6:
                continue
            date_str = str(row[0]) if row[0] else ''
            if not re.match(r'\d{4}-\d{2}-\d{2}', date_str):
                continue
            range_name = str(row[1]) if row[1] else ''
            number     = re.sub(r'[\s\-()+.]', '', str(row[2])) if row[2] else ''
            cli        = str(row[3]) if row[3] else ''
            sms_msg    = str(row[5]) if row[5] else ''

            country = _country_from_range(range_name)
            uid = hashlib.md5(
                f"{date_str}::{number}::{sms_msg[:30]}".encode()
            ).hexdigest()[:16]

            results.append({
                'id':         uid,
                'number':     number.lstrip('+'),
                'message':    sms_msg,
                'country':    country,
                'range_name': range_name,
                'cli':        cli,
                'date':       date_str,
            })

        _sms_panel_logger.info(
            f"[CDR:{login_url}] data_smscdr.php returned {len(results)} record(s) for {today}"
        )
        return results
    except Exception as e:
        _sms_panel_logger.warning(f"_fetch_cdr_sms error ({login_url}): {e}")
        return []


def _fetch_panel_last_messages(panel: dict, count: int = 3):
    """Fetch the latest `count` SMS records from the panel.
    For CDR panels (Seven1Tel, smshadi, etc.) uses the JSON API.
    For generic panels parses the HTML message page.
    Returns a list of dicts or {'error': 'login_failed' | 'session_expired'}."""
    from urllib.parse import urlparse, urlunparse
    panel_id = panel['id']
    session = _panel_sessions.get(panel_id)
    if session is None:
        return {'error': 'login_failed'}
    try:
        if _is_cdr_panel(panel):
            msgs = _fetch_cdr_sms(session, panel, skip_rate_limit=True)
            if msgs is None:
                # None means actual session expiry (redirected to login page)
                _panel_sessions.pop(panel_id, None)
                return {'error': 'session_expired'}
            if msgs == []:
                # Genuinely no messages today — not a session error
                return []
            def _parse_date(d):
                try:
                    return datetime.strptime(d, "%Y-%m-%d %H:%M:%S")
                except Exception:
                    return datetime.min
            msgs_sorted = sorted(msgs, key=lambda x: _parse_date(x.get('date', '')), reverse=True)
            result = []
            for m in msgs_sorted[:count]:
                result.append({
                    'number':  m['number'],
                    'message': m['message'],
                    'country': m.get('country', '') or m.get('range_name', ''),
                    'sender':  m.get('cli', ''),
                    'time':    m.get('date', ''),
                })
            return result
        # Generic HTML panels
        _login_parsed = urlparse(panel['login_url'])
        _msg_parsed   = urlparse(panel['message_url'])
        if _msg_parsed.netloc.replace('www.', '') == _login_parsed.netloc.replace('www.', ''):
            message_url = urlunparse(_msg_parsed._replace(
                scheme=_login_parsed.scheme, netloc=_login_parsed.netloc))
        else:
            message_url = panel['message_url']
        resp = session.get(message_url, timeout=15, allow_redirects=True)
        if 'login' in (getattr(resp, 'url', '') or '').lower():
            return {'error': 'session_expired'}
        raw = _extract_messages(resp.text)
        if not raw:
            return []
        result = []
        for m in raw[:count]:
            result.append({
                'number':  m['number'].lstrip('+'),
                'message': m['message'],
                'country': '',
                'sender':  '',
                'time':    datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
            })
        return result
    except Exception as e:
        _sms_panel_logger.warning(f"_fetch_panel_last_messages error: {e}")
        return []


def _fetch_panel_last_message(panel: dict) -> dict | None:
    """Fetch the single latest SMS (used by Group-forward feature).
    Returns one dict or {'error': ...} or None."""
    result = _fetch_panel_last_messages(panel, count=1)
    if isinstance(result, dict):          # error dict
        return result
    if result:
        return result[0]
    return None


def _build_last_msgs_text(panel_name: str, msgs: list, today_count: int, icon: str) -> str:
    """Build the formatted text for Last 3 Messages display."""
    numbers = ["1️⃣", "2️⃣", "3️⃣"]
    lines = [f"📨 <b>{panel_name} — Last {len(msgs)} Message{'s' if len(msgs) != 1 else ''}</b>\n"]
    for i, m in enumerate(msgs):
        prefix = numbers[i] if i < len(numbers) else f"{i+1}."
        number  = m.get('number', '')
        message = m.get('message', '')
        country = m.get('country', '') or ''
        sender  = m.get('sender', '')
        otp     = ''
        try:
            otp = extract_otp_from_message(message) if message else ''
        except Exception:
            pass
        service = extract_service_from_message(message) if message else ''
        lines.append(f"{prefix}")
        lines.append(f"   Country: {country if country else 'Unknown'}")
        lines.append(f"   Number: <code>{number}</code>")
        if service:
            lines.append(f"   Service: {service}")
        elif sender:
            lines.append(f"   Sender: <code>{sender}</code>")
        if otp:
            lines.append(f"   OTP: <code>{otp}</code>")
        lines.append(f"   Message: <code>{message}</code>")
        lines.append("")
    lines.append(f"📊 Last 24h Total: <b>{today_count}</b>")
    lines.append(f"{icon} Status: {'Active' if icon == '🟢' else 'Inactive'}")
    return "\n".join(lines)


def _spread_cookies_to_both_domains(session: requests.Session):
    """
    Copy cookies from www.X to X and from X to www.X so that requests
    to either subdomain variant are authenticated.
    """
    new_cookies = []
    for cookie in list(session.cookies):
        domain = cookie.domain or ''
        if domain.startswith('www.'):
            bare = domain[4:]
            new_cookies.append((cookie.name, cookie.value, bare, cookie.path))
        elif domain and not domain.startswith('www.'):
            www_domain = 'www.' + domain
            new_cookies.append((cookie.name, cookie.value, www_domain, cookie.path))
    for name, value, domain, path in new_cookies:
        session.cookies.set(name, value, domain=domain, path=path or '/')


def _do_login(session: requests.Session, login_url: str,
              username: str, password: str) -> bool:
    """
    Attempt to login.  Returns True on success.
    Handles the math captcha automatically.
    Spreads cookies to both www and non-www variants after login.
    """
    from urllib.parse import urljoin, urlparse
    try:
        r = session.get(login_url, timeout=15)
        r.raise_for_status()
    except Exception as e:
        _sms_panel_logger.warning(f"Panel login GET failed ({login_url}): {e}")
        return False

    answer = _solve_math_captcha(r.text)

    # Build POST target — strip path from login_url and append form action
    soup = BeautifulSoup(r.text, 'html.parser')
    form = soup.find('form')
    if form:
        action = form.get('action', 'signin')
    else:
        action = 'signin'

    if action.startswith('http'):
        post_url = action
    else:
        post_url = urljoin(login_url, action)

    payload = {'username': username, 'password': password}
    if answer is not None:
        payload['capt'] = str(answer)

    try:
        r2 = session.post(post_url, data=payload, timeout=15, allow_redirects=False,
                          headers={'Referer': login_url,
                                   'Origin': urljoin(login_url, '/').rstrip('/'),
                                   'Content-Type': 'application/x-www-form-urlencoded'})
        # Must get a redirect to proceed
        if r2.status_code not in (301, 302, 303):
            return False

        # Resolve redirect Location — handle relative paths cleanly
        raw_loc = r2.headers.get('Location', '/')
        if raw_loc.startswith('http'):
            redirect_target = raw_loc
        else:
            # Remove leading "./" and resolve against the login base (not post_url)
            cleaned = raw_loc.lstrip('.').lstrip('/')
            base = urljoin(login_url, '/')   # e.g. http://smshadi.net/
            redirect_target = base + cleaned if cleaned else base

        # Remove accidental trailing-dot in hostname (e.g. http://smshadi.net./)
        redirect_target = re.sub(r'(https?://[^/]+)\./', r'\1/', redirect_target)

        try:
            r_final = session.get(redirect_target, timeout=15, allow_redirects=True)
            # Verify we are NOT back on the login page — that means auth failed
            final_url = getattr(r_final, 'url', redirect_target) or redirect_target
            if '/login' in final_url.lower():
                _sms_panel_logger.warning(
                    f"Login verification failed — landed on {final_url} (wrong credentials or IP restriction)"
                )
                return False
        except Exception:
            pass

        # Copy cookies to both www and non-www so either domain variant works
        _spread_cookies_to_both_domains(session)
        return True
    except Exception as e:
        _sms_panel_logger.warning(f"Panel login POST failed ({post_url}): {e}")
        return False


# ─── Message parsing ───────────────────────────────────────────────────────────

def _extract_messages(html: str, column_map: dict = None) -> list[dict]:
    """
    Parse the SMS stats page and return a list of:
        {id, number, message}
    If column_map is provided (admin-configured), uses exact column indices.
    Otherwise falls back to automatic header-based detection.
    column_map format:
        {
          "number":  {"name": "Phone",   "index": 2},  # 1-based
          "service": {"name": "Sender",  "index": 3},  # optional
          "message": {"name": "Message", "index": 5},
        }
    """
    soup = BeautifulSoup(html, 'html.parser')
    results = []

    for table in soup.find_all('table'):
        rows = table.find_all('tr')
        if not rows:
            continue

        # ── Column-map mode (admin-configured) ────────────────────────────────
        if column_map:
            num_idx = int(column_map.get('number', {}).get('index', 0)) - 1   # 0-based
            msg_idx = int(column_map.get('message', {}).get('index', 0)) - 1
            svc_entry = column_map.get('service')
            svc_idx = (int(svc_entry['index']) - 1) if svc_entry and svc_entry.get('index') else None

            data_rows = rows[1:] if len(rows) > 1 else rows
            for row in data_rows:
                cells = row.find_all(['td', 'th'])
                if not cells:
                    continue
                cell_texts = [c.get_text(strip=True) for c in cells]

                if num_idx < 0 or num_idx >= len(cell_texts):
                    continue
                number_raw = cell_texts[num_idx]
                if not number_raw:
                    continue
                number = re.sub(r'[\s\-().]', '', number_raw)

                message = cell_texts[msg_idx] if 0 <= msg_idx < len(cell_texts) else ''
                service = cell_texts[svc_idx] if svc_idx is not None and 0 <= svc_idx < len(cell_texts) else ''

                uid = hashlib.md5(f"{number}::{message}".encode()).hexdigest()[:16]
                entry = {'id': uid, 'number': number, 'message': message}
                if service:
                    entry['sender'] = service
                results.append(entry)

            if results:
                break  # column_map matched this table — no need to try others
            continue   # try next table if this one gave nothing

        # ── Auto-detect mode (no column_map) ──────────────────────────────────
        # Detect header row
        header_cells = rows[0].find_all(['th', 'td'])
        headers = [c.get_text(strip=True).lower() for c in header_cells]

        # Try to guess which column is the phone number and which is the message
        num_col = msg_col = None
        for i, h in enumerate(headers):
            if any(k in h for k in ('number', 'phone', 'mobile', 'from', 'to', 'caller', 'msisdn', 'src', 'destination')):
                if num_col is None:
                    num_col = i
            if any(k in h for k in ('message', 'text', 'sms', 'body', 'content', 'otp', 'msg')):
                if msg_col is None:
                    msg_col = i

        for row_idx, row in enumerate(rows[1:], start=1):
            cells = row.find_all(['td', 'th'])
            if not cells:
                continue
            cell_texts = [c.get_text(strip=True) for c in cells]

            # Find the phone number in this row
            number = None
            message = None

            if num_col is not None and num_col < len(cell_texts):
                candidate = cell_texts[num_col]
                if _PHONE_RE.search(candidate):
                    number = re.sub(r'[\s\-().]', '', candidate)
            if number is None:
                # Scan all cells for a phone-like value
                for ct in cell_texts:
                    m = _PHONE_RE.search(ct)
                    if m:
                        number = re.sub(r'[\s\-().]', '', m.group())
                        break

            if number is None:
                continue

            # Find message text
            if msg_col is not None and msg_col < len(cell_texts):
                message = cell_texts[msg_col]
            if not message:
                # Use any non-number cell that's non-empty
                for i, ct in enumerate(cell_texts):
                    if ct and not _PHONE_RE.search(ct) and len(ct) > 1:
                        message = ct
                        break
            if not message:
                message = " | ".join(cell_texts)

            # Stable unique ID for this record
            uid = hashlib.md5(f"{number}::{message}".encode()).hexdigest()[:16]
            results.append({'id': uid, 'number': number, 'message': message})

    # Fallback: look for phone-like patterns anywhere in the page body
    if not results:
        for m in _PHONE_RE.finditer(soup.get_text()):
            number = re.sub(r'[\s\-().]', '', m.group())
            uid = hashlib.md5(number.encode()).hexdigest()[:16]
            results.append({'id': uid, 'number': number, 'message': ''})

    return results


# ─── Per-panel polling loop ────────────────────────────────────────────────────

async def _poll_panel(bot, panel: dict, notify_chat_id: int = None):
    """
    Runs forever for one panel.
    Logs in, then fetches the message page every 3 seconds.
    Re-logs in automatically when the session expires.
    If notify_chat_id is set, sends a Telegram notification on first successful login.
    """
    from urllib.parse import urlparse, urlunparse
    panel_id = panel['id']
    name = panel['name']
    login_url = panel['login_url']
    username = panel['username']
    password = panel['password']

    # Normalize message_url to use the same host as login_url so session cookies match
    _login_parsed = urlparse(login_url)
    _msg_parsed   = urlparse(panel['message_url'])
    if _msg_parsed.netloc.replace('www.','') == _login_parsed.netloc.replace('www.',''):
        # Same base domain — reuse login host exactly
        message_url = urlunparse(_msg_parsed._replace(
            scheme=_login_parsed.scheme,
            netloc=_login_parsed.netloc
        ))
    else:
        message_url = panel['message_url']
    _sms_panel_logger.info(f"[Panel:{name}] Effective message URL: {message_url}")

    session = requests.Session()
    session.headers.update({'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'})

    logged_in = False
    _first_login_notified = False

    _sms_panel_logger.info(f"[Panel:{name}] Starting monitor → {message_url}")

    while True:
        try:
            # Check if panel is still enabled
            current = await asyncio.to_thread(get_sms_panel, panel_id)
            if not current or not current.get('enabled', True):
                _sms_panel_logger.info(f"[Panel:{name}] Disabled — stopping.")
                break

            # Login if needed
            if not logged_in:
                ok = await asyncio.to_thread(_do_login, session, login_url, username, password)
                if not ok:
                    _sms_panel_logger.warning(f"[Panel:{name}] Login failed — notifying admins and stopping monitor.")
                    admin_ids = await asyncio.to_thread(get_all_admin_ids)
                    for aid in admin_ids:
                        try:
                            await bot.send_message(
                                chat_id=aid,
                                text=(
                                    f"⚠️ <b>Panel Login ব্যর্থ হয়েছে!</b>\n\n"
                                    f"📡 Panel: <b>{name}</b>\n"
                                    f"🔗 URL: <code>{login_url}</code>\n\n"
                                    f"❌ Login করা সম্ভব হয়নি।\n"
                                    f"সম্ভাব্য কারণ:\n"
                                    f"• Username/Password ভুল\n"
                                    f"• সার্ভার এই IP থেকে login allow করছে না\n\n"
                                    f"⚙️ অনুগ্রহ করে প্যানেলটি ম্যানুয়ালি চেক করুন।"
                                ),
                                parse_mode='HTML'
                            )
                        except Exception:
                            pass
                    break  # Stop monitoring — bot will NOT retry on its own
                logged_in = True
                _panel_sessions[panel_id] = session
                _sms_panel_logger.info(f"[Panel:{name}] Logged in ✅")

                # Immediate login success notification → all admins
                if notify_chat_id and not _first_login_notified:
                    _first_login_notified = True
                    admin_ids = await asyncio.to_thread(get_all_admin_ids)
                    notified_set = set()
                    success_text = (
                        f"✅ <b>Panel Login সফল!</b>\n\n"
                        f"📡 Panel: <b>{name}</b>\n"
                        f"🔗 Login URL: <code>{login_url}</code>\n\n"
                        f"✔️ Login সফলভাবে সম্পন্ন হয়েছে।\n"
                        f"<i>এখন থেকে প্রতি {int(panel.get('poll_interval', 3))} সেকেন্ডে message চেক হবে।</i>"
                    )
                    for aid in admin_ids:
                        notified_set.add(aid)
                        try:
                            await bot.send_message(chat_id=aid, text=success_text, parse_mode='HTML')
                        except Exception:
                            pass
                    if notify_chat_id not in notified_set:
                        try:
                            await bot.send_message(chat_id=notify_chat_id, text=success_text, parse_mode='HTML')
                        except Exception:
                            pass

            # Fetch message page
            try:
                resp = await asyncio.to_thread(
                    lambda: session.get(message_url, timeout=15, allow_redirects=True)
                )
            except Exception as e:
                _sms_panel_logger.warning(f"[Panel:{name}] Fetch error: {e}")
                await asyncio.sleep(5)
                continue

            # Session expired → notify all admins and stop monitoring (no re-login)
            def _norm_url(u):
                return u.replace('://www.', '://').rstrip('/')
            final_url = getattr(resp, 'url', '') or ''
            if _norm_url(final_url) != _norm_url(message_url) and 'login' in final_url.lower():
                _sms_panel_logger.warning(f"[Panel:{name}] Session expired (redirected to {final_url}). Notifying admins and stopping.")
                _panel_sessions.pop(panel_id, None)  # Remove stale session so Last Message shows correct error
                admin_ids = await asyncio.to_thread(get_all_admin_ids)
                for aid in admin_ids:
                    try:
                        await bot.send_message(
                            chat_id=aid,
                            text=(
                                f"⚠️ <b>Panel Session Expire হয়েছে!</b>\n\n"
                                f"📡 Panel: <b>{name}</b>\n"
                                f"🔗 URL: <code>{login_url}</code>\n\n"
                                f"❌ Session expire হয়ে গেছে, monitoring বন্ধ হয়েছে।\n\n"
                                f"⚙️ Panel টি পুনরায় চালু করতে বটে গিয়ে panel বন্ধ করে আবার চালু করুন।"
                            ),
                            parse_mode='HTML'
                        )
                    except Exception:
                        pass
                break  # Stop monitoring

            _sms_panel_logger.debug(f"[Panel:{name}] Fetched page OK ({len(resp.text)} chars), url={final_url}")

            # Parse messages — CDR panels use JSON API, others use HTML parsing
            current_panel = await asyncio.to_thread(get_sms_panel, panel_id)
            if current_panel and _is_cdr_panel(current_panel):
                messages = await asyncio.to_thread(_fetch_cdr_sms, session, current_panel)
                if messages is None:
                    # Session expired detected via CDR API redirect — clean up and stop
                    _sms_panel_logger.warning(f"[Panel:{name}] Session expired (CDR API redirect). Stopping monitor.")
                    _panel_sessions.pop(panel_id, None)
                    break
                messages = messages or []
            else:
                col_map = (current_panel or {}).get('column_map') or None
                messages = await asyncio.to_thread(_extract_messages, resp.text, col_map)
            if messages:
                _sms_panel_logger.debug(f"[Panel:{name}] Parsed {len(messages)} message(s) from page")

            for msg_data in messages:
                sms_id = msg_data['id']
                number = msg_data['number'].lstrip('+')
                message = msg_data.get('message', '')
                country = msg_data.get('country', '')

                # Skip already seen
                if await asyncio.to_thread(is_sms_seen, panel_id, sms_id):
                    continue

                await asyncio.to_thread(mark_sms_seen, panel_id, sms_id)
                date_str = msg_data.get('date', '')
                sender_str = msg_data.get('cli', '') or msg_data.get('sender', '')

                # Resolve country & OTP once — used by both user send and group forward
                if country:
                    resolved_country = country
                else:
                    resolved_country = await asyncio.to_thread(get_country_name_by_number, number)
                resolved_otp     = await asyncio.to_thread(extract_otp_from_message, message)
                resolved_service = extract_service_from_message(message)
                resolved_sender  = msg_data.get('cli', '') or msg_data.get('sender', '')
                otp_line         = f"OTP: <code>{resolved_otp}</code>\n" if resolved_otp else ""
                if resolved_service:
                    service_line = f"Service: {resolved_service}\n"
                elif resolved_sender:
                    service_line = f"Sender: <code>{resolved_sender}</code>\n"
                else:
                    service_line = ""

                # ── 1. Auto-forward to group/chat (always — regardless of user) ──
                global_fwd    = get_global_forward_chat()
                per_panel_fwd = await asyncio.to_thread(get_panel_forward_chat, panel_id)
                fwd_targets   = list({c for c in [global_fwd, per_panel_fwd] if c})
                if fwd_targets:
                    _svc_raw = resolved_service or resolved_sender or ''
                    fwd_text, fwd_markup = _build_group_forward(
                        number, message, resolved_country,
                        resolved_otp, resolved_service, resolved_sender
                    )
                    _group_forwarded = False
                    for fwd_chat_id in fwd_targets:
                        try:
                            await bot.send_message(
                                chat_id=fwd_chat_id,
                                text=fwd_text,
                                parse_mode='HTML',
                                reply_markup=fwd_markup
                            )
                            _sms_panel_logger.info(f"[Panel:{name}] Auto-forwarded to chat {fwd_chat_id}")
                            _group_forwarded = True
                        except Exception as fe:
                            _sms_panel_logger.warning(f"[Panel:{name}] Auto-forward failed ({fwd_chat_id}): {fe}")
                    if _group_forwarded:
                        await asyncio.to_thread(increment_panel_group_count, panel_id)

                # ── 2. Send to the user who holds this number (if any) ──
                user_id = await asyncio.to_thread(get_user_id_by_number, number)
                if user_id is None:
                    user_id = await asyncio.to_thread(get_user_id_by_number, '+' + number)

                if user_id is None:
                    _sms_panel_logger.debug(f"[Panel:{name}] No user for number {number} — group-only forward done")
                    continue

                try:
                    # ── Balance reward for this OTP ──
                    reward_amount = await asyncio.to_thread(get_country_otp_reward, resolved_country)
                    old_bal_data  = await asyncio.to_thread(get_user_balance_data, user_id)
                    old_bal       = old_bal_data.get("balance", 0)
                    if reward_amount > 0:
                        await asyncio.to_thread(_add_balance_db, user_id, reward_amount)
                        new_bal = old_bal + reward_amount
                        bal_line = f"\n💰 +{reward_amount}৳ »»»–»»» {new_bal}৳"
                    else:
                        bal_line = ""

                    # ── Bot username (cached) ──
                    bot_username = ""
                    try:
                        bot_info = await bot.get_me()
                        bot_username = f"@{bot_info.username}" if bot_info.username else ""
                    except Exception:
                        pass

                    # ── Service or Sender line ──
                    if resolved_service:
                        svc_display = resolved_service
                    elif resolved_sender:
                        svc_display = resolved_sender
                    else:
                        svc_display = "Unknown"

                    otp_display = resolved_otp if resolved_otp else "—"

                    user_text = (
                        f"🔔 Service: {svc_display}\n\n"
                        f"☎️ NUMBER: +{number}\n\n"
                        f"🔑 OTP : <code>{otp_display}</code>\n\n"
                        f"💬 Full Message :\n{message}"
                        f"{bal_line}\n\n"
                        f" Thanks For using : {bot_username}"
                    )
                    await bot.send_message(chat_id=user_id, text=user_text, parse_mode='HTML')
                    await asyncio.to_thread(increment_panel_user_count, panel_id)
                    _sms_panel_logger.info(f"[Panel:{name}] Sent SMS to user {user_id} for number +{number}")
                except Exception as e:
                    _sms_panel_logger.warning(f"[Panel:{name}] Failed to send to user {user_id}: {e}")


        except asyncio.CancelledError:
            _sms_panel_logger.info(f"[Panel:{name}] Polling cancelled.")
            break
        except Exception as e:
            _sms_panel_logger.error(f"[Panel:{name}] Unexpected error: {e}")

        interval_sec = await asyncio.to_thread(get_sms_panel_interval, panel_id)
        await asyncio.sleep(max(3, interval_sec))


# ─── Public API ────────────────────────────────────────────────────────────────

async def start_all_panels(bot):
    """Start monitoring tasks for all enabled panels. Called at bot startup."""

    panels = await asyncio.to_thread(get_sms_panels)
    for panel in panels:
        if panel.get('enabled', True):
            start_panel_monitor(bot, panel)
    if panels:
        _sms_panel_logger.info(f"[PanelChecker] Started {len(panels)} panel monitor(s).")


def start_panel_monitor(bot, panel: dict, notify_chat_id: int = None):
    """Start (or restart) the polling task for a single panel.
    If notify_chat_id is given, a Telegram message is sent on the first successful login."""
    panel_id = panel['id']
    existing = _active_tasks.get(panel_id)
    if existing and not existing.done():
        existing.cancel()
    task = asyncio.create_task(_poll_panel(bot, panel, notify_chat_id=notify_chat_id))
    _active_tasks[panel_id] = task
    _sms_panel_logger.info(f"[PanelChecker] Monitor started for panel '{panel['name']}' ({panel_id})")


def stop_panel_monitor(panel_id: str):
    """Stop the polling task for a panel."""
    task = _active_tasks.pop(panel_id, None)
    if task and not task.done():
        task.cancel()
        _sms_panel_logger.info(f"[PanelChecker] Monitor stopped for panel {panel_id}")

# ════════════════════════════════════════════════════════════════════════
# DB INIT
# ════════════════════════════════════════════════════════════════════════
db_init(
    otp_link=OTP_GROUP_LINK,
    channel_one_link=CHANNEL_ONE_LINK,
    channel_two_link=CHANNEL_TWO_LINK,
    protected_admin_ids=PROTECTED_ADMIN_IDS,
)

# Set up logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO,
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Admin keyboard
def get_admin_keyboard():
    keyboard = [
        [KeyboardButton("🚩 Country Manager"), KeyboardButton("📬 Message Hub")],
        [KeyboardButton("👑 Admin Manager"), KeyboardButton("⚙️ Settings")],
        [KeyboardButton("📡 SMS Panels"), KeyboardButton("🤖 Bot Status")],
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


def get_sms_panels_keyboard():
    panels = get_sms_panels()
    keyboard = [
        [KeyboardButton("➕ Add SMS Panel"), KeyboardButton("📤 Group OTP Forward")],
        [KeyboardButton("📊 Panel Statistics"), KeyboardButton("🔙 Back to Admin")],
    ]
    row = []
    for p in panels:
        icon = "✅" if p.get("enabled", True) else "❌"
        row.append(KeyboardButton(f"📡 {icon} {p['name']}"))
        if len(row) == 2:
            keyboard.insert(-1, row)
            row = []
    if row:
        keyboard.insert(-1, row)
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


def get_panel_detail_keyboard(panel: dict) -> ReplyKeyboardMarkup:
    """Mobile keyboard shown after clicking a specific panel button."""
    enabled = panel.get("enabled", True)
    toggle_btn = "🔴 বন্ধ করুন" if enabled else "🟢 চালু করুন"
    interval = int(panel.get("poll_interval", 3))
    keyboard = [
        [KeyboardButton(toggle_btn), KeyboardButton("📨 Last Message")],
        [KeyboardButton("📤 Group এ পাঠান"), KeyboardButton(f"⏱ Interval ({interval}s)")],
        [KeyboardButton("✏️ Username পরিবর্তন"), KeyboardButton("🔑 Password পরিবর্তন")],
        [KeyboardButton("🗑️ Delete Panel"), KeyboardButton("🔙 Back to Panels")],
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

# Admin Manager reply keyboard (mobile keyboard)
def get_admin_manager_keyboard():
    keyboard = [
        [KeyboardButton("➕ Add Admin"), KeyboardButton("👥 User Activity")],
        [KeyboardButton("🔙 Back to Admin")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

# Settings reply keyboard (mobile keyboard)
def get_settings_keyboard():
    keyboard = [
        [KeyboardButton("🎁 Referral Settings"), KeyboardButton("💳 Withdraw Settings")],
        [KeyboardButton("🔢 Numbers Per Request"), KeyboardButton("🔗 Link Settings")],
        [KeyboardButton("📝 Custom Message"), KeyboardButton("🔙 Back to Admin")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


def get_custom_message_keyboard():
    keyboard = [
        [KeyboardButton("✏️ Set Message"), KeyboardButton("🗑 Remove Message")],
        [KeyboardButton("🔙 Back to Settings")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


def get_link_settings_keyboard():
    keyboard = [
        [KeyboardButton("📲 Set OTP Link"), KeyboardButton("❌ Remove OTP Link")],
        [KeyboardButton("➕ Add Channel"), KeyboardButton("➖ Remove Channel")],
        [KeyboardButton("📤 Group Button Links"), KeyboardButton("⏱ Check Interval")],
        [KeyboardButton("🔙 Back to Settings")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

def get_withdraw_settings_keyboard():
    keyboard = [
        [KeyboardButton("🔛 Toggle Withdraw"), KeyboardButton("💰 Set Min Amount")],
        [KeyboardButton("📨 Set Group ID"), KeyboardButton("📊 Withdraw Stats")],
        [KeyboardButton("🔙 Back to Settings")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

# Referral Manager reply keyboard (mobile keyboard)
def get_referral_manager_keyboard():
    keyboard = [
        [KeyboardButton("🔛 Toggle Referral"), KeyboardButton("🎁 Set Reward")],
        [KeyboardButton("👤 Check Balance"), KeyboardButton("➕ Add Balance")],
        [KeyboardButton("➖ Remove Balance"), KeyboardButton("🔙 Back to Settings")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

# Country Manager reply keyboard (mobile keyboard)
def get_country_manager_keyboard():
    keyboard = [
        [KeyboardButton("➕ Add Numbers"), KeyboardButton("🌍 Add Country")],
        [KeyboardButton("📱 Add Service"), KeyboardButton("🗺 Service Map")],
        [KeyboardButton("💰 OTP Rewards"), KeyboardButton("🔄 Reset Number")],
        [KeyboardButton("🔙 Back to Admin")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

# User keyboard
def get_user_keyboard():
    keyboard = [
        [KeyboardButton("🔢 Get Numbers"), KeyboardButton("🌍 Available Country")],
        [KeyboardButton("💰 My Balance"), KeyboardButton("📢 Notice")]
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

# Welcome keyboard with channel options (dynamic from DB)
def get_welcome_keyboard():
    channels = get_join_channels()
    keyboard = []
    for cid, title, link in channels:
        keyboard.append([InlineKeyboardButton(title, url=link)])
    keyboard.append([InlineKeyboardButton("✅ Verify", callback_data="verify_user")])
    return InlineKeyboardMarkup(keyboard)

# Main menu keyboard after verification
def get_main_menu_keyboard():
    keyboard = [
        [InlineKeyboardButton("✅ Get Numbers", callback_data="get_numbers")]
    ]
    return InlineKeyboardMarkup(keyboard)

async def cancel_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /cancel — abort any ongoing setup flow and restore admin keyboard."""
    username = update.effective_user.username
    user_id = update.effective_user.id
    panel_setup_keys = [
        'awaiting_panel_name', 'awaiting_panel_login_url', 'awaiting_panel_message_url',
        'awaiting_panel_username', 'awaiting_panel_password',
        'awaiting_panel_num_col_name', 'awaiting_panel_num_col_idx',
        'awaiting_panel_svc_col_name', 'awaiting_panel_svc_col_idx',
        'awaiting_panel_msg_col_name', 'awaiting_panel_msg_col_idx',
        'panel_setup_name', 'panel_setup_login_url', 'panel_setup_message_url', 'panel_setup_username',
        'panel_setup_password', 'panel_setup_num_col_name', 'panel_setup_num_col_idx',
        'panel_setup_svc_col_name', 'panel_setup_svc_col_idx', 'panel_setup_msg_col_name',
        'awaiting_panel_edit_username', 'awaiting_panel_edit_password',
        'awaiting_manual_forward_panel', 'awaiting_manual_forward_step', 'manual_forward_number',
        'awaiting_panel_interval', 'awaiting_forward_chat_id', 'awaiting_global_forward_chat_id',
    ]
    was_in_setup = any(context.user_data.get(k) for k in panel_setup_keys)
    for k in panel_setup_keys:
        context.user_data.pop(k, None)
    if was_in_setup:
        if is_admin(username, user_id):
            await update.message.reply_text(
                f"{_sx('❌')} <b>Panel Setup বাতিল হয়েছে।</b>",
                parse_mode='HTML',
                reply_markup=get_sms_panels_keyboard()
            )
        return
    clear_awaiting_states(context)
    if is_admin(username, user_id):
        await update.message.reply_text(
            f"{_sx('✅')} সব pending input বাতিল হয়েছে।",
            parse_mode='HTML',
            reply_markup=get_admin_keyboard()
        )

# Admin commands
async def admin_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    username = update.effective_user.username
    user_id = update.effective_user.id
    if not is_admin(username, user_id):
        await update.message.reply_text(f"{_sx('❌')} You are not authorized to use admin commands.", parse_mode='HTML')
        return
    
    await update.message.reply_text(
        f"{_sx('🤖')} <b>Admin Panel</b>\n\n"
        f"{_sx('🚩')} Country Manager — Countries &amp; numbers\n"
        f"{_sx('📣')} Broadcast — Message all users\n"
        f"{_sx('👑')} Admin Manager — Manage admins\n"
        f"{_sx('🔑')} Settings — Configure bot\n"
        f"{_sx('🤖')} Bot Status — Statistics",
        parse_mode='HTML',
        reply_markup=get_admin_keyboard()
    )

async def enforce_join(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    """
    Check that the user has joined all required channels.
    Returns True if all channels are joined (or no channels configured).
    Returns False and sends a join-prompt if any channel is missing.
    Uses _ALREADY_VERIFIED for session-level bypass and _MEMBER_CACHE for TTL cache.
    """
    user_id = update.effective_user.id

    # Session-level fast path
    if user_id in _ALREADY_VERIFIED:
        return True

    # TTL cache fast path
    cached = _cache_get(user_id)
    if cached is not None and len(cached) == 0:
        _ALREADY_VERIFIED.add(user_id)
        return True

    channels = get_join_channels()
    if not channels:
        _ALREADY_VERIFIED.add(user_id)
        return True

    unjoined = []
    for cid, title, link in channels:
        try:
            chat_id = await _resolve_chat_id(context.bot, link)
            if chat_id is None:
                # Bot not in this channel/group — can't verify, treat as unjoined
                unjoined.append((cid, title, link))
                continue
            member = await context.bot.get_chat_member(chat_id, user_id)
            from telegram import ChatMemberStatus
            if member.status in (ChatMemberStatus.LEFT, ChatMemberStatus.BANNED):
                unjoined.append((cid, title, link))
        except Exception:
            continue

    _cache_set(user_id, unjoined)

    if not unjoined:
        _ALREADY_VERIFIED.add(user_id)
        return True

    # "Send once" tracking — don't re-send join prompt on every button press
    prompted: set = context.bot_data.setdefault('join_prompted', set())
    if user_id in prompted:
        return False  # already notified this cycle — block silently

    prompted.add(user_id)

    # Show join prompt
    lines = "\n".join([f"👉 <a href='{l}'>{t}</a>" for _, t, l in unjoined])
    msg = (
        "⚠️ <b>আপনি নিচের চ্যানেলে জয়েন নেই:</b>\n\n"
        f"{lines}\n\n"
        "জয়েন করার পর ✅ <b>Verify</b> বাটনে ক্লিক করুন।"
    )
    keyboard = [[InlineKeyboardButton(f"📢 {t}", url=l)] for _, t, l in unjoined]
    keyboard.append([InlineKeyboardButton("✅ Verify", callback_data="verify_user")])
    reply_markup = InlineKeyboardMarkup(keyboard)

    if update.callback_query:
        try:
            await update.callback_query.edit_message_text(
                msg, parse_mode='HTML',
                reply_markup=reply_markup,
                disable_web_page_preview=True
            )
        except Exception:
            await update.callback_query.message.reply_text(
                msg, parse_mode='HTML',
                reply_markup=reply_markup,
                disable_web_page_preview=True
            )
    else:
        await update.message.reply_text(
            msg, parse_mode='HTML',
            reply_markup=reply_markup,
            disable_web_page_preview=True
        )
    return False


# User commands
async def user_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Store user data for broadcast (non-blocking — fire and forget)
    user = update.effective_user
    asyncio.create_task(asyncio.to_thread(
        add_user, user.id, user.username, user.first_name, user.last_name
    ))
    
    # Check if user has joined all required channels
    if await enforce_join(update, context):
        await show_main_menu(update, context)

async def show_welcome_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    welcome_message = """
🤖 *Welcome to Number Bot!*

📢 *Join our channels first to use this bot*

Stay with us, I hope you can learn something good. Join the live regularly. Join all my channels and groups.

🧑‍💻 *Bot Owner:* ADMIN LIMON

*Instructions:*
1. Click on Channel One button to join first channel
2. Click on Channel Two button to join second channel
3. After joining all channels, click on ✅ Verify button
4. Then you can use the bot features
"""
    
    if update.callback_query:
        await update.callback_query.edit_message_text(
            welcome_message,
            parse_mode='Markdown',
            reply_markup=get_welcome_keyboard()
        )
    else:
        await update.message.reply_text(
            welcome_message,
            parse_mode='Markdown',
            reply_markup=get_welcome_keyboard()
        )

async def show_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    menu_message = (
        f"{_sx('🎉')} <b>Verification Successful!</b>\n\n"
        f"{_sx('🤖')} <b>Welcome to Number Bot</b>\n\n"
        "You can now use all features of the bot.\n\n"
        f"🔢 <b>Get Numbers</b> — Get phone numbers by country\n"
        f"🌍 <b>Available Country</b> — See available numbers\n"
        f"💰 <b>My Balance</b> — Check balance &amp; referrals\n"
        f"📢 <b>Notice</b> — Important announcements\n\n"
        "👇 Choose an option below:"
    )
    
    if update.callback_query:
        await update.callback_query.edit_message_text(
            menu_message,
            parse_mode='HTML'
        )
        await update.callback_query.message.reply_text(
            "👇 Choose an option below:",
            reply_markup=get_user_keyboard()
        )
    else:
        await update.message.reply_text(
            menu_message,
            parse_mode='HTML',
            reply_markup=get_user_keyboard()
        )

async def show_services(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show available services for the user to pick from."""
    if not await enforce_join(update, context):
        return

    services = get_services_with_emoji()
    if not services:
        msg = "❌ No services available at the moment."
        if update.callback_query:
            await update.callback_query.edit_message_text(msg)
        else:
            await update.message.reply_text(msg)
        return

    # Build message with animated emoji per service
    lines = "\n".join(
        f"{_svc_animated_tag(name, ceid)} <b>{name}</b>"
        for _, name, ceid in services
    )
    msg_text = f"🎯 <b>Select a Service:</b>"

    # Buttons: plain service name (buttons don't render HTML/animated emoji)
    keyboard = [[InlineKeyboardButton(name, callback_data=f"service_{sid}")] for sid, name, _ in services]
    reply_markup = InlineKeyboardMarkup(keyboard)

    if update.callback_query:
        await update.callback_query.edit_message_text(
            msg_text,
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
    else:
        await update.message.reply_text(
            msg_text,
            reply_markup=reply_markup,
            parse_mode='HTML'
        )

async def show_countries_for_service(query, service_id, service_name):
    """Show countries available under a service."""
    countries = await asyncio.to_thread(get_countries_by_service, service_id)
    if not countries:
        await query.edit_message_text(f"❌ No numbers available for <b>{service_name}</b>.", parse_mode='HTML')
        return

    # Fetch all counts concurrently
    counts = await asyncio.gather(*[
        asyncio.to_thread(get_numbers_count_by_country, cid)
        for cid, _ in countries
    ])

    keyboard = []
    for (country_id, country_name), (_, available) in zip(countries, counts):
        if available > 0:
            plain = _clean_name(country_name)
            flag = get_unicode_flag(plain)
            keyboard.append([InlineKeyboardButton(f"{flag} {plain} (+{available})", callback_data=f"country_{country_id}_{service_id}")])

    if not keyboard:
        await query.edit_message_text(f"❌ No numbers available for <b>{service_name}</b>.", parse_mode='HTML')
        return

    keyboard.append([InlineKeyboardButton("🔙 Back to Services", callback_data="get_numbers")])
    await query.edit_message_text(
        "<b>Select a country:</b>",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='HTML'
    )

async def show_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await enforce_join(update, context):
        return
    
    stats = get_country_stats()
    if not stats:
        message = "❌ No data available."
    else:
        message = "<b>Country Statistics</b>\n\n"
        for country, total, available in stats:
            plain = _clean_name(country)
            flag = get_animated_flag_html(plain)
            message += f"• {flag} <code>{plain}</code>: {available}/{total} available\n"
    
    keyboard = [[InlineKeyboardButton("🔙 Back to Main Menu", callback_data="back_to_main")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if update.callback_query:
        await update.callback_query.edit_message_text(
            message,
            parse_mode='HTML',
            reply_markup=reply_markup
        )
    else:
        await update.message.reply_text(
            message,
            parse_mode='HTML',
            reply_markup=reply_markup
        )

# Handle callback queries
async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    data = query.data
    user = query.from_user
    user_id = user.id
    username = user.username
    admin_check = lambda: is_admin(username, user_id)

    # Register user on any button click (in case they skipped /start)
    asyncio.create_task(asyncio.to_thread(
        add_user, user.id, user.username, user.first_name, user.last_name
    ))
    
    if data in ("broadcast_type_all", "broadcast_type_banned"):
        if not admin_check():
            await query.answer("❌ Unauthorized", show_alert=True)
            return
        target = "all" if data == "broadcast_type_all" else "banned"
        label  = "All Users" if target == "all" else "Banned Users"
        context.user_data['awaiting_broadcast'] = True
        context.user_data['broadcast_target']   = target
        await query.edit_message_text(
            f"{_sx('📢')} <b>Broadcast → {label}</b>\n\n"
            "Send any message to broadcast:\n"
            "• Text, Photo, Video, Audio, Document, Sticker, Voice\n"
            "• Or <b>forward</b> any message from a channel/user\n\n"
            "Forwarded messages will appear as forwarded to users.",
            parse_mode='HTML'
        )

    elif data == "broadcast_confirm":
        if not admin_check():
            await query.answer("❌ Unauthorized", show_alert=True)
            return
        msg_id = context.user_data.get('broadcast_msg_id')
        chat_id = context.user_data.get('broadcast_chat_id')
        is_fwd  = context.user_data.get('broadcast_is_forwarded', False)
        target  = context.user_data.get('broadcast_target', 'all')
        if not msg_id or not chat_id:
            await query.edit_message_text(f"{_sx('❌')} Broadcast data not found. Please try again.", parse_mode='HTML')
            return

        # Build user list based on target
        if target == "banned":
            raw_list = get_banned_users()
            target_label = "🚫 Banned Users"
        else:
            raw_list = get_all_users()
            target_label = "📣 All Users"

        user_ids = []
        for u in raw_list:
            uid = u.get('user_id') if isinstance(u, dict) else u
            if uid:
                try:
                    user_ids.append(int(uid))
                except (ValueError, TypeError):
                    pass

        total = len(user_ids)

        progress_msg = await query.edit_message_text(
            f"{_sx('📢')} <b>Broadcast শুরু হচ্ছে...</b>\n"
            f"🎯 Target: <b>{target_label}</b>\n\n"
            f"[░░░░░░░░░░░░░░] 0%\n"
            f"📤 Sent: <b>0 / {total}</b>",
            parse_mode='HTML'
        )

        success_count, fail_count = await _fast_broadcast(
            bot=context.bot,
            user_ids=user_ids,
            from_chat_id=chat_id,
            message_id=msg_id,
            is_fwd=is_fwd,
            progress_msg=progress_msg
        )

        context.user_data.pop('broadcast_msg_id', None)
        context.user_data.pop('broadcast_chat_id', None)
        context.user_data.pop('broadcast_is_forwarded', None)
        context.user_data.pop('broadcast_target', None)

        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=(
                f"{_sx('📊')} <b>Broadcast Completed!</b>\n"
                f"🎯 Target: <b>{target_label}</b>\n\n"
                f"✅ Success: <b>{success_count}</b>\n"
                f"❌ Failed: <b>{fail_count}</b>\n"
                f"👥 Total: <b>{total}</b>"
            ),
            parse_mode='HTML'
        )

    elif data == "broadcast_cancel":
        if not admin_check():
            await query.answer("❌ Unauthorized", show_alert=True)
            return
        context.user_data.pop('broadcast_msg_id', None)
        context.user_data.pop('broadcast_chat_id', None)
        context.user_data.pop('broadcast_is_forwarded', None)
        context.user_data.pop('broadcast_target', None)
        await query.edit_message_text(f"{_sx('❌')} <b>Broadcast বাতিল করা হয়েছে।</b>", parse_mode='HTML')

    elif data == "verify_user":
        # Real-time membership check on Verify button
        if await enforce_join(update, context):
            # Clear prompted flag so verified user isn't silently blocked later
            context.bot_data.setdefault('join_prompted', set()).discard(query.from_user.id)
            await show_main_menu(update, context)
    
    elif data == "get_numbers":
        await show_services(update, context)
    
    elif data == "view_stats":
        await show_stats(update, context)

    elif data == "view_countries":
        if not await enforce_join(update, context):
            return
        countries = await asyncio.to_thread(get_countries)
        if not countries:
            await query.edit_message_text(f"{_sx('❌')} No countries available.", parse_mode='HTML')
            return
        counts = await asyncio.gather(*[
            asyncio.to_thread(get_numbers_count_by_country, cid)
            for cid, _ in countries
        ])
        rewards = await asyncio.to_thread(get_all_country_otp_rewards)
        lines = []
        for (cid, name), (_, available) in zip(countries, counts):
            if available > 0:
                plain = _clean_name(name)
                flag = get_animated_flag_html(plain)
                bonus = rewards.get(plain.lower(), 0)
                bonus_text = f" | 💰 <b>{bonus}৳</b> bonus" if bonus else ""
                lines.append(f"{flag} <b>{plain}</b> — <b>{available}</b>{bonus_text}")
        if not lines:
            await query.edit_message_text(f"{_sx('❌')} No numbers available in any country.", parse_mode='HTML')
            return
        msg = "📋 <b>Available Countries</b>\n\n" + "\n".join(lines)
        back_kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_main")]])
        await query.edit_message_text(msg, parse_mode='HTML', reply_markup=back_kb)

    elif data == "my_balance":
        if await _notify_flagged_once(update, context):
            return
        await show_balance_panel(update, context)

    elif data == "view_notice":
        msg = await asyncio.to_thread(get_custom_message)
        back_kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back to Menu", callback_data="back_to_main")]])
        if not msg:
            await query.edit_message_text(
                f"{_sx('🔔')} <b>Notice</b>\n\n<i>এখনো কোনো notice সেট করা হয়নি।</i>",
                parse_mode='HTML',
                reply_markup=back_kb
            )
        else:
            await query.edit_message_text(
                f"{_sx('🔔')} <b>Notice</b>\n\n{msg}",
                parse_mode='HTML',
                reply_markup=back_kb
            )

    elif data == "back_to_main":
        await show_main_menu(update, context)
    
    elif data.startswith("service_"):
        if not await enforce_join(update, context):
            return
        service_id = int(data.split("_")[1])
        services = get_services()
        service_name = next((name for sid, name in services if sid == service_id), "Unknown")
        await show_countries_for_service(query, service_id, service_name)

    elif data.startswith("country_"):
        if not await enforce_join(update, context):
            return

        parts = data.split("_")
        country_id = int(parts[1])
        service_id = int(parts[2]) if len(parts) > 2 else None

        # Get country name + numbers concurrently
        countries, count = await asyncio.gather(
            asyncio.to_thread(get_countries),
            asyncio.to_thread(get_numbers_per_request),
        )
        country_name = next((name for cid, name in countries if cid == country_id), "Unknown")
        numbers = await asyncio.to_thread(get_available_numbers_by_country, country_id, count, query.from_user.id)
        if not numbers:
            await query.edit_message_text(f"{_sx('❌')} No numbers available for this country.", parse_mode='HTML')
            return
        back_cb = f"service_{service_id}" if service_id else "get_numbers"
        plain_name = _clean_name(country_name)
        animated_flag = get_animated_flag_html(plain_name)
        keyboard = []
        for num in numbers:
            keyboard.append([InlineKeyboardButton(f"+{num}", copy_text=CopyTextButton(text=f"+{num}"))])
        keyboard.append([InlineKeyboardButton("🔄 Change Number", callback_data=f"another_{country_id}_{service_id or ''}")])
        otp_link = get_otp_link()
        keyboard.append([InlineKeyboardButton("📲 GET OTP", url=otp_link)])
        keyboard.append([InlineKeyboardButton("🔙 Back to Countries", callback_data=back_cb)])
        await query.edit_message_text(
            f"{animated_flag} <b>{plain_name}</b>",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )

    elif data.startswith("another_"):
        if not await enforce_join(update, context):
            return

        parts = data.split("_")
        country_id = int(parts[1])
        service_id = int(parts[2]) if len(parts) > 2 and parts[2] else None

        # Get country name + count concurrently
        countries, count = await asyncio.gather(
            asyncio.to_thread(get_countries),
            asyncio.to_thread(get_numbers_per_request),
        )
        country_name = next((name for cid, name in countries if cid == country_id), "Unknown")
        numbers = await asyncio.to_thread(get_available_numbers_by_country, country_id, count, query.from_user.id)
        if not numbers:
            await query.edit_message_text(f"{_sx('❌')} No more numbers available for this country.", parse_mode='HTML')
            return
        back_cb = f"service_{service_id}" if service_id else "get_numbers"
        plain_name = _clean_name(country_name)
        animated_flag = get_animated_flag_html(plain_name)
        keyboard = []
        for num in numbers:
            keyboard.append([InlineKeyboardButton(f"+{num}", copy_text=CopyTextButton(text=f"+{num}"))])
        keyboard.append([InlineKeyboardButton("🔄 Change Number", callback_data=f"another_{country_id}_{service_id or ''}")])
        otp_link = get_otp_link()
        keyboard.append([InlineKeyboardButton("📲 GET OTP", url=otp_link)])
        keyboard.append([InlineKeyboardButton("🔙 Back to Countries", callback_data=back_cb)])
        await query.edit_message_text(
            f"{animated_flag} <b>{plain_name}</b>",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )

    # Admin callback handlers (existing functionality)
    elif data.startswith("delete_country_"):
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
            
        country_id = int(data.split("_")[2])
        
        # Get country name
        countries = get_countries()
        country_name = next((name for cid, name in countries if cid == country_id), "Unknown")
        
        # Create deletion options
        keyboard = [
            [InlineKeyboardButton("🗑️ Delete All Numbers", callback_data=f"delete_all_{country_id}")],
            [InlineKeyboardButton("🗑️ Delete Country", callback_data=f"delete_country_completely_{country_id}")],
            [InlineKeyboardButton("🔙 Back to Delete Menu", callback_data="back_to_delete")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.edit_message_text(
            f"*Delete Options for {country_name}*\n\n⚠️ *Warning:* This will permanently delete numbers or the entire country!",
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )
    
    elif data.startswith("delete_all_"):
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
            
        country_id = int(data.split("_")[2])
        
        # Get country name
        countries = get_countries()
        country_name = next((name for cid, name in countries if cid == country_id), "Unknown")
        
        # Delete all numbers from this country (but keep the country)
        deleted_count = delete_all_numbers_from_country(country_id)
        
        await query.edit_message_text(f"{_sx('✅')} Successfully deleted {deleted_count} numbers from {country_name}!", parse_mode='HTML')
    
    elif data.startswith("delete_country_completely_"):
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
            
        country_id = int(data.split("_")[3])
        
        # Get country name
        countries = get_countries()
        country_name = next((name for cid, name in countries if cid == country_id), "Unknown")
        
        # Delete the country completely (with all numbers)
        numbers_deleted, country_deleted = delete_country(country_id)
        
        if country_deleted:
            await query.edit_message_text(f"{_sx('✅')} Successfully deleted country '<b>{country_name}</b>' and {numbers_deleted} numbers!", parse_mode='HTML')
        else:
            await query.edit_message_text(f"{_sx('❌')} Failed to delete country '<b>{country_name}</b>'!", parse_mode='HTML')
    
    elif data.startswith("remove_admin_"):
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
            
        admin_to_remove = data.split("_")[2]
        
        # Remove admin
        success, message = remove_admin(admin_to_remove)
        
        if success:
            await query.edit_message_text(f"{_sx('✅')} Admin @{admin_to_remove} has been removed successfully!", parse_mode='HTML')
        else:
            await query.edit_message_text(f"{_sx('❌')} {message}", parse_mode='HTML')
    
    elif data.startswith("protected_admin_"):
        protected_admin = data.split("_")[2]
        await query.answer(f"🛡️ {protected_admin} is a protected admin and cannot be removed!", show_alert=True)
    
    elif data == "back_to_delete":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
            
        # Show countries list for deletion
        countries = get_countries()
        if not countries:
            await query.edit_message_text(f"{_sx('❌')} No countries available.", parse_mode='HTML')
            return
        
        keyboard = []
        country_lines = []
        for country_id, country_name in countries:
            total, available = get_numbers_count_by_country(country_id)
            plain = _clean_name(country_name)
            country_lines.append(f"{get_animated_flag_html(plain)} <b>{plain}</b> ({total})")
            keyboard.append([InlineKeyboardButton(f"{plain} ({total})", callback_data=f"delete_country_{country_id}")])
        
        keyboard.append([InlineKeyboardButton("🔙 Back to Admin Panel", callback_data="back_to_admin")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        countries_text = "\n".join(country_lines)
        await query.edit_message_text(
            f"<b>Delete Numbers/Countries</b>\n\nSelect a country to delete numbers from or delete the country completely:\n\n{countries_text}",
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
    
    elif data == "back_to_admin":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
            
        await query.edit_message_text(
            f"{_sx('🤖')} <b>Admin Panel</b>",
            parse_mode='HTML'
        )
        await query.message.reply_text(
            f"{_sx('🤖')} <b>Admin Panel</b>",
            parse_mode='HTML',
            reply_markup=get_admin_keyboard()
        )

    elif data == "gfwd_set":
        if not admin_check():
            await query.answer("❌ Unauthorized.", show_alert=True)
            return
        context.user_data['awaiting_global_forward_chat_id'] = True
        await query.message.reply_text(
            f"📤 <b>Group OTP Forward — Group Add</b>\n\n"
            f"যে Group বা Channel-এ OTP পাঠাতে চান সেটির <b>Chat ID</b> দিন।\n\n"
            f"<i>উদাহরণ: -1001234567890</i>\n\n"
            f"💡 Chat ID পেতে @userinfobot বা @RawDataBot ব্যবহার করুন।\n\n"
            f"<i>বাতিল করতে /cancel লিখুন।</i>",
            parse_mode='HTML'
        )

    elif data == "gfwd_remove":
        if not admin_check():
            await query.answer("❌ Unauthorized.", show_alert=True)
            return
        remove_global_forward_chat()
        try:
            await query.edit_message_text(
                f"📤 <b>Group OTP Forward</b>\n\n"
                f"✅ Group Remove করা হয়েছে।\n\n"
                f"এখনো কোনো group/channel সেট করা হয়নি।",
                parse_mode='HTML'
            )
        except Exception:
            pass
        await query.answer("✅ Group Remove করা হয়েছে!", show_alert=True)
        reply_kb = ReplyKeyboardMarkup(
            [
                [KeyboardButton("➕ Group Add"), KeyboardButton("🗑 Group Remove")],
                [KeyboardButton("🔙 Back to Panels")],
            ],
            resize_keyboard=True
        )
        await query.message.reply_text(
            f"📤 <b>Group OTP Forward</b>\n\n"
            f"এখনো কোনো group/channel সেট করা হয়নি।\n\n"
            f"নিচের বাটন থেকে একটি Group বা Channel সেট করুন।",
            parse_mode='HTML',
            reply_markup=reply_kb
        )

    elif data.startswith("fwd_set_"):
        if not admin_check():
            await query.answer("❌ Unauthorized.", show_alert=True)
            return
        panel_id = data[len("fwd_set_"):]
        panel = get_sms_panel(panel_id)
        if not panel:
            await query.answer("❌ Panel পাওয়া যায়নি।", show_alert=True)
            return
        context.user_data['awaiting_forward_chat_id'] = panel_id
        await query.answer()
        await query.message.reply_text(
            f"📤 <b>Auto Forward — {panel['name']}</b>\n\n"
            f"Channel বা Group এর <b>Chat ID</b> দিন।\n\n"
            f"<i>উদাহরণ: -1001234567890</i>\n\n"
            f"💡 Chat ID পেতে @userinfobot বা @RawDataBot ব্যবহার করুন।\n\n"
            f"<i>বাতিল করতে /cancel লিখুন।</i>",
            parse_mode='HTML'
        )

    elif data.startswith("fwd_remove_"):
        if not admin_check():
            await query.answer("❌ Unauthorized.", show_alert=True)
            return
        panel_id = data[len("fwd_remove_"):]
        panel = get_sms_panel(panel_id)
        if not panel:
            await query.answer("❌ Panel পাওয়া যায়নি।", show_alert=True)
            return
        remove_panel_forward_chat(panel_id)
        await query.answer("✅ Auto Forward বন্ধ করা হয়েছে!", show_alert=False)
        updated_panel = get_sms_panel(panel_id)
        try:
            await query.edit_message_text(
                f"📤 <b>Auto Forward — {panel['name']}</b>\n\n"
                f"<i>এখনো কোনো chat ID সেট করা হয়নি।</i>\n\n"
                f"নতুন SMS আসলে স্বয়ংক্রিয়ভাবে নির্দিষ্ট channel বা group-এ পাঠাবে।",
                parse_mode='HTML',
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("✅ Set Channel/Group", callback_data=f"fwd_set_{panel_id}")],
                    [InlineKeyboardButton("🗑 Remove Forward", callback_data=f"fwd_remove_{panel_id}")],
                ])
            )
        except Exception:
            pass
        if updated_panel:
            await query.message.reply_text(
                f"✅ <b>Auto Forward বন্ধ করা হয়েছে।</b>",
                parse_mode='HTML',
                reply_markup=get_panel_detail_keyboard(updated_panel)
            )

    elif data.startswith("sms_panel_toggle_"):
        if not admin_check():
            await query.answer("❌ Unauthorized.", show_alert=True)
            return
        panel_id = data[len("sms_panel_toggle_"):]
        new_state = toggle_sms_panel(panel_id)
        if new_state is None:
            await query.answer("❌ Panel পাওয়া যায়নি।", show_alert=True)
            return
        panel = get_sms_panel(panel_id)
        if panel:
            try:
                if new_state:
                    start_panel_monitor(context.bot, panel)
                else:
                    stop_panel_monitor(panel_id)
            except Exception as e:
                logger.error(f"Panel toggle error: {e}")
        icon = "🟢 চালু" if new_state else "🔴 বন্ধ"
        await query.answer(f"Panel {icon} করা হয়েছে!", show_alert=False)
        keyboard = [
            [InlineKeyboardButton(
                "🔴 বন্ধ করুন" if new_state else "🟢 চালু করুন",
                callback_data=f"sms_panel_toggle_{panel_id}"
            )],
            [InlineKeyboardButton("📨 Last Message", callback_data=f"sms_panel_last_msg_{panel_id}")],
            [InlineKeyboardButton("🗑️ Delete Panel", callback_data=f"sms_panel_delete_{panel_id}")],
        ]
        try:
            await query.edit_message_reply_markup(reply_markup=InlineKeyboardMarkup(keyboard))
        except Exception:
            pass

    elif data.startswith("sms_panel_delete_"):
        if not admin_check():
            await query.answer("❌ Unauthorized.", show_alert=True)
            return
        panel_id = data[len("sms_panel_delete_"):]
        panel = get_sms_panel(panel_id)
        name = panel['name'] if panel else panel_id
        delete_sms_panel(panel_id)
        try:
            stop_panel_monitor(panel_id)
        except Exception:
            pass
        await query.edit_message_text(
            f"{_sx('✅')} Panel <b>{name}</b> মুছে ফেলা হয়েছে।",
            parse_mode='HTML',
            reply_markup=None
        )
        await query.message.reply_text(
            f"{_sx('📡')} <b>SMS Panels</b>",
            parse_mode='HTML',
            reply_markup=get_sms_panels_keyboard()
        )

    elif data.startswith("sms_panel_last_msg_"):
        if not admin_check():
            await query.answer("❌ Unauthorized.", show_alert=True)
            return
        panel_id = data[len("sms_panel_last_msg_"):]
        panel = get_sms_panel(panel_id)
        if not panel:
            await query.answer("❌ Panel পাওয়া যায়নি।", show_alert=True)
            return
        await query.answer("⏳ চেক করা হচ্ছে...", show_alert=False)
        checking_msg = await query.message.reply_text(
            f"⏳ <b>{panel['name']}</b> — panel এ চেক করা হচ্ছে...",
            parse_mode='HTML'
        )
        msgs = await asyncio.to_thread(_fetch_panel_last_messages, panel, 3)
        icon = "🟢" if panel.get("enabled", True) else "🔴"
        today_count = await asyncio.to_thread(get_today_panel_message_count, panel_id)
        try:
            await checking_msg.delete()
        except Exception:
            pass
        if isinstance(msgs, dict) and msgs.get('error'):
            err = msgs['error']
            if err == 'login_failed':
                err_detail = "Panel এ login করা হয়নি। Panel চালু করুন।"
            else:
                err_detail = "Session expire হয়ে গেছে। Panel বন্ধ করে আবার চালু করুন।"
            msg_text = (
                f"❌ <b>{panel['name']} — Login হয়নি</b>\n\n"
                f"{err_detail}\n\n"
                f"📊 Last 24h Total: <b>{today_count}</b>\n"
                f"{icon} Status: {'Active' if panel.get('enabled', True) else 'Inactive'}"
            )
        elif msgs:
            msg_text = _build_last_msgs_text(panel['name'], msgs, today_count, icon)
        else:
            msg_text = (
                f"📨 <b>{panel['name']} — Last Messages</b>\n\n"
                f"<i>No message received yet.</i>\n\n"
                f"📊 Last 24h Total: <b>{today_count}</b>\n"
                f"{icon} Status: {'Active' if panel.get('enabled', True) else 'Inactive'}"
            )
        refresh_markup = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 Refresh", callback_data=f"sms_panel_refresh_msg_{panel_id}")]
        ])
        try:
            await query.message.reply_text(msg_text, parse_mode='HTML', reply_markup=refresh_markup)
        except Exception as e:
            logger.error(f"sms_panel_last_msg reply error: {e}")

    elif data.startswith("sms_panel_refresh_msg_"):
        if not admin_check():
            await query.answer("❌ Unauthorized.", show_alert=True)
            return
        panel_id = data[len("sms_panel_refresh_msg_"):]
        panel = get_sms_panel(panel_id)
        if not panel:
            await query.answer("❌ Panel পাওয়া যায়নি।", show_alert=True)
            return
        await query.answer("⏳ Refreshing...", show_alert=False)
        msgs = await asyncio.to_thread(_fetch_panel_last_messages, panel, 3)
        icon = "🟢" if panel.get("enabled", True) else "🔴"
        today_count = await asyncio.to_thread(get_today_panel_message_count, panel_id)
        if isinstance(msgs, dict) and msgs.get('error'):
            err = msgs['error']
            if err == 'login_failed':
                err_detail = "Panel এ login করা হয়নি। Panel চালু করুন।"
            else:
                err_detail = "Session expire হয়ে গেছে। Panel বন্ধ করে আবার চালু করুন।"
            msg_text = (
                f"❌ <b>{panel['name']} — Login হয়নি</b>\n\n"
                f"{err_detail}\n\n"
                f"📊 Last 24h Total: <b>{today_count}</b>\n"
                f"{icon} Status: {'Active' if panel.get('enabled', True) else 'Inactive'}"
            )
        elif msgs:
            msg_text = _build_last_msgs_text(panel['name'], msgs, today_count, icon)
        else:
            msg_text = (
                f"📨 <b>{panel['name']} — Last Messages</b>\n\n"
                f"<i>Panel থেকে কোনো message পাওয়া যায়নি।</i>\n\n"
                f"📊 Last 24h Total: <b>{today_count}</b>\n"
                f"{icon} Status: {'Active' if panel.get('enabled', True) else 'Inactive'}"
            )
        refresh_markup = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 Refresh", callback_data=f"sms_panel_refresh_msg_{panel_id}")]
        ])
        try:
            await query.message.edit_text(msg_text, parse_mode='HTML', reply_markup=refresh_markup)
        except Exception as e:
            logger.error(f"sms_panel_refresh_msg edit error: {e}")

    elif data == "cm_view_stats":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        stats = get_country_stats()
        if not stats:
            message = f"{_sx('❌')} No data available."
        else:
            message = f"{_sx('📊')} <b>Country Statistics</b>\n\n"
            for country, total, available in stats:
                message += f"• {country}: {available}/{total} available\n"
        keyboard = [[InlineKeyboardButton("🔙 Back to Country Manager", callback_data="cm_back")]]
        await query.edit_message_text(message, parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))
    
    elif data == "cm_delete_country":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        countries = get_countries()
        if not countries:
            await query.edit_message_text(f"{_sx('❌')} No countries available.", parse_mode='HTML', reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="cm_back")]]))
            return
        keyboard = []
        country_lines = []
        for country_id, country_name in countries:
            total, available = get_numbers_count_by_country(country_id)
            plain = _clean_name(country_name)
            country_lines.append(f"{get_animated_flag_html(plain)} <b>{plain}</b> ({total})")
            keyboard.append([InlineKeyboardButton(f"{plain} ({total})", callback_data=f"delete_country_{country_id}")])
        keyboard.append([InlineKeyboardButton("🔙 Back to Country Manager", callback_data="cm_back")])
        countries_text = "\n".join(country_lines)
        await query.edit_message_text(
            f"<b>Delete Country</b>\n\nSelect a country to delete:\n\n{countries_text}",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )
    
    elif data == "cm_edit_number":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        countries = get_countries()
        if not countries:
            await query.edit_message_text(f"{_sx('❌')} No countries available.", parse_mode='HTML', reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="cm_back")]]))
            return
        keyboard = []
        country_lines = []
        for country_id, country_name in countries:
            total, available = get_numbers_count_by_country(country_id)
            plain = _clean_name(country_name)
            country_lines.append(f"{get_animated_flag_html(plain)} <b>{plain}</b> ({total})")
            keyboard.append([InlineKeyboardButton(f"{plain} ({total})", callback_data=f"edit_country_{country_id}")])
        keyboard.append([InlineKeyboardButton("🔙 Back to Country Manager", callback_data="cm_back")])
        countries_text = "\n".join(country_lines)
        await query.edit_message_text(
            f"<b>Edit Numbers</b>\n\nSelect a country to add more numbers:\n\n{countries_text}",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )
    
    elif data == "cm_back":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        await query.edit_message_text(
            f"{_sx('🚩')} <b>Country Manager</b>\n\n"
            f"{_sx('➕')} Add Numbers — Upload new numbers\n"
            f"{_sx('🌐')} Add Country — Register a country\n"
            f"{_sx('📱')} Add Service — Create a service\n"
            f"{_sx('🗺')} Service Map — Link countries to services\n"
            f"{_sx('🔄')} Reset Number — Reset usage\n\n"
            "Select an option:",
            parse_mode='HTML',
            reply_markup=get_country_manager_keyboard()
        )

    elif data == "grp_btn_number":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        current = get_group_number_btn_link()
        context.user_data['awaiting_grp_number_link'] = True
        await query.edit_message_text(
            f"📱 <b>Number Button Link</b>\n\n"
            f"বর্তমান লিংক: <code>{current}</code>\n\n"
            f"নতুন লিংক পাঠান:",
            parse_mode='HTML'
        )

    elif data == "grp_btn_channel":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        current = get_group_channel_btn_link()
        context.user_data['awaiting_grp_channel_link'] = True
        await query.edit_message_text(
            f"📢 <b>Channel Button Link</b>\n\n"
            f"বর্তমান লিংক: <code>{current}</code>\n\n"
            f"নতুন লিংক পাঠান:",
            parse_mode='HTML'
        )

    elif data == "do_add_admin":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        context.user_data['awaiting_new_admin'] = True
        await query.edit_message_text(
            "*Add New Admin*\n\nSend the Telegram UID of the new admin:\n(Numbers only, e.g: 123456789)",
            parse_mode='Markdown'
        )

    elif data == "do_remove_admin":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        admins = get_all_admins()
        keyboard = []
        for uid in PROTECTED_ADMIN_IDS:
            keyboard.append([InlineKeyboardButton(f"🛡️ {uid} (Protected)", callback_data=f"protected_admin_id_{uid}")])
        for uid in admins:
            if uid not in PROTECTED_ADMIN_IDS:
                keyboard.append([InlineKeyboardButton(f"❌ {uid}", callback_data=f"remove_admin_id_{uid}")])
        if not keyboard:
            await query.edit_message_text(f"{_sx('❌')} No admins found.", parse_mode='HTML')
            return
        keyboard.append([InlineKeyboardButton("🔙 Close", callback_data="am_noop")])
        await query.edit_message_text(
            f"{_sx('👑')} <b>Remove Admin</b>\n\n🛡️ Protected admins cannot be removed",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )

    elif data == "am_add_admin":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        context.user_data['awaiting_new_admin'] = True
        context.user_data['am_msg_id'] = query.message.message_id
        await query.edit_message_text(
            f"{_sx('👑')} <b>Add Admin</b>\n\nSend the Telegram UID of the new admin:\n(Numbers only, e.g: 123456789)",
            parse_mode='HTML'
        )

    elif data == "am_remove_admin":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        admins = get_all_admins()
        keyboard = []
        for uid in PROTECTED_ADMIN_IDS:
            keyboard.append([InlineKeyboardButton(f"🛡️ {uid} (Protected)", callback_data=f"protected_admin_id_{uid}")])
        for uid in admins:
            if uid not in PROTECTED_ADMIN_IDS:
                keyboard.append([InlineKeyboardButton(f"❌ {uid}", callback_data=f"remove_admin_id_{uid}")])
        if not keyboard:
            keyboard.append([InlineKeyboardButton("— No admins —", callback_data="am_noop")])
        keyboard.append([InlineKeyboardButton("🔙 Back", callback_data="am_back")])
        await query.edit_message_text(
            f"{_sx('👑')} <b>Remove Admin</b>\n\n🛡️ Protected admins cannot be removed",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )

    elif data == "am_user_count":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        count = get_user_count()
        keyboard = [[InlineKeyboardButton("🔙 Back", callback_data="am_back")]]
        await query.edit_message_text(
            f"*Total Users:* {count}",
            parse_mode='Markdown',
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif data == "am_back":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        await query.edit_message_text(
            f"{_sx('👑')} <b>Admin Manager</b>\n\n"
            f"{_sx('➕')} Add Admin — Add a new admin\n"
            f"{_sx('👥')} User Activity — Activity report & Excel export\n\n"
            "Select an option:",
            parse_mode='HTML',
            reply_markup=get_admin_manager_keyboard()
        )

    elif data.startswith("remove_admin_id_"):
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        uid_to_remove = int(data.split("remove_admin_id_")[1])
        success, msg = remove_admin(uid_to_remove)
        if success:
            await query.edit_message_text(f"{_sx('✅')} Admin <code>{uid_to_remove}</code> has been removed!", parse_mode='HTML')
        else:
            await query.answer(f"❌ {msg}", show_alert=True)

    elif data.startswith("protected_admin_id_"):
        uid = data.split("protected_admin_id_")[1]
        await query.answer(f"🛡️ UID {uid} is a protected admin — cannot be removed!", show_alert=True)
    
    # New EDIT NUMBER callback handlers
    elif data == "edit_numbers":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
            
        # Show countries list for editing
        countries = get_countries()
        if not countries:
            await query.edit_message_text(f"{_sx('❌')} No countries available.", parse_mode='HTML')
            return
        
        keyboard = []
        country_lines = []
        for country_id, country_name in countries:
            total, available = get_numbers_count_by_country(country_id)
            plain = _clean_name(country_name)
            country_lines.append(f"{get_animated_flag_html(plain)} <b>{plain}</b> ({total})")
            keyboard.append([InlineKeyboardButton(f"{plain} ({total})", callback_data=f"edit_country_{country_id}")])
        
        keyboard.append([InlineKeyboardButton("🔙 Back to Admin Panel", callback_data="back_to_admin")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        countries_text = "\n".join(country_lines)
        await query.edit_message_text(
            f"<b>Edit Numbers</b>\n\nSelect a country to add more numbers:\n\n{countries_text}",
            reply_markup=reply_markup,
            parse_mode='HTML'
        )
    
    elif data.startswith("edit_country_"):
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
            
        country_id = int(data.split("_")[2])
        
        # Get country name
        countries = get_countries()
        country_name = next((name for cid, name in countries if cid == country_id), "Unknown")
        
        # Store country ID in user data for file upload
        context.user_data['edit_country_id'] = country_id
        context.user_data['edit_country_name'] = country_name
        
        await query.edit_message_text(
            f"*Edit Numbers for {country_name}*\n\nPlease send a file with numbers (TXT, CSV, Excel — any format):",
            parse_mode='Markdown'
        )

    # ─── Referral callbacks ────────────────────────────────────────────────────
    elif data == "ref_toggle":
        if not admin_check():
            await query.answer("❌ Not authorized.", show_alert=True)
            return
        settings = get_referral_settings()
        new_val = 0 if settings['enabled'] else 1
        set_referral_setting('enabled', new_val)
        settings = get_referral_settings()
        status = f"{_sx('✅')} Enabled" if settings['enabled'] else f"{_sx('❌')} Disabled"
        keyboard = [
            [InlineKeyboardButton(f"Power: {'✅ Enabled' if settings['enabled'] else '❌ Disabled'}", callback_data="ref_toggle")],
            [InlineKeyboardButton(f"🎁 Reward: {settings['reward']} {settings['label']} per referral", callback_data="ref_set_reward")],
            [InlineKeyboardButton("👤 Check User Balance", callback_data="ref_check_user")],
            [InlineKeyboardButton("➕ Add Balance to User", callback_data="ref_add_balance")],
            [InlineKeyboardButton("➖ Remove Balance from User", callback_data="ref_remove_balance")],
        ]
        total, unique_ref, total_rewards = get_total_referral_stats()
        msg = (
            f"{_sx('🎁')} <b>Referral Settings</b>\n\n"
            f"Status: <b>{status}</b>\n"
            f"Reward per referral: <b>{settings['reward']} {settings['label']}</b>\n\n"
            f"<b>Overall Stats:</b>\n"
            f"├ Total referrals: <b>{total}</b>\n"
            f"├ Active referrers: <b>{unique_ref}</b>\n"
            f"└ Total rewards given: <b>{total_rewards} {settings['label']}</b>"
        )
        await query.edit_message_text(msg, parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))

    elif data == "ref_set_reward":
        if not admin_check():
            await query.answer("❌ Not authorized.", show_alert=True)
            return
        context.user_data['awaiting_ref_reward'] = True
        await query.edit_message_text(f"{_sx('🎁')} <b>Set Reward Amount</b>\n\nSend the number of points to award per referral (e.g. <code>10</code>):", parse_mode='HTML')

    elif data == "ref_check_user":
        if not admin_check():
            await query.answer("❌ Not authorized.", show_alert=True)
            return
        context.user_data['awaiting_ref_check_user'] = True
        await query.edit_message_text(f"👤 <b>Check User Balance</b>\n\nSend the user's Telegram UID:", parse_mode='HTML')

    elif data == "ref_add_balance":
        if not admin_check():
            await query.answer("❌ Not authorized.", show_alert=True)
            return
        context.user_data['awaiting_ref_add_balance'] = True
        await query.edit_message_text(f"{_sx('➕')} <b>Add Balance</b>\n\nSend: <code>USER_ID AMOUNT</code>\nExample: <code>123456789 50</code>", parse_mode='HTML')

    elif data == "ref_remove_balance":
        if not admin_check():
            await query.answer("❌ Not authorized.", show_alert=True)
            return
        context.user_data['awaiting_ref_remove_balance'] = True
        await query.edit_message_text(f"{_sx('➖')} <b>Remove Balance</b>\n\nSend: <code>USER_ID AMOUNT</code>\nExample: <code>123456789 20</code>", parse_mode='HTML')

    # ─── Withdraw callbacks ────────────────────────────────────────────────────
    elif data == "withdraw_start":
        uid = query.from_user.id
        bal = get_user_balance_data(uid)
        settings = get_referral_settings()
        wcfg = get_withdraw_config()
        if not wcfg['enabled']:
            await query.answer("❌ Withdraw এখন বন্ধ আছে!", show_alert=True)
            return
        if bal['balance'] <= 0:
            await query.answer("❌ You have no balance to withdraw!", show_alert=True)
            return
        if wcfg['min_amount'] > 0 and bal['balance'] < wcfg['min_amount']:
            await query.answer(
                f"❌ Minimum {wcfg['min_amount']} {settings['label']} লাগবে!\nআপনার ব্যালেন্স: {bal['balance']} {settings['label']}",
                show_alert=True
            )
            return
        keyboard = [
            [InlineKeyboardButton("🟢 Bkash", callback_data="withdraw_wallet_Bkash"),
             InlineKeyboardButton("🟠 Nagad", callback_data="withdraw_wallet_Nagad")],
            [InlineKeyboardButton("🟡 Binance", callback_data="withdraw_wallet_Binance")]
        ]
        await query.edit_message_text(
            f"💳 <b>Withdraw</b>\n\n"
            f"{_sx('💵')} Your balance: <b>{bal['balance']} {settings['label']}</b>\n\n"
            f"Select your wallet:",
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif data.startswith("withdraw_wallet_"):
        wallet = data[16:]
        uid = query.from_user.id
        bal = get_user_balance_data(uid)
        settings = get_referral_settings()
        context.user_data['withdraw_wallet'] = wallet
        context.user_data['awaiting_withdraw_details'] = True
        await query.edit_message_text(
            f"💳 <b>{wallet} Withdraw</b>\n\n"
            f"💵 Available: <b>{bal['balance']} {settings['label']}</b>\n\n"
            f"Send your {wallet} number and amount:\n"
            f"Format: <code>NUMBER AMOUNT</code>\n"
            f"Example: <code>01712345678 100</code>",
            parse_mode='HTML'
        )

    elif data.startswith("wd_approve_"):
        if not admin_check():
            await query.answer("❌ Not authorized.", show_alert=True)
            return
        req_id = int(data[11:])
        req = get_withdraw_by_id(req_id)
        if not req:
            await query.answer("❌ Request not found!", show_alert=True)
            return
        rid, uid, uname, wtype, waddr, amount, label, status = req
        if status != 'pending':
            await query.answer(f"Already {status}!", show_alert=True)
            return
        update_withdraw_status(req_id, 'approved')
        try:
            await context.bot.send_message(
                chat_id=uid,
                text=f"✅ <b>Withdraw Approved!</b>\n\n"
                     f"💳 Wallet: <b>{wtype}</b>\n"
                     f"📱 Address: <code>{waddr}</code>\n"
                     f"💰 Amount: <b>{amount} {label}</b>\n\n"
                     f"Your withdrawal has been processed! 🎉",
                parse_mode='HTML'
            )
        except Exception:
            pass
        display = f"@{uname}" if uname else str(uid)
        await query.edit_message_text(
            f"✅ <b>Approved!</b>\n\nRequest #{req_id} — {display}\n{wtype} | <code>{waddr}</code> | {amount} {label}",
            parse_mode='HTML'
        )

    elif data.startswith("wd_reject_"):
        if not admin_check():
            await query.answer("❌ Not authorized.", show_alert=True)
            return
        req_id = int(data[10:])
        req = get_withdraw_by_id(req_id)
        if not req:
            await query.answer("❌ Request not found!", show_alert=True)
            return
        rid, uid, uname, wtype, waddr, amount, label, status = req
        if status != 'pending':
            await query.answer(f"Already {status}!", show_alert=True)
            return
        update_withdraw_status(req_id, 'rejected', refund_user_id=uid, refund_amount=amount)
        try:
            await context.bot.send_message(
                chat_id=uid,
                text=f"❌ <b>Withdraw Rejected</b>\n\n"
                     f"💳 Wallet: <b>{wtype}</b>\n"
                     f"💰 Amount: <b>{amount} {label}</b>\n\n"
                     f"Your balance has been refunded. 🔄",
                parse_mode='HTML'
            )
        except Exception:
            pass
        display = f"@{uname}" if uname else str(uid)
        await query.edit_message_text(
            f"❌ <b>Rejected!</b>\n\nRequest #{req_id} — {display}\n{amount} {label} refunded to user.",
            parse_mode='HTML'
        )


    elif data == "custom_msg_set":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        context.user_data['awaiting_custom_message'] = True
        await query.answer()
        await query.message.reply_text(
            f"{_sx('✏️')} <b>Custom Message সেট করুন</b>\n\n"
            "এখন যে message টি পাঠাবেন সেটি User Panel-এর 📋 Notice বাটনে দেখাবে।\n\n"
            "⚠️ যেকোনো ফরম্যাটে message লিখতে পারবেন।",
            parse_mode='HTML'
        )

    elif data == "custom_msg_remove":
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        await asyncio.to_thread(remove_custom_message)
        await query.answer("✅ সব message রিমুভ হয়েছে!", show_alert=True)
        await query.edit_message_text(
            f"{_sx('✏️')} <b>Custom Message</b>\n\n"
            "<i>সব message সফলভাবে রিমুভ করা হয়েছে।</i>",
            parse_mode='HTML'
        )

    elif data.startswith("del_custmsg_"):
        if not admin_check():
            await query.answer("❌ You are not authorized.", show_alert=True)
            return
        msg_id = data[len("del_custmsg_"):]
        await asyncio.to_thread(remove_custom_message_by_id, msg_id)
        await query.answer("✅ Message delete হয়েছে!", show_alert=True)
        # Refresh the list in the same message
        msgs = await asyncio.to_thread(get_custom_messages)
        if not msgs:
            await query.edit_message_text(
                f"{_sx('🗑')} <b>Message Delete করুন</b>\n\n"
                "<i>✅ সব message ডিলিট হয়েছে।</i>",
                parse_mode='HTML'
            )
            return
        keyboard = []
        msg_lines = []
        for i, (mid, txt) in enumerate(msgs):
            preview = txt[:100] + ("…" if len(txt) > 100 else "")
            msg_lines.append(f"<b>#{i+1}.</b> {preview}")
            keyboard.append([InlineKeyboardButton(
                f"❌ #{i+1} Delete",
                callback_data=f"del_custmsg_{mid}"
            )])
        body = "\n\n".join(msg_lines)
        await query.edit_message_text(
            f"{_sx('🗑')} <b>Message Delete করুন</b>\n\n{body}\n\nনির্দিষ্ট message এর পাশের বাটনে ক্লিক করুন:",
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup(keyboard)
        )


def clear_awaiting_states(context: ContextTypes.DEFAULT_TYPE):
    for key in [
        'awaiting_new_country', 'awaiting_add_numbers_country', 'awaiting_numbers_file',
        'current_country_name', 'awaiting_number_delete', 'awaiting_new_admin',
        'awaiting_broadcast', 'awaiting_reset_country', 'awaiting_new_service',
        'awaiting_service_map', 'edit_country_id', 'edit_country_name',
        'awaiting_country_otp_reward', 'country_otp_reward_name',
        'awaiting_ref_reward', 'awaiting_ref_check_user',
        'awaiting_ref_add_balance', 'awaiting_ref_remove_balance',
        'awaiting_numbers_per_request', 'awaiting_otp_link',
        'awaiting_grp_number_link', 'awaiting_grp_channel_link',
        'awaiting_channel_title', 'awaiting_channel_link', 'awaiting_remove_channel',
        'pending_channel_title', 'awaiting_check_interval', 'awaiting_custom_message',
        'awaiting_panel_name', 'awaiting_panel_login_url', 'awaiting_panel_message_url',
        'awaiting_panel_username', 'awaiting_panel_password',
        'awaiting_panel_num_col_name', 'awaiting_panel_num_col_idx',
        'awaiting_panel_svc_col_name', 'awaiting_panel_svc_col_idx',
        'awaiting_panel_msg_col_name', 'awaiting_panel_msg_col_idx',
        'panel_setup_name', 'panel_setup_login_url', 'panel_setup_message_url',
        'panel_setup_username', 'panel_setup_password',
        'panel_setup_num_col_name', 'panel_setup_num_col_idx',
        'panel_setup_svc_col_name', 'panel_setup_svc_col_idx', 'panel_setup_msg_col_name',
        'awaiting_panel_edit_username', 'awaiting_panel_edit_password',
        'awaiting_manual_forward_panel', 'awaiting_manual_forward_step', 'manual_forward_number',
        'awaiting_panel_interval',
        'awaiting_direct_uid', 'awaiting_direct_message', 'direct_msg_uid',
    ]:
        context.user_data.pop(key, None)

# Handle button clicks
async def handle_button_click(update: Update, context: ContextTypes.DEFAULT_TYPE):
    username = update.effective_user.username
    user_id = update.effective_user.id
    text = update.message.text

    # Guard: if panel setup or manual forward is in progress, block button presses
    _panel_setup_keys = [
        'awaiting_panel_name', 'awaiting_panel_login_url', 'awaiting_panel_message_url',
        'awaiting_panel_username', 'awaiting_panel_password',
        'awaiting_panel_num_col_name', 'awaiting_panel_num_col_idx',
        'awaiting_panel_svc_col_name', 'awaiting_panel_svc_col_idx',
        'awaiting_panel_msg_col_name', 'awaiting_panel_msg_col_idx',
    ]
    if any(context.user_data.get(k) for k in _panel_setup_keys):
        step_map = {
            'awaiting_panel_name':         '1/11 — Panel নাম',
            'awaiting_panel_login_url':     '2/11 — Login URL',
            'awaiting_panel_message_url':   '3/11 — Message URL',
            'awaiting_panel_username':      '4/11 — Username',
            'awaiting_panel_password':      '5/11 — Password',
            'awaiting_panel_num_col_name':  '6/11 — নাম্বার কলামের নাম',
            'awaiting_panel_num_col_idx':   '7/11 — নাম্বার কলাম নম্বর',
            'awaiting_panel_svc_col_name':  '8/11 — Service কলামের নাম',
            'awaiting_panel_svc_col_idx':   '9/11 — Service কলাম নম্বর',
            'awaiting_panel_msg_col_name':  '10/11 — Message কলামের নাম',
            'awaiting_panel_msg_col_idx':   '11/11 — Message কলাম নম্বর',
        }
        current_step = next((v for k, v in step_map.items() if context.user_data.get(k)), 'অজানা')
        await update.message.reply_text(
            f"⚠️ <b>Panel Setup চলছে!</b>\n\n"
            f"এখন Step {current_step} দেওয়ার অপেক্ষায় আছি।\n"
            f"Setup শেষ করুন অথবা /cancel দিয়ে বাতিল করুন।",
            parse_mode='HTML'
        )
        return
    if context.user_data.get('awaiting_panel_interval'):
        await update.message.reply_text(
            f"⚠️ <b>Interval সেট চলছে!</b>\n\n"
            f"একটি সংখ্যা দিন (সেকেন্ড)।\n"
            f"বাতিল করতে /cancel দিন।",
            parse_mode='HTML'
        )
        return

    # Clear all pending states before processing any new button
    clear_awaiting_states(context)

    # Allow regular users to view stats
    if text == "📊 View Stats" and not is_admin(username, user_id):
        await show_stats(update, context)
        return

    if not is_admin(username, user_id):
        await update.message.reply_text(f"{_sx('❌')} You are not authorized to use admin commands.", parse_mode='HTML')
        return
    
    if text == "➕ Add Numbers":
        countries = get_countries()
        if not countries:
            await update.message.reply_text(f"{_sx('❌')} No countries saved yet. Please add a country first using Add Country.", parse_mode='HTML')
            return
        country_list = "\n".join([
            f"{get_animated_flag_html(_clean_name(name))} <code>{_clean_name(name)}</code>"
            for _, name in countries
        ])
        context.user_data['awaiting_add_numbers_country'] = True
        await update.message.reply_text(
            f"📋 <b>Saved Countries:</b>\n\n{country_list}\n\n"
            f"Type the exact country name from the list above:",
            parse_mode='HTML'
        )

    elif text == "🌍 Add Country":
        context.user_data['awaiting_new_country'] = True
        stats = get_country_stats()
        if stats:
            country_lines = "\n".join(
                [f"{get_animated_flag_html(_clean_name(name))} <code>{_clean_name(name)}</code> — {int(available or 0)} remaining"
                 for name, total, available in stats]
            )
            country_section = f"\n\n📋 <b>Saved Countries:</b>\n{country_lines}"
        else:
            country_section = "\n\n📋 <b>No countries added yet.</b>"
        await update.message.reply_text(
            f"<b>Add Country</b>\n\nSend the name of the new country:{country_section}\n\n🗑️ <b>Delete a country:</b>\nType <code>delete</code> followed by country name.\nExample: <code>delete India</code>",
            parse_mode='HTML'
        )

    elif text == "📊 View Stats":
        stats = get_country_stats()
        if not stats:
            await update.message.reply_text(f"{_sx('❌')} No data available.", parse_mode='HTML')
            return
        message = f"{_sx('📊')} <b>Country Statistics</b>\n\n"
        for country, total, available in stats:
            plain = _clean_name(country)
            flag = get_animated_flag_html(plain)
            message += f"• {flag} <code>{plain}</code>: {available}/{total} available\n"
        await update.message.reply_text(message, parse_mode='HTML')
    
    elif text == "🤖 Bot Status":
        # Fetch all data concurrently
        (
            ref_settings,
            (total_refs, unique_refs, total_rewards),
            otp_link,
            channels,
            check_interval,
            w_config,
            country_stats,
            all_users,
            all_panels,
        ) = await asyncio.gather(
            asyncio.to_thread(get_referral_settings),
            asyncio.to_thread(get_total_referral_stats),
            asyncio.to_thread(get_otp_link),
            asyncio.to_thread(get_join_channels),
            asyncio.to_thread(get_check_interval),
            asyncio.to_thread(get_withdraw_config),
            asyncio.to_thread(get_country_stats),
            asyncio.to_thread(get_all_users),
            asyncio.to_thread(get_sms_panels),
        )

        user_count = len(all_users)
        ref_status = f"{_sx('✅')} Enabled" if ref_settings['enabled'] else f"{_sx('❌')} Disabled"
        w_status   = f"{_sx('✅')} Enabled" if w_config['enabled']     else f"{_sx('❌')} Disabled"
        w_group    = w_config['group_chat_id'] or "(Not set)"

        if country_stats:
            country_lines = "\n".join(
                f"  • {_clean_name(name)}: <b>{available}</b> available / <b>{total}</b> total"
                for name, total, available in country_stats
            )
        else:
            country_lines = "  (No countries added)"

        ch_lines = "\n".join([f"  • {t}: <code>{l}</code>" for _, t, l in channels]) or "  (None)"

        # Panel summary
        total_panels = len(all_panels)
        enabled_panels  = [p for p in all_panels if p.get('enabled', True)]
        disabled_panels = [p for p in all_panels if not p.get('enabled', True)]
        logged_in_ids   = set(_panel_sessions.keys())
        monitoring_ids  = {pid for pid, t in _active_tasks.items() if not t.done()}
        logged_in_count   = sum(1 for p in all_panels if p['id'] in logged_in_ids)
        monitoring_count  = sum(1 for p in all_panels if p['id'] in monitoring_ids)

        if total_panels == 0:
            panel_lines = "  (No panels added yet)"
        else:
            panel_lines = ""
            for p in all_panels:
                pid = p['id']
                p_name = p.get('name', pid)
                is_enabled   = p.get('enabled', True)
                is_logged_in = pid in logged_in_ids
                is_running   = pid in monitoring_ids
                status_icon  = "🟢" if is_enabled else "🔴"
                login_icon   = "✅" if is_logged_in else "❌"
                monitor_icon = "📡" if is_running   else "⏸"
                panel_lines += (
                    f"  {status_icon} <b>{p_name}</b>\n"
                    f"      Login: {login_icon}  |  Monitor: {monitor_icon}\n"
                )

        msg = (
            f"{_sx('🤖')} <b>Bot Status</b>\n"
            "━━━━━━━━━━━━━━━━━━━━\n\n"
            f"{_sx('👥')} <b>Total Users:</b> <b>{user_count}</b>\n\n"
            f"🌍 <b>Country Numbers:</b>\n{country_lines}\n\n"
            f"📡 <b>SMS Panels</b>\n"
            f"  • Total Panels: <b>{total_panels}</b>  |  Active: <b>{len(enabled_panels)}</b>  |  Inactive: <b>{len(disabled_panels)}</b>\n"
            f"  • Login Success: <b>{logged_in_count}</b>  |  Monitoring: <b>{monitoring_count}</b>\n"
            f"{panel_lines}\n"
            f"{_sx('🎁')} <b>Referral Settings</b>\n"
            f"  • Status: {ref_status}\n"
            f"  • Reward: <b>{ref_settings['reward']} {ref_settings['label']}</b>\n"
            f"  • Total Referrals: <b>{total_refs}</b>\n"
            f"  • Unique Referrers: <b>{unique_refs}</b>\n"
            f"  • Total Distributed: <b>{total_rewards} {ref_settings['label']}</b>\n\n"
            f"💳 <b>Withdraw Settings</b>\n"
            f"  • Status: {w_status}\n"
            f"  • Min Amount: <b>{w_config['min_amount']}</b>\n"
            f"  • Group ID: <code>{w_group}</code>\n\n"
            f"{_sx('🔗')} <b>Link Settings</b>\n"
            f"  • OTP Link: <code>{otp_link}</code>\n"
            f"  • Join Channels:\n{ch_lines}\n\n"
            f"⏱ <b>Auto Check</b>\n"
            f"  • Check Interval: every <b>{check_interval} minutes</b>\n\n"
            f"🕐 <b>Report Time:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        )
        await update.message.reply_text(msg, parse_mode='HTML', reply_markup=get_admin_keyboard())

        # ── Build Excel user list and send ────────────────────────────────────

        if all_users:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Users"

            # Header styling
            hdr_font  = Font(bold=True, color="FFFFFF", size=11)
            hdr_fill  = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            hdr_align = Alignment(horizontal="center", vertical="center")

            headers = ["#", "User ID", "Username", "First Name", "Last Name",
                       "Verified", "Balance", "Total Earned", "Joined At"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font  = hdr_font
                cell.fill  = hdr_fill
                cell.alignment = hdr_align

            ws.row_dimensions[1].height = 20

            # Data rows
            alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for i, u in enumerate(all_users, 1):
                uid    = u.get('user_id', '')
                uname  = u.get('username') or ''
                fname  = u.get('first_name') or ''
                lname  = u.get('last_name') or ''
                veri   = "Yes" if u.get('is_verified') else "No"
                bal    = u.get('balance', 0)
                earned = u.get('total_earned', 0)
                joined = str(u.get('joined_at', ''))
                row_data = [i, uid, uname, fname, lname, veri, bal, earned, joined]
                ws.append(row_data)
                if i % 2 == 0:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=i + 1, column=col).fill = alt_fill

            # Auto-size columns
            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 3), 40)

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            fname_xlsx = f"users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            await update.message.reply_document(
                document=bio,
                filename=fname_xlsx,
                caption=(
                    f"{_sx('📊')} <b>User List (Excel)</b>\n"
                    f"{_sx('👥')} Total: <b>{user_count}</b> users\n"
                    f"🕐 {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                ),
                parse_mode='HTML'
            )
        return

    elif text == "👥 User Activity":
        stats, inactive_users, all_users = await asyncio.gather(
            asyncio.to_thread(get_activity_stats),
            asyncio.to_thread(get_inactive_users, 30),
            asyncio.to_thread(get_all_users),
        )
        active_pct = round(stats['today'] / stats['total'] * 100) if stats['total'] else 0
        week_pct   = round(stats['week'] / stats['total'] * 100) if stats['total'] else 0

        msg = (
            f"👥 <b>User Activity Report</b>\n"
            "━━━━━━━━━━━━━━━━━━━━\n\n"
            f"📊 <b>Total Users:</b> <b>{stats['total']}</b>\n\n"
            f"✅ <b>Active Today:</b> <b>{stats['today']}</b> ({active_pct}%)\n"
            f"📅 <b>Active Last 7 Days:</b> <b>{stats['week']}</b> ({week_pct}%)\n"
            f"😴 <b>Inactive 30+ Days:</b> <b>{stats['inactive_30d']}</b>\n"
            f"🆕 <b>Never Used Bot:</b> <b>{stats['never']}</b>\n\n"
            f"🕐 <b>Report Time:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        )
        await update.message.reply_text(msg, parse_mode='HTML', reply_markup=get_admin_keyboard())

        # Excel export — full activity list
        if all_users:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "User Activity"

            hdr_font  = Font(bold=True, color="FFFFFF", size=11)
            hdr_fill  = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
            hdr_align = Alignment(horizontal="center", vertical="center")

            headers = ["#", "User ID", "Username", "First Name", "Last Name",
                       "Joined At", "Last Active", "Usage Count", "Balance"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font      = hdr_font
                cell.fill      = hdr_fill
                cell.alignment = hdr_align
            ws.row_dimensions[1].height = 20

            inactive_ids = {str(u.get('user_id', '')) for u in inactive_users}
            inactive_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
            alt_fill      = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # Sort: most recently active first
            def _sort_key(u):
                la = u.get("last_active")
                return la if la else ""
            sorted_users = sorted(all_users, key=_sort_key, reverse=True)

            for i, u in enumerate(sorted_users, 1):
                uid   = u.get('user_id', '')
                uname = u.get('username') or ''
                fname = u.get('first_name') or ''
                lname = u.get('last_name') or ''
                joined = str(u.get('joined_at', ''))
                last_a = str(u.get('last_active') or 'Never')
                usage  = u.get('usage_count', 0)
                bal    = u.get('balance', 0)
                ws.append([i, uid, uname, fname, lname, joined, last_a, usage, bal])
                row_fill = inactive_fill if str(uid) in inactive_ids else (alt_fill if i % 2 == 0 else None)
                if row_fill:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=i + 1, column=col).fill = row_fill

            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max(12, max_len + 3), 40)

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            fname_xlsx = f"activity_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            await update.message.reply_document(
                document=bio,
                filename=fname_xlsx,
                caption=(
                    f"📊 <b>User Activity (Excel)</b>\n"
                    f"👥 Total: <b>{stats['total']}</b> users\n"
                    f"✅ Active Today: <b>{stats['today']}</b>  |  "
                    f"😴 Inactive 30d+: <b>{stats['inactive_30d']}</b>\n"
                    f"🔴 Inactive rows are highlighted in red\n"
                    f"🕐 {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
                ),
                parse_mode='HTML'
            )
        return

    elif text == "👑 Admin Manager":
        await update.message.reply_text(
            f"{_sx('👑')} <b>Admin Manager</b>\n\n"
            f"{_sx('➕')} Add Admin — Add a new admin\n"
            f"{_sx('👥')} User Activity — Activity report & Excel export\n\n"
            "Select an option:",
            parse_mode='HTML',
            reply_markup=get_admin_manager_keyboard()
        )
        return

    elif text == "⚙️ Settings":
        await update.message.reply_text(
            "⚙️ <b>Settings</b>",
            parse_mode='HTML',
            reply_markup=get_settings_keyboard()
        )
        return

    elif text == "🔢 Numbers Per Request":
        current = get_numbers_per_request()
        context.user_data['awaiting_numbers_per_request'] = True
        await update.message.reply_text(
            f"🔢 <b>Numbers Per Request</b>\n\nCurrent: <b>{current}</b>\n\nSend the number of phone numbers to show per request (e.g. <code>1</code>, <code>3</code>, <code>5</code>):",
            parse_mode='HTML'
        )
        return

    elif text == "🔗 Link Settings":
        otp = get_otp_link()
        channels = get_join_channels()
        ch_lines = "\n".join([f"  {i+1}. {t}\n      <code>{l}</code>" for i, (_, t, l) in enumerate(channels)]) or "  (none)"
        num_link = get_group_number_btn_link()
        ch_link  = get_group_channel_btn_link()
        await update.message.reply_text(
            f"🔗 <b>Link Settings</b>\n\n"
            f"📲 OTP Link: <code>{otp}</code>\n\n"
            f"📢 Join Channels:\n{ch_lines}\n\n"
            f"📤 <b>Group Button Links:</b>\n"
            f"  📱 Number: <code>{num_link}</code>\n"
            f"  📢 Channel: <code>{ch_link}</code>",
            parse_mode='HTML',
            reply_markup=get_link_settings_keyboard()
        )
        return

    elif text == "📤 Group Button Links":
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized.", parse_mode='HTML')
            return
        num_link = get_group_number_btn_link()
        ch_link  = get_group_channel_btn_link()
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton(f"📱 Number Button Link", callback_data="grp_btn_number")],
            [InlineKeyboardButton(f"📢 Channel Button Link", callback_data="grp_btn_channel")],
        ])
        await update.message.reply_text(
            f"📤 <b>Group Button Links</b>\n\n"
            f"📱 Number: <code>{num_link}</code>\n"
            f"📢 Channel: <code>{ch_link}</code>\n\n"
            f"পরিবর্তন করতে নিচের বাটন চাপুন:",
            parse_mode='HTML',
            reply_markup=keyboard
        )
        return

    elif text == "📲 Set OTP Link":
        context.user_data['awaiting_otp_link'] = True
        current = get_otp_link()
        await update.message.reply_text(
            f"📲 <b>Set OTP Link</b>\n\nCurrent: <code>{current}</code>\n\nSend the new OTP group link:",
            parse_mode='HTML'
        )
        return

    elif text == "❌ Remove OTP Link":
        set_otp_link("")
        await update.message.reply_text(
            f"{_sx('✅')} OTP link removed successfully!",
            parse_mode='HTML',
            reply_markup=get_link_settings_keyboard()
        )
        return

    elif text == "➕ Add Channel":
        context.user_data['awaiting_channel_title'] = True
        await update.message.reply_text(
            f"{_sx('➕')} <b>Add Channel</b>\n\nStep 1/2 — Send the channel <b>title/name</b> (e.g. <code>My Channel</code>):",
            parse_mode='HTML'
        )
        return

    elif text == "➖ Remove Channel":
        channels = get_join_channels()
        if not channels:
            await update.message.reply_text(f"{_sx('❌')} No channels to remove.", parse_mode='HTML', reply_markup=get_link_settings_keyboard())
            return
        lines = "\n".join([f"{i+1}. {t}" for i, (_, t, l) in enumerate(channels)])
        context.user_data['awaiting_remove_channel'] = True
        context.user_data['channel_list'] = channels
        await update.message.reply_text(
            f"{_sx('➖')} <b>Remove Channel</b>\n\nCurrent channels:\n{lines}\n\nSend the <b>number</b> of the channel to remove:",
            parse_mode='HTML'
        )
        return

    elif text == "⏱ Check Interval":
        current = get_check_interval()
        context.user_data['awaiting_check_interval'] = True
        await update.message.reply_text(
            f"⏱ <b>Auto Check Interval</b>\n\n"
            f"বর্তমান সেটিং: প্রতি <b>{current} মিনিট</b> পর পর ইউজার চেক হয়।\n\n"
            f"নতুন মিনিট সংখ্যা পাঠান (১ - ১৪৪০):",
            parse_mode='HTML',
            reply_markup=get_link_settings_keyboard()
        )
        return

    elif text == "📝 Custom Message":
        msgs = await asyncio.to_thread(get_custom_messages)
        if msgs:
            lines = "\n\n".join(
                f"<b>#{i+1}.</b> {txt[:120]}{'…' if len(txt) > 120 else ''}"
                for i, (_, txt) in enumerate(msgs)
            )
            current_text = f"\n\n📋 <b>সেট করা Messages ({len(msgs)}):</b>\n\n{lines}"
        else:
            current_text = "\n\n<i>কোনো message সেট নেই।</i>"
        await update.message.reply_text(
            f"{_sx('✏️')} <b>Custom Message</b>{current_text}\n\n"
            "নিচের বাটন দিয়ে message যোগ বা রিমুভ করুন:",
            parse_mode='HTML',
            reply_markup=get_custom_message_keyboard()
        )
        return

    elif text == "✏️ Set Message":
        if not is_admin(username, user_id):
            return
        context.user_data['awaiting_custom_message'] = True
        await update.message.reply_text(
            f"{_sx('✏️')} <b>নতুন Message যোগ করুন</b>\n\n"
            "এখন যে message টি পাঠাবেন সেটি User Panel-এর 📋 Notice বাটনে দেখাবে।\n\n"
            "⚠️ যেকোনো ফরম্যাটে message লিখতে পারবেন।",
            parse_mode='HTML'
        )
        return

    elif text == "🗑 Remove Message":
        if not is_admin(username, user_id):
            return
        msgs = await asyncio.to_thread(get_custom_messages)
        if not msgs:
            await update.message.reply_text(
                f"{_sx('❌')} কোনো message সেট নেই।",
                parse_mode='HTML',
                reply_markup=get_custom_message_keyboard()
            )
            return
        # Show each message with its own Delete button
        keyboard = []
        msg_lines = []
        for i, (msg_id, txt) in enumerate(msgs):
            preview = txt[:100] + ("…" if len(txt) > 100 else "")
            msg_lines.append(f"<b>#{i+1}.</b> {preview}")
            keyboard.append([InlineKeyboardButton(
                f"❌ #{i+1} Delete",
                callback_data=f"del_custmsg_{msg_id}"
            )])
        body = "\n\n".join(msg_lines)
        await update.message.reply_text(
            f"{_sx('🗑')} <b>Message Delete করুন</b>\n\n{body}\n\nনির্দিষ্ট message এর পাশের বাটনে ক্লিক করুন:",
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return

    elif text == "🎁 Referral Settings":
        settings = get_referral_settings()
        status = f"{_sx('✅')} Enabled" if settings['enabled'] else f"{_sx('❌')} Disabled"
        total, unique_ref, total_rewards = get_total_referral_stats()
        msg = (
            f"{_sx('🎁')} <b>Referral Settings</b>\n\n"
            f"Status: <b>{status}</b>\n"
            f"Reward per referral: <b>{settings['reward']} {settings['label']}</b>\n\n"
            f"<b>Overall Stats:</b>\n"
            f"├ Total referrals: <b>{total}</b>\n"
            f"├ Active referrers: <b>{unique_ref}</b>\n"
            f"└ Total rewards given: <b>{total_rewards} {settings['label']}</b>"
        )
        await update.message.reply_text(msg, parse_mode='HTML', reply_markup=get_referral_manager_keyboard())
        return

    elif text == "🔛 Toggle Referral":
        settings = get_referral_settings()
        new_val = 0 if settings['enabled'] else 1
        set_referral_setting('enabled', new_val)
        settings = get_referral_settings()
        status = f"{_sx('✅')} Enabled" if settings['enabled'] else f"{_sx('❌')} Disabled"
        await update.message.reply_text(
            f"🔛 Referral program is now <b>{status}</b>.",
            parse_mode='HTML',
            reply_markup=get_referral_manager_keyboard()
        )
        return

    elif text == "🎁 Set Reward":
        context.user_data['awaiting_ref_reward'] = True
        settings = get_referral_settings()
        await update.message.reply_text(
            f"{_sx('🎁')} <b>Set Reward Amount</b>\n\nCurrent: <b>{settings['reward']} {settings['label']}</b>\n\nSend the new number of points to award per referral (e.g. <code>10</code>):",
            parse_mode='HTML'
        )
        return

    elif text == "👤 Check Balance":
        context.user_data['awaiting_ref_check_user'] = True
        await update.message.reply_text("👤 <b>Check User Balance</b>\n\nSend the user's Telegram UID:", parse_mode='HTML')
        return

    elif text == "➕ Add Balance":
        context.user_data['awaiting_ref_add_balance'] = True
        await update.message.reply_text(f"{_sx('➕')} <b>Add Balance</b>\n\nSend: <code>USER_ID AMOUNT</code>\nExample: <code>123456789 50</code>", parse_mode='HTML')
        return

    elif text == "➖ Remove Balance":
        context.user_data['awaiting_ref_remove_balance'] = True
        await update.message.reply_text(f"{_sx('➖')} <b>Remove Balance</b>\n\nSend: <code>USER_ID AMOUNT</code>\nExample: <code>123456789 20</code>", parse_mode='HTML')
        return

    elif text == "🔙 Back to Settings":
        await update.message.reply_text(
            "⚙️ <b>Settings</b>\n\nSelect a settings category:",
            parse_mode='HTML',
            reply_markup=get_settings_keyboard()
        )
        return

    elif text == "💳 Withdraw Settings":
        cfg = get_withdraw_config()
        status_icon = f"{_sx('🟢')} চালু" if cfg['enabled'] else f"{_sx('🔴')} বন্ধ"
        label = get_referral_settings()['label']
        await update.message.reply_text(
            f"💳 <b>Withdraw Settings</b>\n\n"
            f"{_sx('📌')} Status: <b>{status_icon}</b>\n"
            f"{_sx('💰')} Min Amount: <b>{cfg['min_amount']} {label}</b>\n\n"
            f"Select an option:",
            parse_mode='HTML',
            reply_markup=get_withdraw_settings_keyboard()
        )
        return

    elif text == "🔛 Toggle Withdraw":
        cfg = get_withdraw_config()
        new_state = not cfg['enabled']
        set_withdraw_enabled(new_state)
        icon = f"{_sx('🟢')} চালু" if new_state else f"{_sx('🔴')} বন্ধ"
        await update.message.reply_text(
            f"{_sx('✅')} Withdraw এখন <b>{icon}</b> করা হয়েছে!",
            parse_mode='HTML',
            reply_markup=get_withdraw_settings_keyboard()
        )
        return

    elif text == "💰 Set Min Amount":
        context.user_data['awaiting_withdraw_min_amount'] = True
        label = get_referral_settings()['label']
        await update.message.reply_text(
            f"{_sx('💰')} <b>Set Minimum Withdraw Amount</b>\n\n"
            f"ন্যূনতম কত <b>{label}</b> হলে Withdraw করা যাবে তা লিখুন:\n"
            f"Example: <code>100</code>",
            parse_mode='HTML'
        )
        return

    elif text == "📨 Set Group ID":
        cfg = get_withdraw_config()
        current = cfg['group_chat_id'] if cfg['group_chat_id'] else "Not set"
        context.user_data['awaiting_withdraw_group_id'] = True
        await update.message.reply_text(
            f"📨 <b>Set Withdraw Group ID</b>\n\n"
            f"Current Group ID: <code>{current}</code>\n\n"
            f"Send the Chat ID of the group where you want to receive Withdraw Requests:\n"
            f"(Example: <code>-1001234567890</code>)\n\n"
            f"💡 Add the bot to the group and send any message to get the Group ID.",
            parse_mode='HTML'
        )
        return

    elif text == "📊 Withdraw Stats":
        stats = get_withdraw_stats()
        await update.message.reply_text(
            f"{_sx('📊')} <b>Withdraw Statistics</b>\n\n"
            f"🗓 <b>আজকের তথ্য:</b>\n"
            f"├ 📤 মোট রিকোয়েস্ট: <b>{stats['today_count']}টি</b>\n"
            f"└ ⏳ Pending: <b>{stats['today_pending']}টি</b>\n\n"
            f"📈 <b>সর্বমোট তথ্য:</b>\n"
            f"├ 📤 মোট রিকোয়েস্ট: <b>{stats['total_count']}টি</b>\n"
            f"├ ⏳ Pending: <b>{stats['total_pending']}টি</b>\n"
            f"├ {_sx('✅')} Approved: <b>{stats['total_approved']}টি</b>\n"
            f"└ {_sx('❌')} Rejected: <b>{stats['total_rejected']}টি</b>",
            parse_mode='HTML',
            reply_markup=get_withdraw_settings_keyboard()
        )
        return

    elif text == "➕ Add Admin":
        keyboard = [
            [InlineKeyboardButton("➕ Add New Admin", callback_data="do_add_admin")],
            [InlineKeyboardButton("❌ Remove Admin", callback_data="do_remove_admin")]
        ]
        await update.message.reply_text(
            f"{_sx('👑')} <b>Admin Manager</b>\n\n"
            f"{_sx('➕')} Add New Admin\n"
            f"{_sx('❌')} Remove Admin\n\n"
            "Select an option:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )

    elif text == "📬 Message Hub":
        all_count    = len(get_all_users())
        banned_count = len(get_banned_users())
        kb = ReplyKeyboardMarkup(
            [
                [KeyboardButton(f"📢 Broadcast All ({all_count})"),
                 KeyboardButton("👤 One User")],
                [KeyboardButton("🔙 Back to Admin")],
            ],
            resize_keyboard=True
        )
        await update.message.reply_text(
            f"📬 <b>Message Hub</b>\n\n"
            f"📢 <b>Broadcast All:</b> {all_count} জন\n"
            f"👤 <b>One User:</b> Direct Message\n\n"
            "Broadcast target select করুন 👇",
            parse_mode='HTML',
            reply_markup=kb
        )

    elif text == "🚩 Country Manager":
        await update.message.reply_text(
            f"{_sx('🚩')} <b>Country Manager</b>\n\n"
            f"{_sx('➕')} Add Numbers — Upload new numbers\n"
            f"{_sx('🌐')} Add Country — Register a country\n"
            f"{_sx('📱')} Add Service — Create a service\n"
            f"{_sx('🗺')} Service Map — Link countries to services\n"
            f"💰 OTP Rewards — Per-country OTP balance\n"
            f"{_sx('🔄')} Reset Number — Reset usage\n\n"
            "Select an option:",
            parse_mode='HTML',
            reply_markup=get_country_manager_keyboard()
        )
        return

    elif text == "💰 OTP Rewards":
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized.", parse_mode='HTML')
            return
        rewards = get_all_country_otp_rewards()
        countries = get_countries()
        lines = []
        for _, cname in countries:
            plain = _clean_name(cname)
            amt = rewards.get(plain.lower(), 0)
            flag = get_animated_flag_html(plain)
            if amt:
                lines.append(f"{flag} <code>{plain}</code>: <code>{amt}৳</code>")
            else:
                lines.append(f"{flag} <code>{plain}</code>: <i>not set</i>")
        reward_list = "\n".join(lines) if lines else "<i>No countries added yet.</i>"
        context.user_data['awaiting_country_otp_reward'] = True
        await update.message.reply_text(
            f"💰 <b>OTP Rewards</b>\n\n"
            f"প্রতিটি দেশের OTP পেলে ইউজার কত ব্যালেন্স পাবে সেটা এখানে সেট করুন।\n\n"
            f"<b>বর্তমান রিওয়ার্ড:</b>\n{reward_list}\n\n"
            f"📝 <b>সেট করতে লিখুন:</b>\n"
            f"<code>CountryName Amount</code>\n"
            f"উদাহরণ: <code>Bangladesh 0.20</code>\n\n"
            f"একাধিক দেশ একসাথে:\n"
            f"<code>Bangladesh 0.20\nMyanmar 0.15\nIndia 0.10</code>",
            parse_mode='HTML',
            reply_markup=get_country_manager_keyboard()
        )
        return

    elif text == "📱 Add Service":
        services = get_services_with_emoji()
        if services:
            svc_lines = "\n".join([
                f"{_svc_animated_tag(name, ceid)} <code>{name}</code>"
                for _, name, ceid in services
            ])
            svc_section = f"\n\n📋 <b>Existing Services:</b>\n{svc_lines}"
        else:
            svc_section = "\n\n📋 <b>No services added yet.</b>"
        context.user_data['awaiting_new_service'] = True
        context.user_data.pop('pending_service_emoji', None)
        context.user_data.pop('pending_service_custom_emoji_id', None)
        await update.message.reply_text(
            f"<b>Add Service</b>\n\n"
            f"Send the name of the new service:"
            f"{svc_section}\n\n"
            f"🗑️ <b>Remove a service:</b>\n"
            f"Type <code>remove</code> followed by service name.\n"
            f"Example: <code>remove WhatsApp</code>",
            parse_mode='HTML'
        )
        return

    elif text == "🗺 Service Map":
        service_map = get_service_map()
        if not service_map:
            await update.message.reply_text(f"{_sx('❌')} No services found. Add a service first using Add Service.", parse_mode='HTML')
            return
        all_countries = get_countries()
        # Build country line with flag + name + number count
        country_lines_list = []
        for cid, cname in all_countries:
            plain = _clean_name(cname)
            flag = get_animated_flag_html(plain)
            total, available = get_numbers_count_by_country(cid)
            country_lines_list.append(f"{flag} <code>{plain}</code> — {int(available or 0)}/{int(total or 0)} numbers")
        country_lines = "\n".join(country_lines_list) if country_lines_list else "<i>None</i>"
        msg = "<b>Service Map</b>\n\n"
        for sid, sname, countries in service_map:
            if countries:
                clist = "\n".join([f"  {get_animated_flag_html(_clean_name(c))} <code>{_clean_name(c)}</code>" for c in countries])
            else:
                clist = "  <i>No countries</i>"
            msg += f"<code>{sname}</code>\n{clist}\n\n"
        msg += f"─────────────────\n📋 <b>All Countries:</b>\n{country_lines}\n\n"
        msg += "🔗 Link: ServiceName CountryName\nExample: <code>Telegram Bangladesh</code>\n\n"
        msg += "🔗 <code>unmap</code> ServiceName CountryName\nExample: <code>unmap WhatsApp Bangladesh</code>"
        context.user_data['awaiting_service_map'] = True
        await update.message.reply_text(msg, parse_mode='HTML')
        return

    elif text == "📡 SMS Panels":
        panels = get_sms_panels()
        if not panels:
            msg = (
                f"{_sx('📡')} <b>SMS Panels</b>\n\n"
                "কোনো panel যোগ করা নেই।\n\n"
                "➕ <b>Add SMS Panel</b> বাটন চাপুন নতুন panel যোগ করতে।"
            )
        else:
            lines = [f"{_sx('📡')} <b>SMS Panels</b> ({len(panels)} টি)\n"]
            for p in panels:
                icon = "🟢" if p.get("enabled", True) else "🔴"
                lines.append(f"{icon} <b>{p['name']}</b>\n"
                             f"   🔗 Login: <code>{p['login_url']}</code>\n"
                             f"   📄 Page: <code>{p['message_url']}</code>\n"
                             f"   👤 User: <code>{p['username']}</code>\n")
            msg = "\n".join(lines)
        await update.message.reply_text(
            msg, parse_mode='HTML',
            reply_markup=get_sms_panels_keyboard()
        )
        return

    elif text == "📊 Panel Statistics":
        panels = get_sms_panels()
        if not panels:
            await update.message.reply_text(
                f"📊 <b>Panel Statistics</b>\n\nকোনো panel যোগ করা নেই।",
                parse_mode='HTML',
                reply_markup=get_sms_panels_keyboard()
            )
            return
        total = len(panels)
        enabled = sum(1 for p in panels if p.get("enabled", True))
        disabled = total - enabled
        lines = [
            f"📊 <b>Panel Statistics</b>\n",
            f"📡 মোট Panel: <b>{total}</b>",
            f"🟢 Active: <b>{enabled}</b>   🔴 Inactive: <b>{disabled}</b>\n",
        ]
        for p in panels:
            pid = p['id']
            icon = "🟢" if p.get("enabled", True) else "🔴"
            counts = get_panel_send_counts(pid)
            today = get_today_panel_message_count(pid)
            fwd_chat = get_panel_forward_chat(pid) or get_global_forward_chat()
            fwd_info = f"<code>{fwd_chat}</code>" if fwd_chat else "সেট করা নেই"
            lines.append(
                f"{'─'*28}\n"
                f"{icon} <b>{p['name']}</b>\n"
                f"   🔗 Login URL: <code>{p['login_url']}</code>\n"
                f"   👤 Username: <code>{p['username']}</code>\n"
                f"   📬 আজকের OTP (24h): <b>{today}</b>\n"
                f"   👤 User Panel এ পাঠানো: <b>{counts['user']}</b>\n"
                f"   📢 Group এ পাঠানো: <b>{counts['group']}</b>\n"
                f"   📤 Forward Chat: {fwd_info}"
            )
        await update.message.reply_text(
            "\n".join(lines),
            parse_mode='HTML',
            reply_markup=get_sms_panels_keyboard()
        )
        return

    elif text == "➕ Add SMS Panel":
        context.user_data['awaiting_panel_name'] = True
        await update.message.reply_text(
            f"{_sx('📡')} <b>New SMS Panel</b>\n\n"
            "Step 1/11 — Panel এর নাম দিন:\n"
            "<i>উদাহরণ: Number Panel</i>\n\n"
            "<i>বাতিল করতে /cancel লিখুন।</i>",
            parse_mode='HTML',
            reply_markup=ReplyKeyboardRemove()
        )
        return

    elif text.startswith("📡 ✅ ") or text.startswith("📡 ❌ "):
        panel_name = text[4:].strip()
        panels = get_sms_panels()
        matched = next((p for p in panels if p['name'] == panel_name), None)
        if not matched:
            await update.message.reply_text(f"{_sx('❌')} Panel পাওয়া যায়নি।", parse_mode='HTML')
            return
        context.user_data['viewing_panel_id'] = matched['id']
        icon = "🟢 চালু" if matched.get("enabled", True) else "🔴 বন্ধ"
        await update.message.reply_text(
            f"{_sx('📡')} <b>{matched['name']}</b>\n\n"
            f"স্ট্যাটাস: {icon}\n"
            f"🔗 Login URL: <code>{matched['login_url']}</code>\n"
            f"📄 Message URL: <code>{matched['message_url']}</code>\n"
            f"👤 Username: <code>{matched['username']}</code>",
            parse_mode='HTML',
            reply_markup=get_panel_detail_keyboard(matched)
        )
        return

    elif text in ("🔴 বন্ধ করুন", "🟢 চালু করুন"):
        panel_id = context.user_data.get('viewing_panel_id')
        if not panel_id:
            await update.message.reply_text(f"{_sx('❌')} কোনো panel নির্বাচন করা হয়নি।", parse_mode='HTML',
                                            reply_markup=get_sms_panels_keyboard())
            return
        new_state = toggle_sms_panel(panel_id)
        if new_state is None:
            await update.message.reply_text(f"{_sx('❌')} Panel পাওয়া যায়নি।", parse_mode='HTML')
            return
        panel = get_sms_panel(panel_id)
        if panel:
            try:
                if new_state:
                    start_panel_monitor(context.bot, panel)
                else:
                    stop_panel_monitor(panel_id)
            except Exception as e:
                logger.error(f"Panel toggle error: {e}")
        icon = "🟢 চালু" if new_state else "🔴 বন্ধ"
        await update.message.reply_text(
            f"{_sx('📡')} <b>{panel['name']}</b>\n\n"
            f"স্ট্যাটাস: {icon}\n"
            f"🔗 Login URL: <code>{panel['login_url']}</code>\n"
            f"📄 Message URL: <code>{panel['message_url']}</code>\n"
            f"👤 Username: <code>{panel['username']}</code>",
            parse_mode='HTML',
            reply_markup=get_panel_detail_keyboard(panel)
        )
        return

    elif text == "📨 Last Message":
        panel_id = context.user_data.get('viewing_panel_id')
        if not panel_id:
            await update.message.reply_text(f"{_sx('❌')} কোনো panel নির্বাচন করা হয়নি।", parse_mode='HTML',
                                            reply_markup=get_sms_panels_keyboard())
            return
        panel = get_sms_panel(panel_id)
        if not panel:
            await update.message.reply_text(f"{_sx('❌')} Panel পাওয়া যায়নি।", parse_mode='HTML')
            return
        icon = "🟢" if panel.get("enabled", True) else "🔴"
        loading_msg = await update.message.reply_text(
            f"⏳ <b>{panel['name']}</b> — panel থেকে চেক করা হচ্ছে...",
            parse_mode='HTML'
        )
        msgs = await asyncio.to_thread(_fetch_panel_last_messages, panel, 3)
        today_count = await asyncio.to_thread(get_today_panel_message_count, panel_id)
        try:
            await loading_msg.delete()
        except Exception:
            pass
        if isinstance(msgs, dict) and msgs.get('error'):
            err = msgs['error']
            if err == 'login_failed':
                err_detail = "Panel এ login করা হয়নি। Panel চালু করুন।"
            else:
                err_detail = "Session expire হয়ে গেছে। Panel বন্ধ করে আবার চালু করুন।"
            msg_text = (
                f"❌ <b>{panel['name']} — Login হয়নি</b>\n\n"
                f"{err_detail}\n\n"
                f"📊 Last 24h Total: <b>{today_count}</b>\n"
                f"{icon} Status: {'Active' if panel.get('enabled', True) else 'Inactive'}"
            )
        elif msgs:
            msg_text = _build_last_msgs_text(panel['name'], msgs, today_count, icon)
        else:
            msg_text = (
                f"📨 <b>{panel['name']} — Last Messages</b>\n\n"
                f"<i>Panel থেকে কোনো message পাওয়া যায়নি।</i>\n\n"
                f"📊 Last 24h Total: <b>{today_count}</b>\n"
                f"{icon} Status: {'Active' if panel.get('enabled', True) else 'Inactive'}"
            )
        refresh_markup = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 Refresh", callback_data=f"sms_panel_refresh_msg_{panel_id}")]
        ])
        await update.message.reply_text(msg_text, parse_mode='HTML', reply_markup=refresh_markup)
        return

    elif text == "📤 Group এ পাঠান":
        panel_id = context.user_data.get('viewing_panel_id')
        if not panel_id:
            await update.message.reply_text(
                f"{_sx('❌')} কোনো panel নির্বাচন করা হয়নি।",
                parse_mode='HTML',
                reply_markup=get_sms_panels_keyboard()
            )
            return
        panel = get_sms_panel(panel_id)
        if not panel:
            await update.message.reply_text(f"{_sx('❌')} Panel পাওয়া যায়নি।", parse_mode='HTML')
            return

        # Determine forward target (per-panel first, then global)
        fwd_chat = get_panel_forward_chat(panel_id) or get_global_forward_chat()
        if not fwd_chat:
            await update.message.reply_text(
                f"⚠️ কোনো Group/Chat সেট করা নেই।\n"
                f"প্রথমে <b>Forward Chat ID</b> সেট করুন।",
                parse_mode='HTML'
            )
            return

        loading_msg = await update.message.reply_text(
            f"⏳ <b>{panel['name']}</b> — Last Message নেওয়া হচ্ছে...",
            parse_mode='HTML'
        )
        last = await asyncio.to_thread(_fetch_panel_last_message, panel)
        try:
            await loading_msg.delete()
        except Exception:
            pass

        if not last or (isinstance(last, dict) and last.get('error')):
            await update.message.reply_text(
                f"❌ <b>Group এ পাঠানো সম্ভব হয়নি।</b>\n\n"
                f"Panel থেকে কোনো message পাওয়া যায়নি অথবা login error।",
                parse_mode='HTML'
            )
            return

        # Build message text
        number = last['number']
        message = last.get('message', '')
        country_name = last.get('country', '') or await asyncio.to_thread(get_country_name_by_number, number)
        otp = await asyncio.to_thread(extract_otp_from_message, message) if message else ""
        service = extract_service_from_message(message)
        sender = last.get('sender', '')
        otp_line = f"OTP: <code>{otp}</code>\n" if otp else ""
        if service:
            service_line = f"Service: {service}\n"
        elif sender:
            service_line = f"Sender: <code>{sender}</code>\n"
        else:
            service_line = ""
        group_msg_text, group_msg_markup = _build_group_forward(
            number, message, country_name, otp, service, sender
        )

        # Send to group
        sent_ok = False
        try:
            await context.bot.send_message(
                chat_id=int(fwd_chat),
                text=group_msg_text,
                parse_mode='HTML',
                reply_markup=group_msg_markup
            )
            sent_ok = True
        except Exception as e:
            logger.error(f"Group send error: {e}")

        if sent_ok:
            await update.message.reply_text(
                f"✅ <b>Group এ পাঠানো সফল হয়েছে!</b>\n\n"
                f"📡 Panel: <b>{panel['name']}</b>\n"
                f"💬 Chat ID: <code>{fwd_chat}</code>",
                parse_mode='HTML'
            )
            # Notify all admins
            admin_ids = get_all_admin_ids()
            sender_name = update.effective_user.full_name or "Admin"
            notify_text = (
                f"🔔 <b>Group এ Message পাঠানো হয়েছে</b>\n\n"
                f"👤 পাঠিয়েছেন: <b>{sender_name}</b> (<code>{user_id}</code>)\n"
                f"📡 Panel: <b>{panel['name']}</b>\n"
                f"💬 Chat ID: <code>{fwd_chat}</code>\n\n"
                f"📨 Message:\n"
                f"Number: <code>{number}</code>\n"
                f"{service_line}"
                f"{otp_line}"
                f"Content: <code>{message}</code>"
            )
            for aid in admin_ids:
                if aid == user_id:
                    continue  # নিজেকে notify করার দরকার নেই
                try:
                    await context.bot.send_message(chat_id=aid, text=notify_text, parse_mode='HTML')
                except Exception:
                    pass
        else:
            await update.message.reply_text(
                f"❌ <b>Group এ পাঠাতে ব্যর্থ হয়েছে।</b>\n\n"
                f"Chat ID <code>{fwd_chat}</code> এ message পাঠানো সম্ভব হয়নি।\n"
                f"বট কি এই group এর member?",
                parse_mode='HTML'
            )
        return

    elif text == "✏️ Username পরিবর্তন":
        panel_id = context.user_data.get('viewing_panel_id')
        if not panel_id:
            await update.message.reply_text(f"{_sx('❌')} কোনো panel নির্বাচন করা হয়নি।", parse_mode='HTML',
                                            reply_markup=get_sms_panels_keyboard())
            return
        panel = get_sms_panel(panel_id)
        if not panel:
            await update.message.reply_text(f"{_sx('❌')} Panel পাওয়া যায়নি।", parse_mode='HTML')
            return
        context.user_data['awaiting_panel_edit_username'] = True
        await update.message.reply_text(
            f"{_sx('✏️')} <b>Username পরিবর্তন</b>\n\n"
            f"Panel: <b>{panel['name']}</b>\n"
            f"বর্তমান Username: <code>{panel['username']}</code>\n\n"
            "নতুন username টাইপ করুন:",
            parse_mode='HTML'
        )
        return

    elif text == "🔑 Password পরিবর্তন":
        panel_id = context.user_data.get('viewing_panel_id')
        if not panel_id:
            await update.message.reply_text(f"{_sx('❌')} কোনো panel নির্বাচন করা হয়নি।", parse_mode='HTML',
                                            reply_markup=get_sms_panels_keyboard())
            return
        panel = get_sms_panel(panel_id)
        if not panel:
            await update.message.reply_text(f"{_sx('❌')} Panel পাওয়া যায়নি।", parse_mode='HTML')
            return
        context.user_data['awaiting_panel_edit_password'] = True
        await update.message.reply_text(
            f"{_sx('🔑')} <b>Password পরিবর্তন</b>\n\n"
            f"Panel: <b>{panel['name']}</b>\n\n"
            "নতুন password টাইপ করুন:",
            parse_mode='HTML'
        )
        return

    elif text == "🗑️ Delete Panel":
        panel_id = context.user_data.get('viewing_panel_id')
        if not panel_id:
            await update.message.reply_text(f"{_sx('❌')} কোনো panel নির্বাচন করা হয়নি।", parse_mode='HTML',
                                            reply_markup=get_sms_panels_keyboard())
            return
        panel = get_sms_panel(panel_id)
        name = panel['name'] if panel else panel_id
        delete_sms_panel(panel_id)
        try:
            stop_panel_monitor(panel_id)
        except Exception:
            pass
        context.user_data.pop('viewing_panel_id', None)
        await update.message.reply_text(
            f"{_sx('✅')} Panel <b>{name}</b> মুছে ফেলা হয়েছে।",
            parse_mode='HTML',
            reply_markup=get_sms_panels_keyboard()
        )
        return

    elif text == "📤 Group OTP Forward":
        gfwd = get_global_forward_chat()
        if gfwd:
            gfwd_name = gfwd
            try:
                chat_info = await context.bot.get_chat(int(gfwd))
                gfwd_name = chat_info.title or chat_info.username or gfwd
            except Exception:
                pass
            msg = (
                f"📤 <b>Group OTP Forward</b>\n\n"
                f"✅ বর্তমানে সেট করা Group/Channel:\n"
                f"📌 <b>{gfwd_name}</b>\n"
                f"<code>{gfwd}</code>\n\n"
                f"সকল panel এর নতুন OTP/SMS স্বয়ংক্রিয়ভাবে এই group-এ পাঠানো হচ্ছে।"
            )
        else:
            msg = (
                f"📤 <b>Group OTP Forward</b>\n\n"
                f"এখনো কোনো group/channel সেট করা হয়নি।\n\n"
                f"নিচের বাটন থেকে একটি Group বা Channel সেট করুন।"
            )
        keyboard = ReplyKeyboardMarkup(
            [
                [KeyboardButton("➕ Group Add"), KeyboardButton("🗑 Group Remove")],
                [KeyboardButton("🔙 Back to Panels")],
            ],
            resize_keyboard=True
        )
        await update.message.reply_text(msg, parse_mode='HTML', reply_markup=keyboard)
        return

    elif text == "➕ Group Add":
        context.user_data['awaiting_global_forward_chat_id'] = True
        await update.message.reply_text(
            f"📤 <b>Group OTP Forward — Group Add</b>\n\n"
            f"যে Group বা Channel-এ OTP পাঠাতে চান সেটির <b>Chat ID</b> দিন।\n\n"
            f"<i>উদাহরণ: -1001234567890</i>\n\n"
            f"💡 Chat ID পেতে @userinfobot বা @RawDataBot ব্যবহার করুন।\n\n"
            f"<i>বাতিল করতে /cancel লিখুন।</i>",
            parse_mode='HTML'
        )
        return

    elif text == "🗑 Group Remove":
        remove_global_forward_chat()
        keyboard = ReplyKeyboardMarkup(
            [
                [KeyboardButton("➕ Group Add"), KeyboardButton("🗑 Group Remove")],
                [KeyboardButton("🔙 Back to Panels")],
            ],
            resize_keyboard=True
        )
        await update.message.reply_text(
            f"📤 <b>Group OTP Forward</b>\n\n"
            f"✅ Group Remove করা হয়েছে।\n\n"
            f"এখনো কোনো group/channel সেট করা হয়নি।",
            parse_mode='HTML',
            reply_markup=keyboard
        )
        return

    elif text == "🔙 Back to Panels":
        context.user_data.pop('viewing_panel_id', None)
        await update.message.reply_text(
            f"{_sx('📡')} <b>SMS Panels</b>",
            parse_mode='HTML',
            reply_markup=get_sms_panels_keyboard()
        )
        return

    elif text == "🔄 Reset Number":
        stats = get_country_stats()
        countries_with_numbers = [(name, int(total or 0)) for name, total, available in stats if int(total or 0) > 0]
        if not countries_with_numbers:
            await update.message.reply_text(f"{_sx('❌')} No countries with numbers found.", parse_mode='HTML')
            return
        country_lines = "\n".join([
            f"{get_animated_flag_html(_clean_name(name))} <code>{name}</code> — {total} numbers"
            for name, total in countries_with_numbers
        ])
        context.user_data['awaiting_reset_country'] = True
        await update.message.reply_text(
            f"<b>Reset Numbers</b>\n\nType the country name to remove all its numbers:\n\n{country_lines}",
            parse_mode='HTML'
        )
        return

    elif text == "🔙 Back to Admin":
        # Clear all pending states to avoid conflicts
        context.user_data.pop('awaiting_new_country', None)
        context.user_data.pop('awaiting_add_numbers_country', None)
        context.user_data.pop('awaiting_numbers_file', None)
        context.user_data.pop('current_country_name', None)
        context.user_data.pop('awaiting_number_delete', None)
        context.user_data.pop('awaiting_new_admin', None)
        context.user_data.pop('awaiting_broadcast', None)
        context.user_data.pop('awaiting_reset_country', None)
        context.user_data.pop('awaiting_new_service', None)
        context.user_data.pop('awaiting_service_map', None)
        context.user_data.pop('awaiting_country_otp_reward', None)
        context.user_data.pop('country_otp_reward_name', None)
        await update.message.reply_text(
            f"{_sx('🤖')} <b>Admin Panel</b>",
            parse_mode='HTML',
            reply_markup=get_admin_keyboard()
        )
        return

# Handle user button clicks
async def show_balance_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show balance, referral, and withdraw panel to user."""
    user = update.effective_user

    # Cache bot username — avoid repeated Telegram API calls
    bot_username = context.bot_data.get('_bot_username')
    if not bot_username:
        bot_info = await context.bot.get_me()
        bot_username = bot_info.username
        context.bot_data['_bot_username'] = bot_username

    # Run all DB calls concurrently
    settings, bal, count, history = await asyncio.gather(
        asyncio.to_thread(get_referral_settings),
        asyncio.to_thread(get_user_balance_data, user.id),
        asyncio.to_thread(get_user_referral_count, user.id),
        asyncio.to_thread(get_user_withdraw_history, user.id, 3),
    )
    ref_link = f"https://t.me/{bot_username}?start=ref_{user.id}"
    label = settings['label']
    reward = settings['reward']

    msg = (
        f"{_sx('💰')} <b>My Balance</b>\n\n"
        f"├ {_sx('💵')} Balance: <b>{bal['balance']} {label}</b>\n"
        f"└ 🏆 Total Earned: <b>{bal['total_earned']} {label}</b>\n\n"
    )

    if settings['enabled']:
        msg += (
            f"{_sx('🎁')} <b>Referral Program</b>\n"
            f"{_sx('👥')} Referred: <b>{count} people</b>\n"
            f"💡 Earn <b>{reward} {label}</b> per referral!\n\n"
            f"{_sx('🔗')} <b>Your Referral Link:</b>\n"
            f"<code>{ref_link}</code>\n\n"
        )

    if history:
        msg += "📜 <b>Recent Withdrawals:</b>\n"
        for h in history:
            _, wtype, waddr, hamount, hlabel, hstatus, _ = h
            icon = _sx('✅') if hstatus == "approved" else (_sx('❌') if hstatus == "rejected" else "⏳")
            msg += f"{icon} {wtype} — {hamount} {hlabel} — {hstatus}\n"
        msg += "\n"

    keyboard = [
        [InlineKeyboardButton("💳 Withdraw", callback_data="withdraw_start")]
    ]
    if settings['enabled']:
        share_url = f"https://t.me/share/url?url={ref_link}&text=Join+this+bot!"
        keyboard.append([InlineKeyboardButton("🔗 Share Referral Link", url=share_url)])
    reply_markup = InlineKeyboardMarkup(keyboard)

    if update.callback_query:
        await update.callback_query.edit_message_text(msg, parse_mode='HTML', reply_markup=reply_markup)
    else:
        await update.message.reply_text(msg, parse_mode='HTML', reply_markup=reply_markup)


async def _notify_flagged_once(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    """Send join-channel message at most once per check cycle for flagged users.
    Returns True if user is flagged (caller should stop processing).
    Returns False if user is not flagged (caller continues normally)."""
    user_id = update.effective_user.id
    flagged: set = context.bot_data.get('join_flagged', set())
    if user_id not in flagged:
        return False
    notified: set = context.bot_data.setdefault('join_notified', set())
    if user_id in notified:
        return True  # already notified this cycle — block silently
    # First notification this cycle
    notified.add(user_id)
    channels = await asyncio.to_thread(get_join_channels)
    lines = "\n".join([f"👉 <a href='{l}'>{t}</a>" for _, t, l in channels])
    msg = (
        "⚠️ <b>আপনি নিচের চ্যানেলে জয়েন নেই:</b>\n\n"
        f"{lines}\n\n"
        "জয়েন করার পর ✅ <b>Verify</b> বাটনে ক্লিক করুন।"
    )
    keyboard = [[InlineKeyboardButton(f"📢 {t}", url=l)] for _, t, l in channels]
    keyboard.append([InlineKeyboardButton("✅ Verify", callback_data="verify_user")])
    await update.message.reply_text(
        msg, parse_mode='HTML',
        reply_markup=InlineKeyboardMarkup(keyboard),
        disable_web_page_preview=True
    )
    return True

async def handle_user_button_click(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    user_id = update.effective_user.id

    # Track activity in background
    asyncio.create_task(asyncio.to_thread(update_user_activity, user_id))

    # Universal join guard — applies to every user panel button
    if await _notify_flagged_once(update, context):
        return
    if not await enforce_join(update, context):
        return

    if text in ("✅ Get Numbers", "🔢 Get Numbers", "Get Numbers"):
        await show_services(update, context)

    elif text in ("💰 Balance", "💎 Balance", "💰 My Balance"):
        await show_balance_panel(update, context)

    elif text == "📊 View Stats":
        await show_stats(update, context)

    elif text in ("🌍 Available country", "🌐 Available country", "🌍 Available Country"):
        countries = await asyncio.to_thread(get_countries)
        if not countries:
            await update.message.reply_text(f"{_sx('❌')} No countries available.", parse_mode='HTML')
            return
        counts = await asyncio.gather(*[
            asyncio.to_thread(get_numbers_count_by_country, cid)
            for cid, _ in countries
        ])
        rewards = await asyncio.to_thread(get_all_country_otp_rewards)
        lines = []
        for (cid, name), (_, available) in zip(countries, counts):
            if available > 0:
                plain = _clean_name(name)
                flag = get_animated_flag_html(plain)
                bonus = rewards.get(plain.lower(), 0)
                bonus_text = f" | 💰 <b>{bonus}৳</b> bonus" if bonus else ""
                lines.append(f"{flag} <b>{plain}</b> — <b>{available}</b>{bonus_text}")
        if not lines:
            await update.message.reply_text(f"{_sx('❌')} No numbers available in any country.", parse_mode='HTML')
            return
        msg = "📋 <b>Available Countries</b>\n\n" + "\n".join(lines)
        await update.message.reply_text(msg, parse_mode='HTML')

    elif text in ("📋 Notice", "🔔 Notice", "📢 Notice"):
        msg = await asyncio.to_thread(get_custom_message)
        if not msg:
            await update.message.reply_text(
                f"{_sx('🔔')} <b>Notice</b>\n\n<i>এখনো কোনো notice সেট করা হয়নি।</i>",
                parse_mode='HTML'
            )
            return
        await update.message.reply_text(
            f"{_sx('🔔')} <b>Notice</b>\n\n{msg}",
            parse_mode='HTML'
        )

async def show_broadcast_preview(update: Update, context: ContextTypes.DEFAULT_TYPE, msg):
    """Show a preview of the broadcast message with Confirm/Cancel buttons."""
    context.user_data['broadcast_msg_id'] = msg.message_id
    context.user_data['broadcast_chat_id'] = msg.chat_id
    context.user_data['broadcast_is_forwarded'] = bool(msg.forward_origin) if hasattr(msg, 'forward_origin') else False

    target = context.user_data.get('broadcast_target', 'all')
    if target == "banned":
        raw_list     = get_banned_users()
        target_label = "🚫 Banned Users"
    else:
        raw_list     = get_all_users()
        target_label = "📣 All Users"
    user_count = len([u for u in raw_list if u])

    # Show the actual message as preview
    if context.user_data['broadcast_is_forwarded']:
        await context.bot.forward_message(
            chat_id=update.effective_chat.id,
            from_chat_id=msg.chat_id,
            message_id=msg.message_id
        )
    else:
        await msg.copy_to(chat_id=update.effective_chat.id)

    keyboard = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Confirm Broadcast", callback_data="broadcast_confirm"),
            InlineKeyboardButton("❌ Cancel", callback_data="broadcast_cancel")
        ]
    ])
    await update.message.reply_text(
        f"{_sx('📊')} মোট <b>{user_count}</b> জনকে পাঠানো হবে ({target_label})।\nনিশ্চিত করুন:",
        parse_mode='HTML',
        reply_markup=keyboard
    )


def _progress_bar(done: int, total: int, width: int = 14) -> str:
    filled = int(width * done / total) if total else 0
    bar = "█" * filled + "░" * (width - filled)
    pct = int(100 * done / total) if total else 0
    return f"[{bar}] {pct}%"


async def _fast_broadcast(bot, user_ids: list, from_chat_id: int, message_id: int,
                           is_fwd: bool, progress_msg=None) -> tuple:
    """Send to all users concurrently with FloodWait retry — no messages dropped."""
    from telegram.error import RetryAfter, TimedOut, NetworkError
    sem = asyncio.Semaphore(29)
    total = len(user_ids)
    success_count = 0
    fail_count = 0
    last_edit = [0.0]

    async def _send_with_retry(uid: int):
        nonlocal success_count, fail_count
        max_retries = 5
        for attempt in range(max_retries):
            try:
                async with sem:
                    if is_fwd:
                        await bot.forward_message(chat_id=uid, from_chat_id=from_chat_id, message_id=message_id)
                    else:
                        await bot.copy_message(chat_id=uid, from_chat_id=from_chat_id, message_id=message_id)
                success_count += 1
                break
            except RetryAfter as e:
                wait = e.retry_after + 1
                await asyncio.sleep(wait)
            except (TimedOut, NetworkError):
                await asyncio.sleep(2 * (attempt + 1))
            except Exception:
                fail_count += 1
                break
        else:
            fail_count += 1

        if progress_msg:
            now = asyncio.get_event_loop().time()
            done = success_count + fail_count
            if now - last_edit[0] >= 2.0 or done == total:
                last_edit[0] = now
                try:
                    await progress_msg.edit_text(
                        f"{_sx('📢')} <b>Broadcasting...</b>\n\n"
                        f"{_progress_bar(done, total)}\n"
                        f"📤 Sent: <b>{done} / {total}</b>\n"
                        f"✅ Success: <b>{success_count}</b>  "
                        f"❌ Failed: <b>{fail_count}</b>",
                        parse_mode='HTML'
                    )
                except Exception:
                    pass

    await asyncio.gather(*[_send_with_retry(uid) for uid in user_ids])
    return success_count, fail_count


async def run_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE, msg):
    """Broadcast any message type to all users. Forwards if forwarded, copies otherwise."""
    raw_users = get_all_users()
    user_ids = []
    for u in raw_users:
        uid = u.get('user_id') if isinstance(u, dict) else u
        if uid:
            try:
                user_ids.append(int(uid))
            except (ValueError, TypeError):
                pass

    is_forwarded = bool(msg.forward_origin) if hasattr(msg, 'forward_origin') else False
    total = len(user_ids)

    progress_msg = await update.message.reply_text(
        f"{_sx('📢')} <b>Broadcast শুরু হচ্ছে...</b>\n\n"
        f"{_progress_bar(0, total)}\n"
        f"📤 Sent: <b>0 / {total}</b>",
        parse_mode='HTML'
    )

    success_count, fail_count = await _fast_broadcast(
        bot=context.bot,
        user_ids=user_ids,
        from_chat_id=msg.chat_id,
        message_id=msg.message_id,
        is_fwd=is_forwarded,
        progress_msg=progress_msg
    )

    await update.message.reply_text(
        f"{_sx('📊')} <b>Broadcast Completed!</b>\n\n"
        f"✅ Success: <b>{success_count}</b>\n"
        f"❌ Failed: <b>{fail_count}</b>\n"
        f"👥 Total: <b>{total}</b>",
        parse_mode='HTML'
    )


async def handle_broadcast_media(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle non-text media messages for broadcast."""
    username = update.effective_user.username
    user_id = update.effective_user.id

    # ── Sticker sent during "Add Service" flow ─────────────────────────────────
    sticker = update.message.sticker if update.message else None
    if sticker and context.user_data.get('awaiting_new_service') and is_admin(username, user_id):
        emoji_char = sticker.emoji or "🔹"
        custom_emoji_id = sticker.custom_emoji_id or ""
        context.user_data['pending_service_emoji'] = emoji_char
        context.user_data['pending_service_custom_emoji_id'] = custom_emoji_id
        # Build display string for confirmation
        if custom_emoji_id:
            display = f'<tg-emoji emoji-id="{custom_emoji_id}">{emoji_char}</tg-emoji>'
        else:
            display = emoji_char
        await update.message.reply_text(
            f"✅ Emoji saved: {display}\n\nNow type the <b>service name</b>:",
            parse_mode='HTML'
        )
        return
    # ──────────────────────────────────────────────────────────────────────────

    # ── Direct Message cancel button ──────────────────────────────────────────
    if text == "❌ বাতিল করুন" and is_admin(username, user_id):
        in_direct_flow = (
            context.user_data.get('awaiting_direct_uid') or
            context.user_data.get('awaiting_direct_message')
        )
        for _k in ('awaiting_direct_uid', 'awaiting_direct_message', 'direct_msg_uid'):
            context.user_data.pop(_k, None)
        if in_direct_flow:
            await update.message.reply_text(
                "❌ <b>Direct Message বাতিল হয়েছে।</b>",
                parse_mode='HTML',
                reply_markup=get_admin_keyboard()
            )
        return

    # ── Direct Message by UID — media/sticker/voice/video/etc. ───────────────
    if context.user_data.get('awaiting_direct_message') and is_admin(username, user_id):
        target_uid = context.user_data.pop('direct_msg_uid', None)
        context.user_data['awaiting_direct_message'] = False
        if not target_uid:
            await update.message.reply_text("❌ UID not found. আবার চেষ্টা করুন।", parse_mode='HTML')
            return
        try:
            await update.message.copy_to(chat_id=target_uid)
            await update.message.reply_text(
                f"✅ <b>Sent!</b> UID <code>{target_uid}</code> কে message পাঠানো হয়েছে।",
                parse_mode='HTML'
            )
        except Exception as e:
            await update.message.reply_text(
                f"❌ <b>Failed:</b> <code>{target_uid}</code>\n<i>{e}</i>",
                parse_mode='HTML'
            )
        return
    # ─────────────────────────────────────────────────────────────────────────

    if not is_admin(username, user_id):
        return
    if not context.user_data.get('awaiting_broadcast'):
        return
    context.user_data['awaiting_broadcast'] = False
    await run_broadcast(update, context, update.message)


# Handle file uploads for adding numbers
async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    username = update.effective_user.username
    user_id = update.effective_user.id

    # If not awaiting a numbers file, pass to broadcast handler
    if 'awaiting_numbers_file' not in context.user_data or 'current_country_name' not in context.user_data:
        if is_admin(username, user_id) and context.user_data.get('awaiting_broadcast'):
            context.user_data['awaiting_broadcast'] = False
            await run_broadcast(update, context, update.message)
        return

    if not is_admin(username, user_id):
        await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
        return

    try:
        import re as _re_num
        import io
        import csv

        country_name = context.user_data['current_country_name']

        # Download file bytes
        document = await update.message.document.get_file()
        file_bytes = await document.download_as_bytearray()

        file_name = (update.message.document.file_name or '').lower()
        mime_type = (update.message.document.mime_type or '').lower()

        numbers_list = []

        # ── Excel (.xlsx) ──────────────────────────────────────────────
        if file_name.endswith('.xlsx') or 'spreadsheetml' in mime_type:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            raw_values = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            raw_values.append(str(cell).strip())
            text_blob = '\n'.join(raw_values)

        # ── Excel legacy (.xls) ────────────────────────────────────────
        elif file_name.endswith('.xls') or 'ms-excel' in mime_type:
            import xlrd
            wb = xlrd.open_workbook(file_contents=bytes(file_bytes))
            raw_values = []
            for sheet in wb.sheets():
                for row_idx in range(sheet.nrows):
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        raw_values.append(str(cell.value).strip())
            text_blob = '\n'.join(raw_values)

        # ── CSV ────────────────────────────────────────────────────────
        elif file_name.endswith('.csv') or 'csv' in mime_type:
            text_blob = file_bytes.decode('utf-8', errors='ignore')
            reader = csv.reader(io.StringIO(text_blob))
            raw_values = [cell.strip() for row in reader for cell in row if cell.strip()]
            text_blob = '\n'.join(raw_values)

        # ── TXT / any other text-based file ───────────────────────────
        else:
            try:
                text_blob = file_bytes.decode('utf-8', errors='ignore')
            except Exception:
                text_blob = file_bytes.decode('latin-1', errors='ignore')

        # Extract numbers: sequences of 5-15 digits (phone numbers)
        numbers_list = _re_num.findall(r'\b\d{5,15}\b', text_blob)
        numbers_list = list(dict.fromkeys(numbers_list))  # deduplicate, preserve order

        if not numbers_list:
            await update.message.reply_text(f"{_sx('❌')} No valid numbers found in the file.", parse_mode='HTML')
            return

        # Add country if not exists
        add_country(country_name)

        # Look up country_id
        countries_list = get_countries()
        country_id = None
        for cid, cname in countries_list:
            if cname == country_name:
                country_id = cid
                break
        if not country_id:
            await update.message.reply_text(f"{_sx('❌')} Error: Country not found after creation.", parse_mode='HTML')
            return

        added_count = add_numbers_to_country(country_id, numbers_list)

        context.user_data.pop('awaiting_numbers_file', None)
        context.user_data.pop('current_country_name', None)

        await update.message.reply_text(
            f"{_sx('✅')} Successfully added {added_count} numbers to {country_name}!\n"
            f"{_sx('📊')} Extracted: {len(numbers_list)} | Added (new): {added_count}",
            parse_mode='HTML'
        )

    except Exception as e:
        logger.error(f"Error processing file: {e}")
        await update.message.reply_text(f"{_sx('❌')} Error processing file: {str(e)}", parse_mode='HTML')

# Handle text messages
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    username = update.effective_user.username
    text = update.message.text.strip()
    user_id = update.effective_user.id

    # ── Message Hub sub-buttons (dynamic text with counts) ─────────────────────
    if (text.startswith("📢 Broadcast All") or text.startswith("📣 All Users")) and is_admin(username, user_id):
        # Clear ALL conflicting awaiting states before setting broadcast
        for _k in ('awaiting_direct_uid', 'awaiting_direct_message', 'direct_msg_uid',
                   'awaiting_new_admin', 'awaiting_ref_check_user', 'awaiting_ref_add_balance',
                   'awaiting_ref_remove_balance', 'awaiting_custom_message'):
            context.user_data.pop(_k, None)
        context.user_data['awaiting_broadcast'] = True
        context.user_data['broadcast_target']   = 'all'
        await update.message.reply_text(
            f"📢 <b>Broadcast All Users</b>\n\n"
            "Message পাঠান (text, photo, video, sticker, voice বা forward করা message):\n"
            "<i>Message পাঠানোর সাথে সাথেই সবার কাছে চলে যাবে।</i>",
            parse_mode='HTML'
        )
        return

    elif text.startswith("👤 One User") and is_admin(username, user_id):
        # Clear ALL conflicting awaiting states before setting direct uid
        for _k in ('awaiting_broadcast', 'broadcast_target', 'broadcast_msg_id',
                   'broadcast_chat_id', 'broadcast_is_forwarded',
                   'awaiting_direct_message', 'direct_msg_uid'):
            context.user_data.pop(_k, None)
        context.user_data['awaiting_direct_uid'] = True
        cancel_kb = ReplyKeyboardMarkup(
            [[KeyboardButton("❌ বাতিল করুন")]],
            resize_keyboard=True, one_time_keyboard=False
        )
        await update.message.reply_text(
            f"👤 <b>Direct Message</b>\n\n"
            "যে user কে message পাঠাতে চান তার <b>Telegram User ID (UID)</b> দিন:",
            parse_mode='HTML',
            reply_markup=cancel_kb
        )
        return

    # ── Direct Message by UID — step 1: receive UID ───────────────────────────
    if context.user_data.get('awaiting_direct_uid') and not context.user_data.get('awaiting_broadcast'):
        if not is_admin(username, user_id):
            return
        try:
            target_uid = int(text.strip())
        except ValueError:
            await update.message.reply_text(
                "❌ Valid UID দিন (শুধু numbers):",
                parse_mode='HTML'
            )
            return
        context.user_data['direct_msg_uid']      = target_uid
        context.user_data['awaiting_direct_uid'] = False
        context.user_data['awaiting_direct_message'] = True
        cancel_kb = ReplyKeyboardMarkup(
            [[KeyboardButton("❌ বাতিল করুন")]],
            resize_keyboard=True, one_time_keyboard=False
        )
        await update.message.reply_text(
            f"✅ UID: <code>{target_uid}</code>\n\n"
            "এখন message পাঠান — text, photo, video, audio, sticker, voice যেকোনো:",
            parse_mode='HTML',
            reply_markup=cancel_kb
        )
        return

    # ── Direct Message by UID — step 2: send text message ─────────────────────
    elif context.user_data.get('awaiting_direct_message'):
        if not is_admin(username, user_id):
            return
        target_uid = context.user_data.pop('direct_msg_uid', None)
        context.user_data['awaiting_direct_message'] = False
        if not target_uid:
            await update.message.reply_text("❌ UID not found. আবার চেষ্টা করুন।", parse_mode='HTML')
            return
        try:
            await context.bot.send_message(chat_id=target_uid, text=text)
            await update.message.reply_text(
                f"✅ <b>Sent!</b> UID <code>{target_uid}</code> কে message পাঠানো হয়েছে।",
                parse_mode='HTML'
            )
        except Exception as e:
            await update.message.reply_text(
                f"❌ <b>Failed:</b> <code>{target_uid}</code>\n<i>{e}</i>",
                parse_mode='HTML'
            )
        return

    # ── SMS Panel setup flow ───────────────────────────────────────────────────
    if context.user_data.get('awaiting_panel_name'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized.", parse_mode='HTML')
            return
        context.user_data['panel_setup_name'] = text
        context.user_data['awaiting_panel_name'] = False
        context.user_data['awaiting_panel_login_url'] = True
        await update.message.reply_text(
            f"{_sx('📡')} Step 2/11 — Login URL দিন:\n"
            f"<i>উদাহরণ: http://smshadi.net/login</i>",
            parse_mode='HTML'
        )
        return

    elif context.user_data.get('awaiting_panel_login_url'):
        if not is_admin(username, user_id):
            return
        context.user_data['panel_setup_login_url'] = text
        context.user_data['awaiting_panel_login_url'] = False
        context.user_data['awaiting_panel_message_url'] = True
        await update.message.reply_text(
            f"{_sx('📡')} Step 3/11 — Message check পেইজের URL দিন:\n"
            f"<i>উদাহরণ: http://smshadi.net/agent/SMSCDRStats</i>",
            parse_mode='HTML'
        )
        return

    elif context.user_data.get('awaiting_panel_message_url'):
        if not is_admin(username, user_id):
            return
        context.user_data['panel_setup_message_url'] = text
        context.user_data['awaiting_panel_message_url'] = False
        context.user_data['awaiting_panel_username'] = True
        await update.message.reply_text(
            f"{_sx('📡')} Step 4/11 — Username দিন:",
            parse_mode='HTML'
        )
        return

    elif context.user_data.get('awaiting_panel_username'):
        if not is_admin(username, user_id):
            return
        context.user_data['panel_setup_username'] = text
        context.user_data['awaiting_panel_username'] = False
        context.user_data['awaiting_panel_password'] = True
        await update.message.reply_text(
            f"{_sx('📡')} Step 5/11 — Password দিন:",
            parse_mode='HTML'
        )
        return

    elif context.user_data.get('awaiting_panel_password'):
        if not is_admin(username, user_id):
            return
        context.user_data['panel_setup_password'] = text
        context.user_data['awaiting_panel_password'] = False
        context.user_data['awaiting_panel_num_col_name'] = True
        await update.message.reply_text(
            f"{_sx('📡')} Step 6/11 — <b>নাম্বার কলামের নাম</b> দিন:\n\n"
            f"পেইজের table-এ যে column-এ ফোন নাম্বার থাকে সেটার header নাম লিখুন।\n"
            f"<i>উদাহরণ: Phone, Number, Mobile, MSISDN</i>",
            parse_mode='HTML'
        )
        return

    elif context.user_data.get('awaiting_panel_num_col_name'):
        if not is_admin(username, user_id):
            return
        context.user_data['panel_setup_num_col_name'] = text.strip()
        context.user_data['awaiting_panel_num_col_name'] = False
        context.user_data['awaiting_panel_num_col_idx'] = True
        await update.message.reply_text(
            f"{_sx('📡')} Step 7/11 — <b>নাম্বার কলামটি কত নম্বরে আছে?</b>\n\n"
            f"বাম দিক থেকে গণনা করুন, ১ থেকে শুরু।\n"
            f"<i>উদাহরণ: ১ম কলাম হলে 1, ৩য় কলাম হলে 3</i>",
            parse_mode='HTML'
        )
        return

    elif context.user_data.get('awaiting_panel_num_col_idx'):
        if not is_admin(username, user_id):
            return
        if not text.strip().isdigit() or int(text.strip()) < 1:
            await update.message.reply_text(
                f"❌ সঠিক সংখ্যা দিন (১ বা তার বেশি)।\n<i>বাতিল করতে /cancel</i>",
                parse_mode='HTML'
            )
            return
        context.user_data['panel_setup_num_col_idx'] = int(text.strip())
        context.user_data['awaiting_panel_num_col_idx'] = False
        context.user_data['awaiting_panel_svc_col_name'] = True
        await update.message.reply_text(
            f"{_sx('📡')} Step 8/11 — <b>Service/Sender কলামের নাম</b> দিন:\n\n"
            f"যে column-এ app বা sender-এর নাম থাকে।\n"
            f"<i>উদাহরণ: Sender, Service, From, App</i>\n\n"
            f"এই কলাম না থাকলে <b>skip</b> লিখুন।",
            parse_mode='HTML'
        )
        return

    elif context.user_data.get('awaiting_panel_svc_col_name'):
        if not is_admin(username, user_id):
            return
        val = text.strip()
        context.user_data['awaiting_panel_svc_col_name'] = False
        if val.lower() == 'skip':
            context.user_data['panel_setup_svc_col_name'] = None
            context.user_data['panel_setup_svc_col_idx'] = None
            context.user_data['awaiting_panel_msg_col_name'] = True
            await update.message.reply_text(
                f"{_sx('📡')} Step 10/11 — <b>Full Message কলামের নাম</b> দিন:\n\n"
                f"যে column-এ পুরো SMS text থাকে।\n"
                f"<i>উদাহরণ: Message, Text, SMS, Body, Content</i>",
                parse_mode='HTML'
            )
        else:
            context.user_data['panel_setup_svc_col_name'] = val
            context.user_data['awaiting_panel_svc_col_idx'] = True
            await update.message.reply_text(
                f"{_sx('📡')} Step 9/11 — <b>Service কলামটি কত নম্বরে আছে?</b>\n\n"
                f"বাম দিক থেকে গণনা করুন, ১ থেকে শুরু।\n"
                f"<i>উদাহরণ: ২য় কলাম হলে 2</i>",
                parse_mode='HTML'
            )
        return

    elif context.user_data.get('awaiting_panel_svc_col_idx'):
        if not is_admin(username, user_id):
            return
        if not text.strip().isdigit() or int(text.strip()) < 1:
            await update.message.reply_text(
                f"❌ সঠিক সংখ্যা দিন (১ বা তার বেশি)।\n<i>বাতিল করতে /cancel</i>",
                parse_mode='HTML'
            )
            return
        context.user_data['panel_setup_svc_col_idx'] = int(text.strip())
        context.user_data['awaiting_panel_svc_col_idx'] = False
        context.user_data['awaiting_panel_msg_col_name'] = True
        await update.message.reply_text(
            f"{_sx('📡')} Step 10/11 — <b>Full Message কলামের নাম</b> দিন:\n\n"
            f"যে column-এ পুরো SMS text থাকে।\n"
            f"<i>উদাহরণ: Message, Text, SMS, Body, Content</i>",
            parse_mode='HTML'
        )
        return

    elif context.user_data.get('awaiting_panel_msg_col_name'):
        if not is_admin(username, user_id):
            return
        context.user_data['panel_setup_msg_col_name'] = text.strip()
        context.user_data['awaiting_panel_msg_col_name'] = False
        context.user_data['awaiting_panel_msg_col_idx'] = True
        await update.message.reply_text(
            f"{_sx('📡')} Step 11/11 — <b>Message কলামটি কত নম্বরে আছে?</b>\n\n"
            f"বাম দিক থেকে গণনা করুন, ১ থেকে শুরু।\n"
            f"<i>উদাহরণ: ৫ম কলাম হলে 5</i>",
            parse_mode='HTML'
        )
        return

    elif context.user_data.get('awaiting_panel_msg_col_idx'):
        if not is_admin(username, user_id):
            return
        if not text.strip().isdigit() or int(text.strip()) < 1:
            await update.message.reply_text(
                f"❌ সঠিক সংখ্যা দিন (১ বা তার বেশি)।\n<i>বাতিল করতে /cancel</i>",
                parse_mode='HTML'
            )
            return
        # Build column_map
        column_map = {
            "number": {
                "name": context.user_data.pop('panel_setup_num_col_name', ''),
                "index": context.user_data.pop('panel_setup_num_col_idx', 1),
            },
            "message": {
                "name": context.user_data.pop('panel_setup_msg_col_name', ''),
                "index": int(text.strip()),
            },
        }
        svc_name = context.user_data.pop('panel_setup_svc_col_name', None)
        svc_idx  = context.user_data.pop('panel_setup_svc_col_idx', None)
        if svc_name and svc_idx:
            column_map["service"] = {"name": svc_name, "index": svc_idx}

        name      = context.user_data.pop('panel_setup_name', '')
        login_url = context.user_data.pop('panel_setup_login_url', '')
        msg_url   = context.user_data.pop('panel_setup_message_url', '')
        uname     = context.user_data.pop('panel_setup_username', '')
        password  = context.user_data.pop('panel_setup_password', '')
        context.user_data['awaiting_panel_msg_col_idx'] = False

        panel_id = add_sms_panel(name, login_url, msg_url, uname, password, column_map)
        panel = get_sms_panel(panel_id)

        # Start monitoring immediately
        try:
            start_panel_monitor(context.bot, panel, notify_chat_id=user_id)
        except Exception as e:
            logger.error(f"Panel monitor start error: {e}")

        svc_line = f"📌 Service কলাম: <b>{svc_name}</b> (কলাম {svc_idx})\n" if svc_name else "📌 Service কলাম: <i>Skip করা হয়েছে</i>\n"
        await update.message.reply_text(
            f"{_sx('✅')} <b>Panel যোগ হয়েছে!</b>\n\n"
            f"📡 Name: <b>{name}</b>\n"
            f"🔗 Login: <code>{login_url}</code>\n"
            f"📄 Page: <code>{msg_url}</code>\n"
            f"👤 User: <code>{uname}</code>\n\n"
            f"<b>Column Config:</b>\n"
            f"📞 Number কলাম: <b>{column_map['number']['name']}</b> (কলাম {column_map['number']['index']})\n"
            f"{svc_line}"
            f"💬 Message কলাম: <b>{column_map['message']['name']}</b> (কলাম {column_map['message']['index']})\n\n"
            f"<i>এখন থেকে প্রতি 3 সেকেন্ডে message চেক হবে।</i>",
            parse_mode='HTML',
            reply_markup=get_sms_panels_keyboard()
        )
        return

    elif context.user_data.get('awaiting_panel_edit_username'):
        if not is_admin(username, user_id):
            return
        panel_id = context.user_data.get('viewing_panel_id')
        if not panel_id:
            await update.message.reply_text(f"{_sx('❌')} কোনো panel নির্বাচন করা হয়নি।", parse_mode='HTML',
                                            reply_markup=get_sms_panels_keyboard())
            context.user_data.pop('awaiting_panel_edit_username', None)
            return
        new_username = text.strip()
        context.user_data.pop('awaiting_panel_edit_username', None)
        success = update_sms_panel_credentials(panel_id, username=new_username)
        panel = get_sms_panel(panel_id)
        if success and panel:
            await update.message.reply_text(
                f"{_sx('✅')} <b>Username আপডেট হয়েছে!</b>\n\n"
                f"Panel: <b>{panel['name']}</b>\n"
                f"👤 নতুন Username: <code>{new_username}</code>",
                parse_mode='HTML',
                reply_markup=get_panel_detail_keyboard(panel)
            )
        else:
            await update.message.reply_text(f"{_sx('❌')} আপডেট ব্যর্থ হয়েছে।", parse_mode='HTML')
        return

    elif context.user_data.get('awaiting_panel_edit_password'):
        if not is_admin(username, user_id):
            return
        panel_id = context.user_data.get('viewing_panel_id')
        if not panel_id:
            await update.message.reply_text(f"{_sx('❌')} কোনো panel নির্বাচন করা হয়নি।", parse_mode='HTML',
                                            reply_markup=get_sms_panels_keyboard())
            context.user_data.pop('awaiting_panel_edit_password', None)
            return
        new_password = text.strip()
        context.user_data.pop('awaiting_panel_edit_password', None)
        success = update_sms_panel_credentials(panel_id, password=new_password)
        panel = get_sms_panel(panel_id)
        if success and panel:
            await update.message.reply_text(
                f"{_sx('✅')} <b>Password আপডেট হয়েছে!</b>\n\n"
                f"Panel: <b>{panel['name']}</b>\n"
                f"🔑 Password পরিবর্তন সম্পন্ন।",
                parse_mode='HTML',
                reply_markup=get_panel_detail_keyboard(panel)
            )
        else:
            await update.message.reply_text(f"{_sx('❌')} আপডেট ব্যর্থ হয়েছে।", parse_mode='HTML')
        return

    # ── SMS Panel Check Interval flow ─────────────────────────────────────────
    elif context.user_data.get('awaiting_panel_interval'):
        if not is_admin(username, user_id):
            return
        panel_id = context.user_data.get('viewing_panel_id')
        context.user_data.pop('awaiting_panel_interval', None)
        if not panel_id:
            await update.message.reply_text(f"{_sx('❌')} কোনো panel নির্বাচন করা হয়নি।", parse_mode='HTML',
                                            reply_markup=get_sms_panels_keyboard())
            return
        panel = get_sms_panel(panel_id)
        if not panel:
            await update.message.reply_text(f"{_sx('❌')} Panel পাওয়া যায়নি।", parse_mode='HTML')
            return
        if not text.strip().isdigit():
            await update.message.reply_text(
                f"❌ শুধু সংখ্যা দিন। উদাহরণ: <code>30</code>",
                parse_mode='HTML',
                reply_markup=get_panel_detail_keyboard(panel)
            )
            return
        seconds = int(text.strip())
        if seconds < 3:
            seconds = 3
        elif seconds > 3600:
            seconds = 3600
        update_sms_panel_interval(panel_id, seconds)
        panel = get_sms_panel(panel_id)
        await update.message.reply_text(
            f"✅ <b>Interval আপডেট হয়েছে!</b>\n\n"
            f"📡 Panel: <b>{panel['name']}</b>\n"
            f"⏱ এখন থেকে প্রতি <b>{seconds} সেকেন্ড</b> পর পর চেক হবে।",
            parse_mode='HTML',
            reply_markup=get_panel_detail_keyboard(panel)
        )
        return

    # Dynamic "⏱ Interval (Xs)" button — per-panel SMS check interval
    elif is_admin(username, user_id) and text.startswith("⏱ Interval ("):
        panel_id = context.user_data.get('viewing_panel_id')
        if not panel_id:
            await update.message.reply_text(f"{_sx('❌')} কোনো panel নির্বাচন করা হয়নি।", parse_mode='HTML',
                                            reply_markup=get_sms_panels_keyboard())
            return
        panel = get_sms_panel(panel_id)
        if not panel:
            await update.message.reply_text(f"{_sx('❌')} Panel পাওয়া যায়নি।", parse_mode='HTML')
            return
        current = int(panel.get("poll_interval", 3))
        context.user_data['awaiting_panel_interval'] = True
        await update.message.reply_text(
            f"⏱ <b>SMS Check Interval</b>\n\n"
            f"Panel: <b>{panel['name']}</b>\n"
            f"বর্তমান: প্রতি <b>{current} সেকেন্ড</b> পর পর চেক হয়।\n\n"
            f"নতুন interval সেকেন্ডে লিখুন (সর্বনিম্ন ৩, সর্বোচ্চ ৩৬০০):\n"
            f"<i>উদাহরণ: 10 → প্রতি ১০ সেকেন্ডে, 60 → প্রতি ১ মিনিটে</i>\n\n"
            f"<i>বাতিল করতে /cancel লিখুন।</i>",
            parse_mode='HTML',
            reply_markup=ReplyKeyboardRemove()
        )
        return

    # Dynamic SMS panel buttons (e.g. "📡 ✅ Number Panel" / "📡 ❌ Number Panel")
    elif is_admin(username, user_id) and (text.startswith("📡 ✅ ") or text.startswith("📡 ❌ ")):
        panel_name = text[4:].strip()
        panels = get_sms_panels()
        matched = next((p for p in panels if p['name'] == panel_name), None)
        if not matched:
            await update.message.reply_text(f"{_sx('❌')} Panel পাওয়া যায়নি।", parse_mode='HTML')
            return
        context.user_data['viewing_panel_id'] = matched['id']
        icon = "🟢 চালু" if matched.get("enabled", True) else "🔴 বন্ধ"
        await update.message.reply_text(
            f"{_sx('📡')} <b>{matched['name']}</b>\n\n"
            f"স্ট্যাটাস: {icon}\n"
            f"🔗 Login URL: <code>{matched['login_url']}</code>\n"
            f"📄 Message URL: <code>{matched['message_url']}</code>\n"
            f"👤 Username: <code>{matched['username']}</code>",
            parse_mode='HTML',
            reply_markup=get_panel_detail_keyboard(matched)
        )
        return

    # Admin functionality
    if context.user_data.get('awaiting_new_country'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        if text.lower().startswith("delete "):
            country_name = _clean_name(text[7:].strip())
            # Find country by name
            countries = get_countries()
            matched = next(((cid, cname) for cid, cname in countries if _clean_name(cname).lower() == country_name.lower()), None)
            if matched:
                numbers_deleted, country_deleted = delete_country(matched[0])
                if country_deleted:
                    await update.message.reply_text(
                        f"{_sx('✅')} Country <b>{matched[1]}</b> and {numbers_deleted} number(s) deleted successfully!",
                        parse_mode='HTML'
                    )
                else:
                    await update.message.reply_text(f"{_sx('❌')} Failed to delete country.", parse_mode='HTML')
            else:
                await update.message.reply_text(f"{_sx('❌')} Country <b>{country_name}</b> not found!", parse_mode='HTML')
        else:
            plain = _clean_name(text)
            animated = get_animated_flag_html(plain)
            unicode_flag = get_unicode_flag(plain)
            if add_country(plain):
                await update.message.reply_text(
                    f'{_sx("✅")} Country {animated} <b>{plain}</b> saved successfully!',
                    parse_mode='HTML'
                )
            else:
                await update.message.reply_text(
                    f'{animated} <b>{plain}</b> already exists!',
                    parse_mode='HTML'
                )
        context.user_data['awaiting_new_country'] = False

    elif context.user_data.get('awaiting_add_numbers_country'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        countries = get_countries()
        plain_input = _clean_name(text).lower()
        matched = next(((cid, cname) for cid, cname in countries if _clean_name(cname).lower() == plain_input), None)
        if not matched:
            country_list = "\n".join([f"{get_animated_flag_html(_clean_name(name))} <code>{_clean_name(name)}</code>" for _, name in countries])
            await update.message.reply_text(
                f"{_sx('❌')} Country <code>{text}</code> not found!\n\n📋 <b>Saved Countries:</b>\n\n{country_list}\n\nType the exact country name from the list:",
                parse_mode='HTML'
            )
            return
        context.user_data['current_country_name'] = matched[1]
        context.user_data['awaiting_add_numbers_country'] = False
        context.user_data['awaiting_numbers_file'] = True
        await update.message.reply_text(
            f"{_sx('✅')} Country: <b>{matched[1]}</b>\n\nNow send the file with numbers (TXT, CSV, Excel — any format):",
            parse_mode='HTML'
        )
    
    elif context.user_data.get('awaiting_new_service'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        if text.lower().startswith("remove "):
            svc_name = text[7:].strip()
            if delete_service(svc_name):
                await update.message.reply_text(
                    f"{_sx('✅')} <code>{svc_name}</code> removed successfully!",
                    parse_mode='HTML'
                )
            else:
                await update.message.reply_text(f"{_sx('❌')} Service <code>{svc_name}</code> not found!", parse_mode='HTML')
            context.user_data['awaiting_new_service'] = False
        else:
            # Retrieve sticker emoji data captured earlier (custom_emoji_id only; no plain char prepended)
            context.user_data.pop('pending_service_emoji', None)
            pending_ceid  = context.user_data.pop('pending_service_custom_emoji_id', '')
            # Service name = exactly what was typed; animated emoji stored separately
            service_name = text
            if add_service(service_name, custom_emoji_id=pending_ceid):
                anim = _svc_animated_tag(service_name, pending_ceid)
                await update.message.reply_text(
                    f"{_sx('✅')} Service {anim} <b>{service_name}</b> added successfully!",
                    parse_mode='HTML'
                )
            else:
                await update.message.reply_text(
                    f"{_sx('❌')} <code>{service_name}</code> already exists!",
                    parse_mode='HTML'
                )
            context.user_data['awaiting_new_service'] = False

    elif context.user_data.get('awaiting_service_map'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        is_unmap = text.lower().startswith("unmap ")
        raw = text[6:].strip() if is_unmap else text.strip()
        parts = raw.split(" ", 1)
        if len(parts) < 2:
            await update.message.reply_text(
                f"{_sx('❌')} Format:\nLink: <code>ServiceName CountryName</code>\nUnmap: <code>unmap ServiceName CountryName</code>",
                parse_mode='HTML'
            )
            return
        svc_name, cnt_name = parts[0].strip(), parts[1].strip()
        services = get_services()
        svc = next(((sid, sname) for sid, sname in services if sname.lower() == svc_name.lower()), None)
        if not svc:
            await update.message.reply_text(f"{_sx('❌')} Service <b>{svc_name}</b> not found!", parse_mode='HTML')
            return
        countries = get_countries()
        cnt = next(((cid, cname) for cid, cname in countries if cname.lower() == cnt_name.lower()), None)
        if not cnt:
            await update.message.reply_text(f"{_sx('❌')} Country <b>{cnt_name}</b> not found!", parse_mode='HTML')
            return
        if is_unmap:
            if unlink_country_from_service(svc[0], cnt[0]):
                await update.message.reply_text(f"{_sx('✅')} <b>{cnt[1]}</b> removed from <b>{svc[1]}</b> successfully!", parse_mode='HTML')
            else:
                await update.message.reply_text(f"{_sx('❌')} <b>{cnt[1]}</b> is not linked to <b>{svc[1]}</b>!", parse_mode='HTML')
        else:
            if add_country_to_service(svc[0], cnt[0]):
                await update.message.reply_text(f"{_sx('✅')} <b>{cnt[1]}</b> added to <b>{svc[1]}</b> successfully!", parse_mode='HTML')
            else:
                await update.message.reply_text(f"{_sx('❌')} <b>{cnt[1]}</b> is already in <b>{svc[1]}</b>!", parse_mode='HTML')

    elif context.user_data.get('awaiting_reset_country'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        countries = get_countries()
        matched = next(((cid, cname) for cid, cname in countries if cname.lower() == text.lower()), None)
        if matched:
            deleted = delete_all_numbers_from_country(matched[0])
            await update.message.reply_text(
                f"{_sx('✅')} <b>{matched[1]}</b> — {deleted} number(s) removed successfully!",
                parse_mode='HTML'
            )
        else:
            await update.message.reply_text(f"{_sx('❌')} Country <b>{text}</b> not found!", parse_mode='HTML')
        context.user_data['awaiting_reset_country'] = False

    elif context.user_data.get('awaiting_country_otp_reward'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        lines = [l.strip() for l in text.strip().splitlines() if l.strip()]
        success_lines = []
        error_lines = []
        for line in lines:
            parts = line.rsplit(None, 1)
            if len(parts) != 2:
                error_lines.append(f"❌ <code>{line}</code> — ফরম্যাট ঠিক নেই")
                continue
            cname_raw, amt_raw = parts
            try:
                amt = float(amt_raw)
            except ValueError:
                error_lines.append(f"❌ <code>{line}</code> — amount সংখ্যা হতে হবে")
                continue
            set_country_otp_reward(cname_raw.strip(), amt)
            success_lines.append(f"✅ <b>{cname_raw.strip()}</b>: <code>{amt}৳</code>")
        msg_parts = []
        if success_lines:
            msg_parts.append("💰 <b>OTP Reward সেট হয়েছে:</b>\n" + "\n".join(success_lines))
        if error_lines:
            msg_parts.append("\n".join(error_lines))
        await update.message.reply_text("\n\n".join(msg_parts) if msg_parts else "কোনো পরিবর্তন হয়নি।", parse_mode='HTML')
        context.user_data['awaiting_country_otp_reward'] = False

    elif context.user_data.get('awaiting_number_delete'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        
        if delete_number(text):
            await update.message.reply_text(f"{_sx('✅')} Number deleted successfully!", parse_mode='HTML')
        else:
            await update.message.reply_text(f"{_sx('❌')} Number not found!", parse_mode='HTML')
        
        context.user_data['awaiting_number_delete'] = False
    
    elif context.user_data.get('awaiting_new_admin'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        
        if not text.isdigit():
            await update.message.reply_text(f"{_sx('❌')} Please send a valid Telegram UID (numbers only). e.g: 123456789", parse_mode='HTML')
            return
        
        new_uid = int(text)
        if add_admin(new_uid):
            await update.message.reply_text(f"{_sx('✅')} Admin <code>{new_uid}</code> has been added successfully!", parse_mode='HTML')
        else:
            await update.message.reply_text(f"{_sx('❌')} UID <code>{new_uid}</code> is already an admin!", parse_mode='HTML')
        
        context.user_data['awaiting_new_admin'] = False
    
    elif context.user_data.get('awaiting_broadcast'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        context.user_data['awaiting_broadcast'] = False
        await run_broadcast(update, context, update.message)
    
    elif context.user_data.get('awaiting_emoji_add'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        parts = text.strip().split()
        if len(parts) < 2:
            await update.message.reply_text(
                f"{_sx('❌')} সঠিক ফরম্যাটে দিন:\n<code>ServiceName EmojiID FallbackUnicode</code>\n\nউদাহরণ: <code>WhatsApp 6300761828330840482 💬</code>",
                parse_mode='HTML'
            )
            return
        service_name = parts[0]
        emoji_id = parts[1]
        if not emoji_id.isdigit():
            await update.message.reply_text(f"{_sx('❌')} Emoji ID শুধু সংখ্যা হতে হবে!", parse_mode='HTML')
            return
        fallback = parts[2] if len(parts) >= 3 else '📱'
        save_custom_emoji(service_name, emoji_id, fallback)
        context.user_data['awaiting_emoji_add'] = False
        await update.message.reply_text(
            f"{_sx('✅')} <b>Emoji সংরক্ষিত!</b>\n\n"
            f"🏷️ Service: <b>{service_name}</b>\n"
            f'🎭 Preview: <tg-emoji emoji-id="{emoji_id}">{fallback}</tg-emoji>\n'
            f"🔢 ID: <code>{emoji_id}</code>\n\n"
            f"এখন থেকে <b>{service_name}</b> এর পাশে এই animated emoji দেখাবে।",
            parse_mode='HTML',
            reply_markup=get_emoji_manager_keyboard()
        )

    elif context.user_data.get('awaiting_emoji_remove'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        service_name = text.strip()
        deleted = delete_custom_emoji(service_name)
        context.user_data['awaiting_emoji_remove'] = False
        if deleted:
            await update.message.reply_text(
                f"{_sx('✅')} <b>{service_name}</b> এর custom emoji মুছে ফেলা হয়েছে।",
                parse_mode='HTML',
                reply_markup=get_emoji_manager_keyboard()
            )
        else:
            await update.message.reply_text(
                f"{_sx('❌')} <b>{service_name}</b> নামে কোনো custom emoji পাওয়া যায়নি।\n"
                f"📋 Emoji List দেখে সঠিক নাম দিন।",
                parse_mode='HTML',
                reply_markup=get_emoji_manager_keyboard()
            )

    elif context.user_data.get('awaiting_withdraw_group_id'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        chat_id_input = text.strip()

        # remove [link] — remove a join channel by its link
        if chat_id_input.lower().startswith("remove "):
            target_link = chat_id_input[7:].strip()
            channels = get_join_channels()
            matched_channel = next(
                ((cid, t, l) for cid, t, l in channels if l.strip() == target_link),
                None
            )
            if matched_channel:
                remove_join_channel(matched_channel[0])
                context.user_data['awaiting_withdraw_group_id'] = False
                await update.message.reply_text(
                    f"{_sx('✅')} Channel <b>{matched_channel[1]}</b> removed successfully!\n\n"
                    f"{_sx('⚠️')} All users have been <b>unverified</b> — they will need to re-verify.",
                    parse_mode='HTML',
                    reply_markup=get_withdraw_settings_keyboard()
                )
            else:
                channels = get_join_channels()
                ch_lines = "\n".join([f"  • {t} — <code>{l}</code>" for _, t, l in channels]) or "  (none)"
                await update.message.reply_text(
                    f"{_sx('❌')} No channel found with that link.\n\n"
                    f"📋 <b>Current join channels:</b>\n{ch_lines}\n\n"
                    f"Copy the exact link from above and try again.",
                    parse_mode='HTML'
                )
            return

        if not chat_id_input.lstrip('-').isdigit():
            await update.message.reply_text(
                f"{_sx('❌')} Please send a valid Chat ID (e.g. <code>-1001234567890</code>)\n"
                "Or type <code>remove [link]</code> to remove a join channel.",
                parse_mode='HTML'
            )
            return
        set_withdraw_group_chat_id(chat_id_input)
        context.user_data['awaiting_withdraw_group_id'] = False
        await update.message.reply_text(
            f"{_sx('✅')} Withdraw Group ID set: <code>{chat_id_input}</code>\n\n"
            f"All withdraw requests will now be sent to this group.",
            parse_mode='HTML',
            reply_markup=get_withdraw_settings_keyboard()
        )

    elif context.user_data.get('awaiting_withdraw_min_amount'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        if not text.isdigit() or int(text) < 0:
            await update.message.reply_text(f"{_sx('❌')} সঠিক সংখ্যা লিখুন (যেমন: <code>100</code>)!", parse_mode='HTML')
            return
        amount = int(text)
        set_withdraw_min_amount(amount)
        context.user_data['awaiting_withdraw_min_amount'] = False
        label = get_referral_settings()['label']
        await update.message.reply_text(
            f"{_sx('✅')} Minimum Withdraw Amount সেট হয়েছে: <b>{amount} {label}</b>",
            parse_mode='HTML',
            reply_markup=get_withdraw_settings_keyboard()
        )

    elif context.user_data.get('awaiting_otp_link'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        link = text.strip()
        set_otp_link(link)
        context.user_data['awaiting_otp_link'] = False
        await update.message.reply_text(
            f"{_sx('✅')} OTP link updated to:\n<code>{link}</code>",
            parse_mode='HTML',
            reply_markup=get_link_settings_keyboard()
        )

    elif context.user_data.get('awaiting_grp_number_link'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        link = text.strip()
        set_group_number_btn_link(link)
        context.user_data['awaiting_grp_number_link'] = False
        await update.message.reply_text(
            f"✅ <b>📱 Number Button Link</b> আপডেট হয়েছে!\n\n<code>{link}</code>",
            parse_mode='HTML',
            reply_markup=get_link_settings_keyboard()
        )

    elif context.user_data.get('awaiting_grp_channel_link'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        link = text.strip()
        set_group_channel_btn_link(link)
        context.user_data['awaiting_grp_channel_link'] = False
        await update.message.reply_text(
            f"✅ <b>📢 Channel Button Link</b> আপডেট হয়েছে!\n\n<code>{link}</code>",
            parse_mode='HTML',
            reply_markup=get_link_settings_keyboard()
        )

    elif context.user_data.get('awaiting_channel_title'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        context.user_data['pending_channel_title'] = text.strip()
        context.user_data['awaiting_channel_title'] = False
        context.user_data['awaiting_channel_link'] = True
        await update.message.reply_text(
            f"{_sx('➕')} <b>Add Channel</b>\n\nStep 2/2 — Now send the channel <b>invite link</b> (e.g. <code>https://t.me/+xxxxx</code>):",
            parse_mode='HTML'
        )

    elif context.user_data.get('awaiting_channel_link'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        link = text.strip()
        title = context.user_data.pop('pending_channel_title', 'Channel')
        context.user_data['awaiting_channel_link'] = False
        success = add_join_channel(title, link)
        if success:
            await update.message.reply_text(
                f"{_sx('✅')} Channel <b>{title}</b> added successfully!\n\n"
                f"{_sx('⚠️')} All users have been <b>unverified</b> — they must join the new channel and click ✅ Verify again.",
                parse_mode='HTML',
                reply_markup=get_link_settings_keyboard()
            )
        else:
            await update.message.reply_text(f"{_sx('❌')} This link already exists!", parse_mode='HTML', reply_markup=get_link_settings_keyboard())

    elif context.user_data.get('awaiting_remove_channel'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        channels = context.user_data.get('channel_list', [])
        if not text.isdigit() or int(text) < 1 or int(text) > len(channels):
            await update.message.reply_text(f"❌ Please send a valid number between 1 and {len(channels)}!")
            return
        cid, title, _ = channels[int(text) - 1]
        remove_join_channel(cid)
        context.user_data['awaiting_remove_channel'] = False
        context.user_data.pop('channel_list', None)
        await update.message.reply_text(
            f"✅ Channel <b>{title}</b> removed successfully!",
            parse_mode='HTML',
            reply_markup=get_link_settings_keyboard()
        )

    elif context.user_data.get('awaiting_check_interval'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        if not text.isdigit() or int(text) < 1 or int(text) > 1440:
            await update.message.reply_text("❌ ১ থেকে ১৪৪০ এর মধ্যে সংখ্যা পাঠান!")
            return
        set_check_interval(int(text))
        context.user_data['awaiting_check_interval'] = False
        # reset the last-check timer so new interval takes effect immediately
        context.bot_data['last_member_check'] = 0
        await update.message.reply_text(
            f"✅ এখন থেকে প্রতি <b>{text} মিনিট</b> পর পর ইউজার চেক হবে!",
            parse_mode='HTML',
            reply_markup=get_link_settings_keyboard()
        )

    elif context.user_data.get('awaiting_numbers_per_request'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        if not text.isdigit() or int(text) < 1 or int(text) > 20:
            await update.message.reply_text("❌ Please send a valid number between 1 and 20!")
            return
        set_numbers_per_request(int(text))
        context.user_data['awaiting_numbers_per_request'] = False
        await update.message.reply_text(
            f"✅ Numbers per request set to <b>{text}</b>!",
            parse_mode='HTML',
            reply_markup=get_settings_keyboard()
        )

    elif context.user_data.get('awaiting_global_forward_chat_id'):
        if not is_admin(username, user_id):
            return
        context.user_data.pop('awaiting_global_forward_chat_id')
        chat_id_input = text.strip()
        if not re.match(r'^-?\d+$', chat_id_input):
            await update.message.reply_text(
                f"❌ সঠিক Chat ID দিন।\n"
                f"<i>উদাহরণ: -1001234567890 (group/channel)</i>\n\n"
                f"<i>বাতিল করতে /cancel লিখুন।</i>",
                parse_mode='HTML'
            )
            context.user_data['awaiting_global_forward_chat_id'] = True
            return
        # Fetch group/channel name to show in notification
        group_name = chat_id_input
        try:
            chat_info = await context.bot.get_chat(int(chat_id_input))
            group_name = chat_info.title or chat_info.username or chat_id_input
        except Exception:
            pass
        set_global_forward_chat(chat_id_input)
        gfwd_kb = ReplyKeyboardMarkup(
            [
                [KeyboardButton("➕ Group Add"), KeyboardButton("🗑 Group Remove")],
                [KeyboardButton("🔙 Back to Panels")],
            ],
            resize_keyboard=True
        )
        await update.message.reply_text(
            f"✅ <b>Group OTP Forward সেট হয়েছে!</b>\n\n"
            f"📌 Group/Channel: <b>{group_name}</b>\n"
            f"📬 Chat ID: <code>{chat_id_input}</code>\n\n"
            f"এখন থেকে সকল panel এর নতুন OTP ঐ group/channel-এ পাঠানো হবে।",
            parse_mode='HTML',
            reply_markup=gfwd_kb
        )

    elif context.user_data.get('awaiting_forward_chat_id'):
        if not is_admin(username, user_id):
            return
        panel_id = context.user_data.pop('awaiting_forward_chat_id')
        chat_id_input = text.strip()
        # Validate: must be a numeric chat_id (optionally with leading -)
        if not re.match(r'^-?\d+$', chat_id_input):
            await update.message.reply_text(
                f"❌ সঠিক Chat ID দিন।\n"
                f"<i>উদাহরণ: -1001234567890 (group/channel) অথবা 123456789 (user)</i>\n\n"
                f"<i>বাতিল করতে /cancel লিখুন।</i>",
                parse_mode='HTML'
            )
            context.user_data['awaiting_forward_chat_id'] = panel_id
            return
        panel = get_sms_panel(panel_id)
        if not panel:
            await update.message.reply_text(f"{_sx('❌')} Panel পাওয়া যায়নি।", parse_mode='HTML')
            return
        # Test sending a message to verify access
        try:
            await context.bot.send_message(
                chat_id=chat_id_input,
                text=f"✅ <b>Auto Forward সেট হয়েছে!</b>\n\n📡 Panel: <b>{panel['name']}</b>\n\nনতুন SMS এখানে আসবে।",
                parse_mode='HTML'
            )
        except Exception as e:
            await update.message.reply_text(
                f"❌ <b>Test message পাঠানো যায়নি!</b>\n\n"
                f"Error: <code>{e}</code>\n\n"
                f"নিশ্চিত করুন:\n"
                f"• Bot টি ওই channel/group-এ admin হিসেবে আছে\n"
                f"• Chat ID সঠিক আছে\n\n"
                f"<i>আবার চেষ্টা করতে পুনরায় Chat ID দিন, বাতিল করতে /cancel।</i>",
                parse_mode='HTML',
                reply_markup=get_panel_detail_keyboard(panel)
            )
            context.user_data['awaiting_forward_chat_id'] = panel_id
            return
        set_panel_forward_chat(panel_id, chat_id_input)
        updated_panel = get_sms_panel(panel_id)
        await update.message.reply_text(
            f"✅ <b>Auto Forward সেট হয়েছে!</b>\n\n"
            f"📡 Panel: <b>{panel['name']}</b>\n"
            f"📬 Chat ID: <code>{chat_id_input}</code>\n\n"
            f"এখন থেকে নতুন SMS ওই channel/group-এ পাঠানো হবে।",
            parse_mode='HTML',
            reply_markup=get_panel_detail_keyboard(updated_panel)
        )

    elif context.user_data.get('awaiting_custom_message'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        await asyncio.to_thread(add_custom_message, text.strip())
        context.user_data['awaiting_custom_message'] = False
        msgs = await asyncio.to_thread(get_custom_messages)
        await update.message.reply_text(
            f"{_sx('✅')} <b>Message যোগ হয়েছে!</b>\n\n"
            f"📋 <b>Message:</b>\n{text.strip()}\n\n"
            f"মোট {len(msgs)}টি message আছে। User Panel-এর 📋 Notice বাটনে দেখাবে।",
            parse_mode='HTML',
            reply_markup=get_custom_message_keyboard()
        )

    elif context.user_data.get('awaiting_ref_reward'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        if not text.isdigit() or int(text) < 0:
            await update.message.reply_text("❌ Please send a valid positive number!")
            return
        set_referral_setting('reward_per_referral', int(text))
        context.user_data['awaiting_ref_reward'] = False
        await update.message.reply_text(f"✅ Reward set to <b>{text} {get_referral_settings()['label']}</b> per referral!", parse_mode='HTML')

    elif context.user_data.get('awaiting_ref_check_user'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        if not text.isdigit():
            await update.message.reply_text("❌ Please send a valid UID (numbers only)!")
            return
        uid = int(text)
        bal = get_user_balance_data(uid)
        count = get_user_referral_count(uid)
        settings = get_referral_settings()
        await update.message.reply_text(
            f"👤 <b>User {uid}</b>\n\n"
            f"├ 👥 Referrals: <b>{count}</b>\n"
            f"├ 💰 Balance: <b>{bal['balance']} {settings['label']}</b>\n"
            f"└ 🏆 Total Earned: <b>{bal['total_earned']} {settings['label']}</b>",
            parse_mode='HTML'
        )
        context.user_data['awaiting_ref_check_user'] = False

    elif context.user_data.get('awaiting_ref_add_balance'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        parts = text.strip().split()
        if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit() or int(parts[1]) <= 0:
            await update.message.reply_text(
                "❌ সঠিক ফরম্যাটে দিন:\n<code>USER_ID AMOUNT</code>\nউদাহরণ: <code>123456789 50</code>",
                parse_mode='HTML'
            )
            return
        uid, amount = int(parts[0]), int(parts[1])
        new_bal = admin_add_balance(uid, amount)
        settings = get_referral_settings()
        context.user_data['awaiting_ref_add_balance'] = False
        await update.message.reply_text(
            f"✅ <b>Balance Added!</b>\n\n"
            f"👤 User ID: <code>{uid}</code>\n"
            f"➕ Added: <b>{amount} {settings['label']}</b>\n"
            f"💰 New Balance: <b>{new_bal} {settings['label']}</b>",
            parse_mode='HTML',
            reply_markup=get_referral_manager_keyboard()
        )

    elif context.user_data.get('awaiting_ref_remove_balance'):
        if not is_admin(username, user_id):
            await update.message.reply_text(f"{_sx('❌')} Unauthorized access.", parse_mode='HTML')
            return
        parts = text.strip().split()
        if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit() or int(parts[1]) <= 0:
            await update.message.reply_text(
                "❌ সঠিক ফরম্যাটে দিন:\n<code>USER_ID AMOUNT</code>\nউদাহরণ: <code>123456789 20</code>",
                parse_mode='HTML'
            )
            return
        uid, amount = int(parts[0]), int(parts[1])
        new_bal, err = admin_remove_balance(uid, amount)
        settings = get_referral_settings()
        context.user_data['awaiting_ref_remove_balance'] = False
        if err == 'not_found':
            await update.message.reply_text(
                f"❌ User <code>{uid}</code> এর কোনো balance record পাওয়া যায়নি।",
                parse_mode='HTML',
                reply_markup=get_referral_manager_keyboard()
            )
        elif err == 'insufficient':
            await update.message.reply_text(
                f"❌ <b>Insufficient balance!</b>\n\n"
                f"👤 User ID: <code>{uid}</code>\n"
                f"💰 Current Balance: <b>{new_bal} {settings['label']}</b>\n"
                f"আপনি চাইছেন: <b>{amount} {settings['label']}</b> remove করতে।",
                parse_mode='HTML',
                reply_markup=get_referral_manager_keyboard()
            )
        else:
            await update.message.reply_text(
                f"✅ <b>Balance Removed!</b>\n\n"
                f"👤 User ID: <code>{uid}</code>\n"
                f"➖ Removed: <b>{amount} {settings['label']}</b>\n"
                f"💰 New Balance: <b>{new_bal} {settings['label']}</b>",
                parse_mode='HTML',
                reply_markup=get_referral_manager_keyboard()
            )

    elif context.user_data.get('awaiting_withdraw_details'):
        wallet = context.user_data.get('withdraw_wallet', 'Unknown')
        parts = text.strip().split()
        if len(parts) < 2:
            await update.message.reply_text(
                f"❌ Format: <code>NUMBER AMOUNT</code>\nExample: <code>01712345678 100</code>",
                parse_mode='HTML'
            )
            return
        wallet_address = parts[0]
        if not parts[1].isdigit() or int(parts[1]) <= 0:
            await update.message.reply_text("❌ Please send a valid amount (e.g. <code>100</code>)!", parse_mode='HTML')
            return
        amount = int(parts[1])
        settings = get_referral_settings()
        req_id, result = create_withdraw_request(
            user_id, update.effective_user.username or '', wallet, wallet_address, amount, settings['label']
        )
        if result == 'insufficient':
            bal = get_user_balance_data(user_id)
            await update.message.reply_text(
                f"❌ <b>Insufficient balance!</b>\n\nYour balance: <b>{bal['balance']} {settings['label']}</b>\nYou requested: <b>{amount} {settings['label']}</b>",
                parse_mode='HTML'
            )
        else:
            await update.message.reply_text(
                f"✅ <b>Withdraw Request Submitted!</b>\n\n"
                f"💳 Wallet: <b>{wallet}</b>\n"
                f"📱 Address: <code>{wallet_address}</code>\n"
                f"💰 Amount: <b>{amount} {settings['label']}</b>\n\n"
                f"⏳ Your request is pending admin approval.",
                parse_mode='HTML'
            )
            u = update.effective_user
            display = f"@{u.username}" if u.username else (u.first_name or str(user_id))
            kb = [
                [InlineKeyboardButton(f"✅ Approve #{req_id}", callback_data=f"wd_approve_{req_id}"),
                 InlineKeyboardButton(f"❌ Reject #{req_id}", callback_data=f"wd_reject_{req_id}")]
            ]
            req_msg = (
                f"📋 <b>New Withdraw Request #{req_id}</b>\n\n"
                f"👤 User: <b>{display}</b> (<code>{user_id}</code>)\n"
                f"💳 Wallet: <b>{wallet}</b>\n"
                f"📱 Address: <code>{wallet_address}</code>\n"
                f"💰 Amount: <b>{amount} {settings['label']}</b>"
            )
            wcfg = get_withdraw_config()
            if wcfg['group_chat_id']:
                try:
                    await context.bot.send_message(
                        chat_id=int(wcfg['group_chat_id']),
                        text=req_msg,
                        parse_mode='HTML',
                        reply_markup=InlineKeyboardMarkup(kb)
                    )
                except Exception:
                    admin_ids = get_all_admin_ids()
                    for admin_id in admin_ids:
                        try:
                            await context.bot.send_message(
                                chat_id=admin_id,
                                text=req_msg,
                                parse_mode='HTML',
                                reply_markup=InlineKeyboardMarkup(kb)
                            )
                        except Exception:
                            pass
            else:
                admin_ids = get_all_admin_ids()
                for admin_id in admin_ids:
                    try:
                        await context.bot.send_message(
                            chat_id=admin_id,
                            text=req_msg,
                            parse_mode='HTML',
                            reply_markup=InlineKeyboardMarkup(kb)
                        )
                    except Exception:
                        pass
        context.user_data['awaiting_withdraw_details'] = False
        context.user_data.pop('withdraw_wallet', None)

    elif context.user_data.get('awaiting_direct_uid'):
        # Fallback: awaiting_direct_uid was set but not caught by the early chain
        if not is_admin(username, user_id):
            return
        try:
            target_uid = int(text.strip())
        except ValueError:
            await update.message.reply_text(
                "❌ Valid UID দিন (শুধু numbers):",
                parse_mode='HTML'
            )
            return
        context.user_data['direct_msg_uid'] = target_uid
        context.user_data['awaiting_direct_uid'] = False
        context.user_data['awaiting_direct_message'] = True
        await update.message.reply_text(
            f"✅ UID: <code>{target_uid}</code>\n\n"
            "এখন message পাঠান — text, photo, video, audio, sticker, voice যেকোনো:",
            parse_mode='HTML'
        )

    elif context.user_data.get('awaiting_direct_message'):
        # Fallback: awaiting_direct_message was set but not caught by the early chain
        if not is_admin(username, user_id):
            return
        target_uid = context.user_data.pop('direct_msg_uid', None)
        context.user_data['awaiting_direct_message'] = False
        if not target_uid:
            await update.message.reply_text("❌ UID not found. আবার চেষ্টা করুন।", parse_mode='HTML')
            return
        try:
            await context.bot.send_message(chat_id=target_uid, text=text)
            await update.message.reply_text(
                f"✅ <b>Sent!</b> UID <code>{target_uid}</code> কে message পাঠানো হয়েছে।",
                parse_mode='HTML'
            )
        except Exception as e:
            await update.message.reply_text(
                f"❌ <b>Failed:</b> <code>{target_uid}</code>\n<i>{e}</i>",
                parse_mode='HTML'
            )

    else:
        # Check if user is admin and show appropriate interface
        if is_admin(username, user_id) and not context.user_data.get('user_panel_mode'):
            await admin_start(update, context)
        else:
            # For regular users (or admin in user panel mode), show welcome message if not verified
            if await enforce_join(update, context):
                await show_main_menu(update, context)

async def userpanel_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /userpanel command — switch admin to user panel view"""
    username = update.effective_user.username
    user_id = update.effective_user.id

    if not is_admin(username, user_id):
        await update.message.reply_text("❌ You are not authorized to use this command.")
        return

    context.user_data['user_panel_mode'] = True

    await update.message.reply_text(
        "*User Panel*\n\nYou are now viewing the user panel.\nSend /start to return to the Admin Panel.",
        parse_mode='Markdown',
        reply_markup=get_user_keyboard()
    )

async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /start command"""
    username = update.effective_user.username
    user_id = update.effective_user.id
    
    # Store user data for broadcast (non-blocking — fire and forget)
    user = update.effective_user
    asyncio.create_task(asyncio.to_thread(
        add_user, user.id, user.username, user.first_name, user.last_name
    ))

    # Handle referral deep link: /start ref_USERID
    if context.args and context.args[0].startswith('ref_'):
        try:
            referrer_id = int(context.args[0][4:])
            if record_referral(referrer_id, user_id):
                settings = get_referral_settings()
                try:
                    await context.bot.send_message(
                        chat_id=referrer_id,
                        text=f"🎉 <b>New Referral!</b>\n\nSomeone joined using your link!\n"
                             f"You earned <b>{settings['reward']} {settings['label']}</b>! 🎁",
                        parse_mode='HTML'
                    )
                except Exception:
                    pass
        except (ValueError, IndexError):
            pass

    # If admin was in user panel mode, clear it and return to admin panel
    if is_admin(username, user_id):
        context.user_data['user_panel_mode'] = False
    
    # Welcome message without channel join requirement
    welcome_message = """
🤖 *Welcome to Number Bot!*

Stay with us, I hope you can learn something good. Join the live regularly. Join all my channels and groups.

🧑‍💻 *Bot Owner:* ADMIN LIMON
"""
    
    if is_admin(username, user_id):
        await update.message.reply_text(
            welcome_message,
            parse_mode='Markdown',
            reply_markup=get_admin_keyboard()
        )
    else:
        if await enforce_join(update, context):
            await update.message.reply_text(
                welcome_message,
                parse_mode='Markdown'
            )
            await show_main_menu(update, context)

# Error handler
async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.error(f"Exception while handling an update: {context.error}")

# Cache for resolved chat_ids (link → chat_id)
_CHAT_ID_CACHE: dict = {}

async def _resolve_chat_id(bot, link: str):
    """Resolve a Telegram channel link to a numeric chat_id.
    Supports public (@username) and invite (+hash) links.
    Returns chat_id int or None on failure."""
    if not link:
        return None
    if link in _CHAT_ID_CACHE:
        return _CHAT_ID_CACHE[link]
    try:
        # Normalize: https://t.me/username → @username
        # Normalize: https://t.me/+hash  → keep as invite link
        identifier = link.strip()
        if 't.me/' in identifier:
            path = identifier.split('t.me/')[-1].lstrip('/')
            if path.startswith('+'):
                # Invite link — use the full link
                identifier = link
            else:
                identifier = '@' + path
        chat = await bot.get_chat(identifier)
        chat_id = chat.id
        _CHAT_ID_CACHE[link] = chat_id
        return chat_id
    except Exception:
        return None

async def auto_check_members_job(context: ContextTypes.DEFAULT_TYPE):
    """Periodic job: every minute checks if enough time has elapsed, then flags
    all users who are no longer in required channels. Resets notified set each cycle."""
    interval_minutes = await asyncio.to_thread(get_check_interval)
    last_check = context.bot_data.get('last_member_check', 0)
    now = time.monotonic()
    if now - last_check < interval_minutes * 60:
        return  # not time yet

    context.bot_data['last_member_check'] = now
    channels = await asyncio.to_thread(get_join_channels)
    if not channels:
        return

    all_users = await asyncio.to_thread(get_all_users)
    flagged = set()
    for user_data in all_users:
        uid = user_data.get('user_id') if isinstance(user_data, dict) else user_data
        if not uid:
            continue
        for _, title, link in channels:
            chat_id = await _resolve_chat_id(context.bot, link)
            if chat_id is None:
                continue
            try:
                member = await context.bot.get_chat_member(chat_id, uid)
                if member.status in ('left', 'kicked', 'banned'):
                    flagged.add(uid)
                    break
            except Exception:
                pass

    context.bot_data['join_flagged'] = flagged
    context.bot_data['join_notified'] = set()  # reset per cycle so each cycle notifies once
    context.bot_data['join_prompted'] = set()  # reset so enforce_join re-notifies once per cycle
    _MEMBER_CACHE.clear()  # invalidate membership cache after bulk check

    # KEY FIX: remove flagged users from the session-level verified bypass.
    # Without this, users who left a channel after being verified once would
    # pass enforce_join() forever (regardless of check_interval), because
    # _ALREADY_VERIFIED is never expired on its own.
    for uid in flagged:
        _ALREADY_VERIFIED.discard(uid)


def main():
    try:
        import concurrent.futures

        # ── Larger thread pool so asyncio.to_thread() DB calls don't queue ──────
        executor = concurrent.futures.ThreadPoolExecutor(max_workers=64)
        loop = asyncio.new_event_loop()
        loop.set_default_executor(executor)
        asyncio.set_event_loop(loop)

        # ── Build Application with performance settings ──────────────────────────
        application = (
            Application.builder()
            .token(BOT_TOKEN)
            .concurrent_updates(True)          # process many updates simultaneously
            .connection_pool_size(32)          # more HTTP connections to Telegram
            .read_timeout(30)
            .write_timeout(30)
            .connect_timeout(15)
            .pool_timeout(10)
            .build()
        )
        
        # Add handlers
        application.add_handler(CommandHandler("start", start_command))
        application.add_handler(CommandHandler("userpanel", userpanel_command))
        application.add_handler(CommandHandler("cancel", cancel_command))
        
        # Button click handlers
        application.add_handler(MessageHandler(filters.Text([
            "➕ Add Numbers", "🌍 Add Country", "📬 Message Hub",
            "🚩 Country Manager", "👑 Admin Manager", "🔙 Back to Admin",
            "➕ Add Admin", "🔄 Reset Number",
            "💰 OTP Rewards",
            "📱 Add Service", "🗺 Service Map",
            "📤 Group Button Links",
            "⚙️ Settings", "🎁 Referral Settings", "🔢 Numbers Per Request",
            "🔗 Link Settings", "📝 Custom Message", "✏️ Set Message", "🗑 Remove Message",
            "📲 Set OTP Link", "❌ Remove OTP Link",
            "➕ Add Channel", "➖ Remove Channel", "⏱ Check Interval",
            "🔛 Toggle Referral", "🎁 Set Reward",
            "👤 Check Balance", "➕ Add Balance", "➖ Remove Balance",
            "🔙 Back to Settings",
            "💳 Withdraw Settings", "🔛 Toggle Withdraw", "💰 Set Min Amount",
            "📨 Set Group ID", "📊 Withdraw Stats",
            "🤖 Bot Status", "👥 User Activity",
            "📡 SMS Panels", "➕ Add SMS Panel", "📤 Group OTP Forward", "📊 Panel Statistics",
            "🔴 বন্ধ করুন", "🟢 চালু করুন", "📨 Last Message",
            "📤 Group এ পাঠান",
            "✏️ Username পরিবর্তন", "🔑 Password পরিবর্তন",
            "🗑️ Delete Panel", "🔙 Back to Panels",
            "➕ Group Add", "🗑 Group Remove",
        ]), handle_button_click))
        application.add_handler(MessageHandler(filters.Text(["✅ Get Numbers", "🔢 Get Numbers", "Get Numbers", "💰 Balance", "💎 Balance", "💰 My Balance", "🌍 Available country", "🌐 Available country", "🌍 Available Country", "📋 Notice", "🔔 Notice", "📢 Notice"]), handle_user_button_click))
        
        application.add_handler(CallbackQueryHandler(handle_callback))
        # Document handler MUST come before broadcast_media_filter so that
        # files sent during "Add Numbers" flow are not swallowed by the broadcast handler
        application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
        # Broadcast media handler — catches photos/video/audio/etc when broadcast is active
        broadcast_media_filter = (
            filters.PHOTO | filters.VIDEO | filters.AUDIO | filters.VOICE |
            filters.ANIMATION | filters.Sticker.ALL | filters.VIDEO_NOTE |
            filters.FORWARDED
        )
        application.add_handler(MessageHandler(broadcast_media_filter, handle_broadcast_media))
        application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
        
        # Error handler
        application.add_error_handler(error_handler)
        
        # Auto member-check job: runs every 60s, internally throttled by check_interval_minutes
        application.job_queue.run_repeating(auto_check_members_job, interval=60, first=30)

        # Start SMS panel monitors after bot is ready
        async def _post_init(app):
            try:
                await start_all_panels(app.bot)
            except Exception as e:
                logger.error(f"Panel startup error: {e}")

        application.post_init = _post_init

        # Start the bot
        print("🤖 Bot is running...")
        application.run_polling(
            drop_pending_updates=True,
        )
        
    except Exception as e:
        logger.error(f"Failed to start bot: {e}")

if __name__ == '__main__':
    main()